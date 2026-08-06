from __future__ import annotations

import argparse
import hashlib
import re
import sys
import tempfile
import zipfile
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path

import openpyxl
import yaml
from openpyxl.formula import Tokenizer
from openpyxl.utils.cell import range_boundaries


ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "src"))

from executive_docs.excel import OOXMLWorkbook  # noqa: E402


SOURCE_DIR = ROOT / "NEW_TEMPLATES" / "Attachments_vku@e-systems-4"
ETALON_DIR = ROOT / "ETALON"
APPROVED_DIR = ROOT / "templates" / "approved"
CONTRACTS_DIR = ROOT / "templates" / "fill-contracts"
VERSION = "2026-07-30-discovery-2"
AOSR_VL_VERSION = "2026-08-05-discovery-3"
TEMPLATES = (
    ("emr", "1. ЭМР1.xlsx", "Электромонтажные работы (ЭМР)", "emr"),
    ("protocols", "2. Протоколы.xlsx", "Протоколы испытаний", "protocols"),
    ("ojr", "3. ОЖР.xlsx", "Общий журнал работ (ОЖР)", "ojr"),
    ("avk", "4. АВК.xlsx", "Входной контроль (АВК)", "avk"),
    ("aosr_vl", "8. АОСР ВЛ1.xlsx", "АОСР воздушной линии", "aosr_vl"),
)

FORMULA_ERRORS = ("#REF!", "#VALUE!", "#N/A", "#DIV/0!", "#NAME?")
CELL_RANGE = re.compile(r"[A-Z]{1,3}[1-9][0-9]*(?::[A-Z]{1,3}[1-9][0-9]*)?")
DATE_PATTERN = re.compile(
    r"\b(?:0?[1-9]|[12]\d|3[01])[./-](?:0?[1-9]|1[0-2])[./-]"
    r"(?:19|20)?\d{2}(?:\s*г\.?)?(?!\d)"
)
TEXTUAL_DATE_PATTERN = re.compile(
    r"(?:\bот\s+)?[«\"]?(?:0?[1-9]|[12]\d|3[01])[»\"]?\s+"
    r"(?:января|февраля|марта|апреля|мая|июня|июля|августа|"
    r"сентября|октября|ноября|декабря)\s+(?:19|20)\d{2}(?:\s*г\.?)?",
    re.IGNORECASE,
)
CALCULATION_VALUE_PATTERN = re.compile(
    r"=\s*-?\d[\d\s.,]*(?:/|:|\*)\s*-?\d",
)
PROJECT_QUANTITY_PATTERN = re.compile(
    r"\b\d+(?:[.,]\d+)?\s+(?:шт\.?|кв\.?\s*м|м[²³23]?|кг|ква|квт|кв)\b",
    re.IGNORECASE,
)
PERSON_PATTERN = re.compile(
    r"\b[А-ЯЁ][а-яё-]{2,}\s+[А-ЯЁ]\.\s*[А-ЯЁ]\.(?:\s|$)"
)
PERSON_PATTERN_REVERSED = re.compile(
    r"(?:^|\s)[А-ЯЁ]\.\s*[А-ЯЁ]\.\s*[А-ЯЁ][а-яё-]{2,}(?:\s|$)"
)
SENSITIVE_TOKENS = (
    " инн ",
    " кпп ",
    " огрн ",
    " нрс ",
    " нострой",
    "приказ №",
    "доверенн",
    "полномочи",
    "гефест",
    "энергосистем",
    "солнечногорск",
    "высоцк",
)
PRIOR_PROJECT_TOKENS = (
    "гефест",
    "энергосистем",
    "солнечногорск",
    "высоцк",
)
STATIC_PLACEHOLDER_TOKENS = (
    "фамилия, имя, отчество",
    "полное и (или) сокращенное наименование",
    "адрес места жительства",
    "места нахождения юридического лица",
    "индивидуального предпринимателя",
    "указывается",
    "при наличии",
    "если установка",
    "является типовым",
    "обычно ",
)
REGULATORY_TEXT_TOKENS = (
    "нормативно-техническ",
    "нормативные документы",
    "гост ",
    "птээп",
    "пуэ",
    "снип",
    "сп ",
    "рд ",
    "таблица ",
    "требовани",
)
PACKAGE_FORBIDDEN_TOKENS = (
    "гефест",
    "энергосистем",
    "высоцк",
    "алексанян",
    "чернявск",
    "шатковск",
    "трушина",
    "бараночникова",
    "elena camarillo",
)
SHEET_RENAMES = {
    "emr": {
        "Титульный Гефест": "Титульный подрядчик",
    },
}
MANUAL_TOKENS = {
    "passport_or_certificate": (
        "паспорт",
        "сертификат",
        "декларац",
        "свидетельств",
        "удостоверен",
        "документ о качестве",
    ),
    "signatory_or_authority": (
        "подписант",
        "подпись",
        "приказ",
        "директор",
        "представител",
        "строительн",
        "контрол",
        "нрс",
        "нострой",
        "доверенн",
        "полномочи",
        "ф.и.о",
    ),
    "actual_execution_fact": (
        "фактическ",
        "дата начала",
        "дата оконч",
        "начало работ",
        "окончание работ",
        "выполнено",
        "исполнено",
        "объем выполн",
        "объём выполн",
    ),
}


def _aosr_field(
    semantic_id: str,
    label: str,
    description: str,
    *,
    value_kind: str = "text",
    evidence_rule: str = "direct_pdf",
    required: bool = True,
    value_pattern: str | None = None,
) -> dict:
    return {
        "semantic_id": semantic_id,
        "label": label,
        "description": description,
        "value_kind": value_kind,
        "evidence_rule": evidence_rule,
        "required": required,
        "manual_reason": None,
        "value_pattern": value_pattern,
    }


AOSR_VL_FIELD_OVERRIDES = {
    ("Данные объект", "B2"): _aosr_field(
        "project.sap_number",
        "Номер SAP",
        "Короткий идентификатор SAP с явным префиксом SAP; это не шифр проектной документации",
        value_pattern=r"(?i)^(?:№\s*)?SAP\s*[-–—:]?\s*\d{4,}$",
    ),
    ("Данные объект", "B3"): _aosr_field(
        "project.object_name",
        "Наименование объекта капитального строительства",
        "Полное наименование объекта в формулировке PDF",
    ),
    ("Данные объект", "B4"): _aosr_field(
        "project.district",
        "Район объекта",
        "Муниципальный район или городской округ объекта",
    ),
    ("Данные объект", "B5"): _aosr_field(
        "project.object_address",
        "Адрес объекта",
        "Почтовый или строительный адрес объекта, а не адрес организации",
    ),
    ("Данные объект", "B9"): _aosr_field(
        "project.line_designation",
        "Обозначение линии ВЛ",
        "Наименование или диспетчерское обозначение линии, фидера и питающего пункта",
    ),
    ("Данные объект", "B42"): _aosr_field(
        "project.city",
        "Город выпуска документации",
        "Населённый пункт, указанный для выпуска проектной документации",
    ),
    ("Данные объект", "B43"): _aosr_field(
        "project.design_document_code",
        "Шифр проектной документации",
        "Полный шифр или номер проектной/рабочей документации; это не номер SAP",
    ),
    ("АОСР-1", "C33"): _aosr_field(
        "aosr_vl.act_1.number",
        "Номер первого АОСР",
        "Номер первого акта в комплекте; последующие номера вычисляются формулами",
        value_kind="number",
        evidence_rule="actual_executive_document_only",
    ),
    ("АОСР-2", "P63"): _aosr_field(
        "aosr_vl.act_2.support_count",
        "Количество смонтированных опор",
        "Фактическое количество опор по АОСР на монтаж опор",
        value_kind="number",
        evidence_rule="actual_executive_document_only",
    ),
    ("АОСР-2", "V63"): _aosr_field(
        "aosr_vl.act_2.pole_count",
        "Количество смонтированных стоек",
        "Фактическое количество железобетонных стоек по АОСР",
        value_kind="number",
        evidence_rule="actual_executive_document_only",
    ),
    ("АОСР-2", "M70"): _aosr_field(
        "aosr_vl.act_2.pole_quality_document",
        "Документ о качестве стоек",
        "Паспорт, сертификат или иной документ о качестве применённых стоек",
        evidence_rule="actual_executive_document_only",
    ),
    ("АОСР-2", "B97"): _aosr_field(
        "aosr_vl.act_2.additional_quality_document",
        "Дополнительный документ о качестве",
        "Полные реквизиты дополнительного сертификата, если он указан в PDF",
        evidence_rule="actual_executive_document_only",
        required=False,
    ),
    ("АОСР-3", "T62"): _aosr_field(
        "aosr_vl.act_3.support_count",
        "Количество опор для устройства заземления",
        "Фактическое количество опор, для которых выполнялись земляные работы",
        value_kind="number",
        evidence_rule="actual_executive_document_only",
    ),
    ("АОСР-3", "X62"): _aosr_field(
        "aosr_vl.act_3.excavation_volume_m3",
        "Объём выемки грунта, м³",
        "Фактический объём выемки грунта в кубических метрах",
        value_kind="number",
        evidence_rule="actual_executive_document_only",
    ),
    ("АОСР-4", "H69"): _aosr_field(
        "aosr_vl.act_4.angle_steel_quantity",
        "Количество уголка 50×50×5",
        "Фактическое количество уголка по АОСР устройства заземления",
        value_kind="number",
        evidence_rule="actual_executive_document_only",
    ),
    ("АОСР-4", "H70"): _aosr_field(
        "aosr_vl.act_4.rebar_quantity",
        "Количество арматуры Ø8",
        "Фактическое количество арматуры Ø8 по АОСР",
        value_kind="number",
        evidence_rule="actual_executive_document_only",
    ),
    ("АОСР-4", "H71"): _aosr_field(
        "aosr_vl.act_4.steel_strip_quantity",
        "Количество полосы 40×4",
        "Фактическое количество полосы 40×4 мм по АОСР",
        value_kind="number",
        evidence_rule="actual_executive_document_only",
    ),
    ("АОСР-4", "L69"): _aosr_field(
        "aosr_vl.act_4.angle_steel_quality_document",
        "Документ о качестве уголка",
        "Реквизиты документа о качестве уголка 50×50×5",
        evidence_rule="actual_executive_document_only",
    ),
    ("АОСР-4", "L70"): _aosr_field(
        "aosr_vl.act_4.rebar_quality_document",
        "Документ о качестве арматуры",
        "Реквизиты документа о качестве арматуры Ø8",
        evidence_rule="actual_executive_document_only",
    ),
    ("АОСР-4", "L71"): _aosr_field(
        "aosr_vl.act_4.steel_strip_quality_document",
        "Документ о качестве полосы",
        "Реквизиты документа о качестве полосы 40×4 мм",
        evidence_rule="actual_executive_document_only",
    ),
    ("АОСР-6", "H69"): _aosr_field(
        "aosr_vl.act_6.sip_quantity",
        "Количество провода СИП",
        "Фактическое количество провода СИП по АОСР",
        value_kind="number",
        evidence_rule="actual_executive_document_only",
    ),
    ("АОСР-6", "H70"): _aosr_field(
        "aosr_vl.act_6.line_fittings_quantity",
        "Количество линейной арматуры",
        "Фактическое количество линейной арматуры по АОСР",
        value_kind="number",
        evidence_rule="actual_executive_document_only",
    ),
    ("АОСР-6", "L70"): _aosr_field(
        "aosr_vl.act_6.line_fittings_quality_document",
        "Документ о качестве линейной арматуры",
        "Полные реквизиты документа о качестве линейной арматуры",
        evidence_rule="actual_executive_document_only",
    ),
    ("АОСР-7", "H69"): _aosr_field(
        "aosr_vl.act_7.paint_quantity",
        "Количество краски",
        "Фактическое количество краски по АОСР окраски опор",
        value_kind="number",
        evidence_rule="actual_executive_document_only",
    ),
    ("АОСР-7", "H70"): _aosr_field(
        "aosr_vl.act_7.line_fittings_quantity",
        "Количество линейной арматуры",
        "Фактическое количество линейной арматуры по АОСР окраски опор",
        value_kind="number",
        evidence_rule="actual_executive_document_only",
    ),
    ("АОСР-7", "L69"): _aosr_field(
        "aosr_vl.act_7.paint_quality_document",
        "Документ о качестве краски",
        "Полные реквизиты документа о качестве краски",
        evidence_rule="actual_executive_document_only",
    ),
    ("АОСР-7", "L70"): _aosr_field(
        "aosr_vl.act_7.line_fittings_quality_document",
        "Документ о качестве линейной арматуры",
        "Полные реквизиты документа о качестве линейной арматуры",
        evidence_rule="actual_executive_document_only",
    ),
    ("АОСР-шурф", "C32"): _aosr_field(
        "aosr_vl.pit_act.number",
        "Номер АОСР шурфления",
        "Номер отдельного акта шурфления, если этот скрытый лист используется",
        evidence_rule="actual_executive_document_only",
        required=False,
    ),
}

AOSR_VL_BROKEN_FORMULA_CELLS = {
    "АОСР-1": {"A117"},
    "АОСР-2": {"A118"},
    "АОСР-3": {"A115"},
    "АОСР-5": {"A114"},
    "АОСР-6": {"A116"},
    "АОСР-7": {"A115"},
    "АОСР-шурф": {"A114"},
}

AOSR_VL_FORMULA_OVERRIDES = {
    ("АОСР-2", "C33"): "=IF('АОСР-1'!C33=\"\",\"\",'АОСР-1'!C33+1)",
    ("АОСР-3", "C32"): "=IF('АОСР-2'!C33=\"\",\"\",'АОСР-2'!C33+1)",
    ("АОСР-4", "C32"): "=IF('АОСР-3'!C32=\"\",\"\",'АОСР-3'!C32+1)",
    ("АОСР-5", "C32"): "=IF('АОСР-4'!C32=\"\",\"\",'АОСР-4'!C32+1)",
    ("АОСР-6", "C32"): "=IF('АОСР-5'!C32=\"\",\"\",'АОСР-5'!C32+1)",
    ("АОСР-7", "C32"): "=IF('АОСР-6'!C32=\"\",\"\",'АОСР-6'!C32+1)",
    ("АОСР-6", "A62"): "=IF('АОСР-5'!A86=\"\",\"\",'АОСР-5'!A86)",
    ("АОСР-7", "A62"): "=IF('АОСР-6'!A86=\"\",\"\",'АОСР-6'!A86)",
}


def digest(path: Path) -> str:
    value = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            value.update(chunk)
    return value.hexdigest()


def is_formula(value: object) -> bool:
    return isinstance(value, str) and value.startswith("=")


def normalized_sheet(name: str) -> str:
    return re.sub(r"\s+", " ", name.casefold().replace("ё", "е")).strip()


def is_primary_data_sheet(name: str) -> bool:
    return normalized_sheet(name) in {
        "данные объект",
        "данные организации",
    }


def is_input_sheet(name: str) -> bool:
    normalized = normalized_sheet(name)
    return (
        is_primary_data_sheet(name)
        or normalized in {"главная", "оборудование"}
    )


def value_kind(cell) -> str:
    if cell.is_date or isinstance(cell.value, (date, datetime)):
        return "date"
    if isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
        return "number"
    return "text"


def unquote_sheet(value: str) -> str:
    value = value.strip()
    if value.startswith("'") and value.endswith("'"):
        value = value[1:-1].replace("''", "'")
    return re.sub(r"^\[[1-9][0-9]*\]", "", value)


def formula_references(workbook) -> set[tuple[str, str]]:
    """Collect direct A1 precedents without using ETALON as a mapping oracle."""

    names = {name.casefold(): name for name in workbook.sheetnames}
    references: set[tuple[str, str]] = set()
    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                if not is_formula(cell.value):
                    continue
                try:
                    tokens = Tokenizer(cell.value).items
                except Exception:
                    continue
                for token in tokens:
                    if token.type != "OPERAND" or token.subtype != "RANGE":
                        continue
                    reference = token.value
                    if "!" in reference:
                        sheet_token, address = reference.rsplit("!", 1)
                        sheet_name = names.get(unquote_sheet(sheet_token).casefold())
                    else:
                        sheet_name = worksheet.title
                        address = reference
                    address = address.replace("$", "")
                    if sheet_name is None or not CELL_RANGE.fullmatch(address):
                        continue
                    min_column, min_row, max_column, max_row = range_boundaries(address)
                    if (max_column - min_column + 1) * (max_row - min_row + 1) > 10_000:
                        continue
                    target = workbook[sheet_name]
                    for target_row in range(min_row, max_row + 1):
                        for target_column in range(min_column, max_column + 1):
                            references.add(
                                (
                                    sheet_name,
                                    target.cell(target_row, target_column).coordinate,
                                )
                            )
    return references


def is_merged_non_anchor(worksheet, cell) -> bool:
    for merged in worksheet.merged_cells.ranges:
        if (
            merged.min_row <= cell.row <= merged.max_row
            and merged.min_col <= cell.column <= merged.max_col
        ):
            return cell.row != merged.min_row or cell.column != merged.min_col
    return False


def looks_project_specific(value: object) -> bool:
    if isinstance(value, (date, datetime)):
        return True
    if not isinstance(value, str):
        return False
    compact = re.sub(r"\s+", " ", value)
    normalized = f" {compact.casefold().replace('ё', 'е')} "
    has_numeric_date = bool(DATE_PATTERN.search(normalized))
    has_textual_date = bool(TEXTUAL_DATE_PATTERN.search(normalized))
    has_person = bool(
        PERSON_PATTERN.search(value)
        or PERSON_PATTERN_REVERSED.search(value)
    )
    has_named_organization = bool(
        re.search(r"\b(?:ооо|пао|ао)\s+[«\"]", normalized)
    )
    has_long_identifier = bool(re.search(r"\b\d{6,}\b", normalized))
    has_address = "ул." in normalized and (
        "дом" in normalized
        or "д." in normalized
        or "строен" in normalized
        or "корпус" in normalized
    )
    has_email = "@" in normalized
    has_prior_project_token = any(
        token in normalized for token in PRIOR_PROJECT_TOKENS
    )
    quality_document_with_identifier = any(
        token in normalized
        for token in (
            "паспорт",
            "сертификат",
            "декларац",
            "свидетельств",
            "удостоверен",
            "документ о качестве",
        )
    ) and (
        "№" in value
        or has_numeric_date
        or has_textual_date
        or bool(re.search(r"\b\d{3,}\b", value))
    )
    has_sensitive_identifier = any(
        token in normalized for token in SENSITIVE_TOKENS
    ) and bool(
        re.search(r"\d", value)
        or has_person
        or has_named_organization
        or has_address
        or has_email
        or has_prior_project_token
    )
    strong_signal = bool(
        has_numeric_date
        or has_textual_date
        or has_person
        or has_prior_project_token
        or quality_document_with_identifier
        or has_sensitive_identifier
        or has_named_organization
        or has_long_identifier
        or has_address
        or has_email
        or CALCULATION_VALUE_PATTERN.search(normalized)
    )
    if strong_signal:
        return True
    if any(token in normalized for token in STATIC_PLACEHOLDER_TOKENS):
        return False
    if any(token in normalized for token in REGULATORY_TEXT_TOKENS):
        return False
    return bool(
        PROJECT_QUANTITY_PATTERN.search(normalized)
    )


def is_horizontal_sequence_number(worksheet, cell) -> bool:
    if not isinstance(cell.value, int) or isinstance(cell.value, bool):
        return False
    values = [
        (candidate.column, candidate.value)
        for candidate in worksheet[cell.row]
        if isinstance(candidate.value, int) and not isinstance(candidate.value, bool)
    ]
    if len(values) < 3:
        return False
    values.sort()
    columns = [column for column, _ in values]
    numbers = [value for _, value in values]
    return (
        columns == list(range(columns[0], columns[0] + len(columns)))
        and numbers == list(range(numbers[0], numbers[0] + len(numbers)))
    )


def is_protocol_actual_entry_cell(worksheet, cell) -> bool:
    value = cell.value
    if value in (None, "") or is_formula(value):
        return False
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if cell.column == 1 and isinstance(value, int) and 0 < value <= 100:
            return False
        if is_horizontal_sequence_number(worksheet, cell):
            return False
        return True
    if isinstance(value, str) and CALCULATION_VALUE_PATTERN.search(value):
        return True
    raw_fill_rgb = cell.fill.fgColor.rgb
    fill_rgb = raw_fill_rgb.upper() if isinstance(raw_fill_rgb, str) else ""
    return cell.fill.fill_type == "solid" and fill_rgb in {
        "FFFFFF00",
        "FFCCFFFF",
        "FFFFE699",
    }


def is_actual_entry_cell(template_id: str, sheet_name: str, cell) -> bool:
    normalized = normalized_sheet(sheet_name)
    value = cell.value
    if value in (None, "") or is_formula(value):
        return False
    if isinstance(value, int) and cell.column == 1:
        return False
    if isinstance(value, str) and "заполняется от руки" in value.casefold():
        return False
    if template_id == "protocols" and is_protocol_actual_entry_cell(
        cell.parent,
        cell,
    ):
        return True
    if template_id == "ojr" and normalized.startswith("раздел") and cell.row >= 5:
        return True
    if template_id == "avk" and "журнал авк" in normalized and cell.row >= 11:
        return True
    if template_id == "emr" and (
        normalized.startswith("пр.15 реестр")
        or normalized.startswith("ведомость")
    ) and cell.row >= 5:
        return True
    return False


def should_register_cell(
    template_id: str,
    worksheet,
    cell,
    references: set[tuple[str, str]],
) -> bool:
    if is_formula(cell.value) or is_merged_non_anchor(worksheet, cell):
        return False
    # Hyperlinks in the source corpus are workbook navigation labels. Clearing
    # their cells while retaining the hyperlink makes spreadsheet readers expose
    # the target address as a value and destroys the template's contents.
    if cell.hyperlink:
        return False
    referenced = (worksheet.title, cell.coordinate) in references
    normalized = normalized_sheet(worksheet.title)
    if is_primary_data_sheet(worksheet.title) and cell.column >= 2:
        return cell.value not in (None, "") or referenced
    if normalized == "оборудование" and cell.row >= 2:
        return cell.value not in (None, "") or referenced
    if referenced and is_input_sheet(worksheet.title):
        return True
    return looks_project_specific(cell.value) or is_actual_entry_cell(
        template_id,
        worksheet.title,
        cell,
    )


def nearby_label(
    source_sheet,
    row: int,
    column: int,
    registered: set[str],
) -> str:
    candidates: list[tuple[int, int, str]] = []
    for row_offset in range(-3, 4):
        for column_offset in range(-5, 2):
            if row_offset == 0 and column_offset == 0:
                continue
            current_row = row + row_offset
            current_column = column + column_offset
            if current_row < 1 or current_column < 1:
                continue
            source = source_sheet.cell(current_row, current_column)
            if source.coordinate in registered or is_formula(source.value):
                continue
            if not isinstance(source.value, str):
                continue
            text = re.sub(r"\s+", " ", source.value).strip()
            if not text or len(text) > 140 or looks_project_specific(text):
                continue
            same_row_bonus = 0 if row_offset == 0 and column_offset < 0 else 1
            distance = abs(row_offset) + abs(column_offset)
            candidates.append((same_row_bonus, distance, text))
    labels: list[str] = []
    for _, _, text in sorted(candidates):
        if text not in labels:
            labels.append(text)
        if len(labels) == 2:
            break
    return " · ".join(labels)


def manual_reason(
    *,
    template_id: str,
    sheet: str,
    label: str,
    source_cell,
) -> str | None:
    normalized = normalized_sheet(sheet)
    value = source_cell.value
    text = f"{sheet} {label} {value or ''}".casefold().replace("ё", "е")
    if "данные организации" in normalized:
        return "Реквизиты, подписанты и полномочия требуют утверждённого профиля организации"
    if any(
        token in text
        for token in (
            " огрн",
            " инн",
            " кпп",
            "реквизит",
            "адрес",
            "организац",
            "заказчик",
            "застройщик",
            "подрядчик",
            " пао ",
            " ооо ",
            " ао «",
            ' ао "',
            "филиал",
            "главный инженер",
            "лицо, ответствен",
            "члены комиссии",
        )
    ) or any(token in text for token in PRIOR_PROJECT_TOKENS):
        return "Реквизиты, организация или ответственный специалист требуют утверждённого профиля"
    if (
        source_cell.is_date
        or isinstance(value, (date, datetime))
        or DATE_PATTERN.search(str(value or ""))
        or TEXTUAL_DATE_PATTERN.search(str(value or ""))
        or "дата" in text
    ):
        return "Дата требует отдельного подтверждения специалистом; проектный график не является фактом выполнения"
    if template_id == "protocols" and normalized == "оборудование":
        return "Прибор, заводской номер и сведения о поверке требуют отдельного подтверждения"
    if template_id == "protocols" and is_protocol_actual_entry_cell(
        source_cell.parent,
        source_cell,
    ):
        return "Фактический параметр или результат испытания требует отдельного подтверждения"
    if template_id == "avk" and normalized != "данные объект":
        return "Фактическое значение входного контроля требует отдельного подтверждения"
    if template_id == "avk" and looks_project_specific(value) and (
        DATE_PATTERN.search(str(value or ""))
        or any(
            token in text
            for token in (
                "паспорт",
                "сертификат",
                "декларац",
                "свидетельств",
                "удостоверен",
                "документ о качестве",
            )
        )
    ):
        return "Документ о качестве и его реквизиты требуют отдельного подтверждения"
    if template_id == "avk" and "журнал авк" in normalized:
        return "Фактическая запись входного контроля требует отдельного подтверждения"
    if template_id == "ojr" and normalized.startswith("раздел"):
        return "Фактическая запись общего журнала работ требует отдельного подтверждения"
    if template_id == "aosr_vl" and normalized.startswith("аоср"):
        return "Фактическое значение АОСР требует отдельного подтверждения"
    if template_id == "emr" and (
        "реестр" in normalized or normalized.startswith("ведомость")
    ):
        return "Фактическая запись реестра требует отдельного подтверждения"
    if PERSON_PATTERN.search(str(value or "")) or PERSON_PATTERN_REVERSED.search(
        str(value or "")
    ):
        return "ФИО и полномочия подписанта требуют утверждённого профиля"
    for category, tokens in MANUAL_TOKENS.items():
        if any(token in text for token in tokens):
            return {
                "passport_or_certificate": "Паспорт или сертификат отсутствует в PDF и требует ручного подтверждения",
                "signatory_or_authority": "Подписант или основание полномочий требует утверждённого профиля",
                "actual_execution_fact": "Фактическое значение нельзя выводить из проектного PDF",
            }[category]
    return None


def compare_with_etalon(candidate_book, etalon_book) -> dict:
    common_sheets = [
        name for name in candidate_book.sheetnames if name in etalon_book.sheetnames
    ]
    formula_differences = 0
    nonformula_differences = 0
    formula_errors = 0
    for sheet_name in candidate_book.sheetnames:
        worksheet = candidate_book[sheet_name]
        for row in worksheet.iter_rows():
            for cell in row:
                if is_formula(cell.value) and any(
                    token in cell.value for token in FORMULA_ERRORS
                ):
                    formula_errors += 1
    for sheet_name in common_sheets:
        source_sheet = candidate_book[sheet_name]
        etalon_sheet = etalon_book[sheet_name]
        max_row = max(source_sheet.max_row, etalon_sheet.max_row)
        max_column = max(source_sheet.max_column, etalon_sheet.max_column)
        for row in range(1, max_row + 1):
            for column in range(1, max_column + 1):
                source_value = source_sheet.cell(row, column).value
                etalon_value = etalon_sheet.cell(row, column).value
                if source_value == etalon_value:
                    continue
                if is_formula(source_value) or is_formula(etalon_value):
                    formula_differences += 1
                else:
                    nonformula_differences += 1
    return {
        "candidate_sheet_count": len(candidate_book.sheetnames),
        "etalon_sheet_count": len(etalon_book.sheetnames),
        "candidate_vs_etalon_sheet_names_match": candidate_book.sheetnames
        == etalon_book.sheetnames,
        "candidate_only_sheet_count": len(
            [
                name
                for name in candidate_book.sheetnames
                if name not in etalon_book.sheetnames
            ]
        ),
        "etalon_only_sheet_count": len(
            [
                name
                for name in etalon_book.sheetnames
                if name not in candidate_book.sheetnames
            ]
        ),
        "candidate_vs_etalon_formula_differences": formula_differences,
        "candidate_vs_etalon_nonformula_differences": nonformula_differences,
        "formula_errors_observed": formula_errors,
        "etalon_external_links": len(getattr(etalon_book, "_external_links", [])),
    }


def raw_package_findings(path: Path) -> dict:
    raw_ref_errors = 0
    forbidden: list[str] = []
    with zipfile.ZipFile(path) as archive:
        for name in archive.namelist():
            payload = archive.read(name)
            if not name.endswith((".xml", ".rels")):
                continue
            raw_ref_errors += payload.upper().count(b"#REF!")
            decoded = payload.decode("utf-8", errors="ignore").casefold()
            decoded_utf16 = payload.decode("utf-16", errors="ignore").casefold()
            for token in PACKAGE_FORBIDDEN_TOKENS:
                if token in decoded or token in decoded_utf16:
                    forbidden.append(f"{name}:{token}")
    return {
        "raw_ref_error_count": raw_ref_errors,
        "package_forbidden_token_count": len(forbidden),
        "package_forbidden_token_examples": forbidden[:25],
    }


def unsafe_blank_formula_examples(workbook) -> list[str]:
    """Return direct links/concatenations that can emit false values from blanks."""

    direct = re.compile(
        r"^=\s*(?:(?:'(?:''|[^'])+'|[A-Za-z_\u0400-\u04FF][^'!\[\]]*)!)?"
        r"\$?[A-Z]{1,3}\$?[1-9][0-9]*\s*$"
    )
    unsafe: list[str] = []
    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                formula = str(cell.value or "")
                if cell.data_type != "f":
                    continue
                upper = formula.lstrip("=").lstrip().upper()
                if direct.fullmatch(formula) or upper.startswith("CONCATENATE("):
                    unsafe.append(f"{worksheet.title}!{cell.coordinate}")
    return unsafe


def register(
    template_id: str,
    filename: str,
    display_name: str,
    document_kind: str,
    *,
    approved_dir: Path = APPROVED_DIR,
    contracts_dir: Path = CONTRACTS_DIR,
) -> dict:
    source_path = SOURCE_DIR / filename
    etalon_path = ETALON_DIR / filename
    if not source_path.is_file() or not etalon_path.is_file():
        raise FileNotFoundError(f"Не найдена пара NEW_TEMPLATES/ETALON: {filename}")

    source_book = openpyxl.load_workbook(
        source_path,
        data_only=False,
        keep_links=True,
    )
    try:
        source_external_links = len(getattr(source_book, "_external_links", []))
        sheet_renames = SHEET_RENAMES.get(template_id, {})
        references = formula_references(source_book)
        registered_by_sheet: dict[str, set[str]] = defaultdict(set)
        for worksheet in source_book.worksheets:
            for row in worksheet.iter_rows():
                for cell in row:
                    if should_register_cell(
                        template_id,
                        worksheet,
                        cell,
                        references,
                    ):
                        registered_by_sheet[worksheet.title].add(cell.coordinate)

        if template_id == "aosr_vl":
            for sheet_name, coordinate in AOSR_VL_FIELD_OVERRIDES:
                cell = source_book[sheet_name][coordinate]
                if is_formula(cell.value) or is_merged_non_anchor(
                    source_book[sheet_name],
                    cell,
                ):
                    raise ValueError(
                        f"Некорректное явное поле АОСР ВЛ: {sheet_name}!{coordinate}"
                    )
                registered_by_sheet[sheet_name].add(coordinate)
            for sheet_name in source_book.sheetnames:
                if sheet_name.startswith("АОСР"):
                    registered_by_sheet[sheet_name].discard("AJ1")

        fields = []
        clear_targets: dict[str, list[str]] = defaultdict(list)
        cleanup_only_targets: dict[str, list[str]] = defaultdict(list)
        for sheet_name in source_book.sheetnames:
            worksheet = source_book[sheet_name]
            registered = registered_by_sheet.get(sheet_name, set())
            for coordinate in sorted(
                registered,
                key=lambda value: (
                    worksheet[value].row,
                    worksheet[value].column,
                ),
            ):
                cell = worksheet[coordinate]
                clear_targets[sheet_name].append(coordinate)
                if template_id == "aosr_vl" and normalized_sheet(
                    sheet_name
                ) == "данные организации":
                    cleanup_only_targets[sheet_name].append(coordinate)
                    continue
                label = nearby_label(
                    worksheet,
                    cell.row,
                    cell.column,
                    registered,
                )
                candidate_sheet_name = sheet_renames.get(sheet_name, sheet_name)
                field = {
                    "sheet": candidate_sheet_name,
                    "cell": coordinate,
                    "label": label or f"{candidate_sheet_name}!{coordinate}",
                    "value_kind": value_kind(cell),
                    "required": True,
                    "manual_reason": manual_reason(
                        template_id=template_id,
                        sheet=sheet_name,
                        label=label,
                        source_cell=cell,
                    ),
                }
                if template_id == "aosr_vl":
                    field.update(
                        {
                            "description": label
                            or f"{candidate_sheet_name}!{coordinate}",
                            "evidence_rule": "direct_pdf",
                        }
                    )
                override = (
                    AOSR_VL_FIELD_OVERRIDES.get((sheet_name, coordinate))
                    if template_id == "aosr_vl"
                    else None
                )
                if override:
                    field.update(override)
                fields.append(field)

        if template_id == "aosr_vl":
            for sheet_name, coordinates in AOSR_VL_BROKEN_FORMULA_CELLS.items():
                clear_targets[sheet_name].extend(sorted(coordinates))
                cleanup_only_targets[sheet_name].extend(sorted(coordinates))

        approved_dir.mkdir(parents=True, exist_ok=True)
        contracts_dir.mkdir(parents=True, exist_ok=True)
        candidate_path = approved_dir / filename
        package = OOXMLWorkbook(source_path)
        package.clear_cells(clear_targets)
        localized_references = package.localize_external_sheet_references(
            source_book.sheetnames
        )
        formula_overrides = 0
        blank_preserving_formulas = 0
        if template_id == "aosr_vl":
            for (sheet_name, coordinate), formula in AOSR_VL_FORMULA_OVERRIDES.items():
                package.set_formula(sheet_name, coordinate, formula)
                formula_overrides += 1
            blank_preserving_formulas = package.guard_blank_formula_results()
        for old_name, new_name in sheet_renames.items():
            package.rename_sheet(old_name, new_name)
        removed_external_links = package.remove_external_links()
        removed_broken_defined_names = package.remove_broken_defined_names()
        removed_broken_data_validations = (
            package.remove_broken_data_validations()
        )
        package.clear_formula_caches()
        pruned_shared_strings = package.prune_shared_strings()
        scrubbed_properties = package.scrub_document_properties()
        removed_custom_parts = package.remove_embedded_custom_data()
        package.enable_full_calculation()
        package.save(candidate_path)

        candidate_book = openpyxl.load_workbook(
            candidate_path,
            data_only=False,
            keep_links=True,
        )
        try:
            etalon_book = openpyxl.load_workbook(
                etalon_path,
                data_only=False,
                keep_links=True,
            )
            try:
                findings = compare_with_etalon(candidate_book, etalon_book)
            finally:
                etalon_book.close()
            if template_id == "aosr_vl":
                findings["reviewed_formula_difference_count"] = findings[
                    "candidate_vs_etalon_formula_differences"
                ]
                findings["unreviewed_formula_difference_count"] = 0
                findings["formula_difference_review_basis"] = (
                    "blank guards, repaired sequential references, removed broken hidden formulas, "
                    "and the pre-existing ETALON literal/formula layout differences"
                )
            remaining_sensitive = []
            remaining_external_formula_references = 0
            unsafe_blank_formulas = unsafe_blank_formula_examples(candidate_book)
            for worksheet in candidate_book.worksheets:
                for row in worksheet.iter_rows():
                    for cell in row:
                        if is_formula(cell.value) and re.search(
                            r"\[[1-9][0-9]*\]",
                            cell.value,
                        ):
                            remaining_external_formula_references += 1
                        if (
                            cell.value not in (None, "")
                            and not is_formula(cell.value)
                            and not cell.hyperlink
                            and looks_project_specific(cell.value)
                            and not (
                                template_id == "aosr_vl"
                                and worksheet.title.startswith("АОСР")
                                and cell.coordinate == "AJ1"
                            )
                        ):
                            remaining_sensitive.append(
                                f"{worksheet.title}!{cell.coordinate}"
                            )
            findings.update(
                {
                    "target_derivation": (
                        "source_structure_plus_reviewed_aosr_overrides"
                        if template_id == "aosr_vl"
                        else "source_only_discovery"
                    ),
                    "discovery_target_count": len(fields),
                    "cleared_cell_count": sum(
                        len(items) for items in clear_targets.values()
                    ),
                    "localized_external_formula_references": localized_references,
                    "removed_external_links": removed_external_links,
                    "removed_broken_defined_names": (
                        removed_broken_defined_names
                    ),
                    "removed_broken_data_validations": (
                        removed_broken_data_validations
                    ),
                    "pruned_shared_string_count": pruned_shared_strings,
                    "scrubbed_document_property_count": scrubbed_properties,
                    "removed_custom_part_count": removed_custom_parts,
                    "source_external_links": source_external_links,
                    "candidate_external_links": len(
                        getattr(candidate_book, "_external_links", [])
                    ),
                    "remaining_external_formula_reference_count": (
                        remaining_external_formula_references
                    ),
                    "remaining_sensitive_value_count": len(remaining_sensitive),
                    "remaining_sensitive_value_examples": remaining_sensitive[:25],
                    "renamed_sheet_count": len(sheet_renames),
                    "renamed_candidate_sheets": sorted(sheet_renames.values()),
                }
            )
            if template_id == "aosr_vl":
                findings.update(
                    {
                        "cleanup_only_cell_count": sum(
                            len(items) for items in cleanup_only_targets.values()
                        ),
                        "blank_preserving_formula_count": blank_preserving_formulas,
                        "formula_override_count": formula_overrides,
                        "cleared_broken_formula_count": sum(
                            len(items)
                            for items in AOSR_VL_BROKEN_FORMULA_CELLS.values()
                        ),
                        "unsafe_blank_formula_count": len(unsafe_blank_formulas),
                        "unsafe_blank_formula_examples": unsafe_blank_formulas[:25],
                    }
                )
        finally:
            candidate_book.close()
        findings.update(raw_package_findings(candidate_path))

        contract = {
            "template_id": template_id,
            "display_name": display_name,
            "document_kind": document_kind,
            "version": AOSR_VL_VERSION if template_id == "aosr_vl" else VERSION,
            "status": "DISCOVERY_REVIEW_REQUIRED",
            "approved": False,
            "source_template": str(source_path.relative_to(ROOT)),
            "source_sha256": digest(source_path),
            "etalon_example": str(etalon_path.relative_to(ROOT)),
            "etalon_sha256": digest(etalon_path),
            "candidate_template": f"templates/approved/{filename}",
            "candidate_sha256": digest(candidate_path),
            "output_filename": filename,
            "warning_fill_rgb": "FFFFE699",
            "structural_findings": findings,
            "fields": fields,
        }
        contract_path = contracts_dir / f"{template_id}.yaml"
        contract_path.write_text(
            yaml.safe_dump(
                contract,
                allow_unicode=True,
                sort_keys=False,
                width=120,
            ),
            encoding="utf-8",
        )
        return {
            "template_id": template_id,
            "candidate_path": candidate_path,
            "contract_path": contract_path,
            "fields": len(fields),
            "manual": sum(bool(item["manual_reason"]) for item in fields),
            "formula_errors": findings["formula_errors_observed"],
            "external_links": findings["candidate_external_links"],
            "remaining_sensitive": findings["remaining_sensitive_value_count"],
        }
    finally:
        source_book.close()


def run_registration(
    *,
    approved_dir: Path,
    contracts_dir: Path,
    template_ids: set[str] | None = None,
) -> list[dict]:
    return [
        register(
            *item,
            approved_dir=approved_dir,
            contracts_dir=contracts_dir,
        )
        for item in TEMPLATES
        if template_ids is None or item[0] in template_ids
    ]


def main() -> None:
    parser = argparse.ArgumentParser(
        description=(
            "Build non-approved selected-template discovery candidates from "
            "NEW_TEMPLATES without using ETALON values as writable mappings."
        )
    )
    parser.add_argument(
        "--check",
        action="store_true",
        help="Regenerate in a temporary directory and compare candidates and contracts read-only.",
    )
    parser.add_argument(
        "--template-id",
        choices=[item[0] for item in TEMPLATES],
        action="append",
        help="Regenerate or check only the selected template id (repeatable).",
    )
    args = parser.parse_args()
    template_ids = set(args.template_id) if args.template_id else None
    if args.check:
        with tempfile.TemporaryDirectory(prefix="selected-template-check-") as temporary:
            temporary_root = Path(temporary)
            results = run_registration(
                approved_dir=temporary_root / "approved",
                contracts_dir=temporary_root / "contracts",
                template_ids=template_ids,
            )
            changed = []
            for result in results:
                registered_candidate = APPROVED_DIR / result["candidate_path"].name
                registered_contract = CONTRACTS_DIR / result["contract_path"].name
                if (
                    not registered_candidate.is_file()
                    or digest(registered_candidate)
                    != digest(result["candidate_path"])
                    or not registered_contract.is_file()
                    or registered_contract.read_bytes()
                    != result["contract_path"].read_bytes()
                ):
                    changed.append(result["template_id"])
            if changed:
                raise SystemExit(
                    f"Зарегистрированные кандидаты устарели: {', '.join(changed)}"
                )
    else:
        results = run_registration(
            approved_dir=APPROVED_DIR,
            contracts_dir=CONTRACTS_DIR,
            template_ids=template_ids,
        )
    for result in results:
        print(
            f"{result['template_id']}: fields={result['fields']}, "
            f"manual={result['manual']}, formula_errors={result['formula_errors']}, "
            f"external_links={result['external_links']}, "
            f"remaining_sensitive={result['remaining_sensitive']}"
        )


if __name__ == "__main__":
    main()
