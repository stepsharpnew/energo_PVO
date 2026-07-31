from __future__ import annotations

import copy
import hashlib
import io
import re
import shutil
import tempfile
import zipfile
from dataclasses import dataclass
from pathlib import Path, PurePosixPath

import openpyxl
import yaml
from lxml import etree

from .domain import Claim, DocumentPlan, NeedInputQuestion, WorkItem


MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
PKG_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
NS = {"x": MAIN_NS, "r": REL_NS, "pr": PKG_REL_NS}


@dataclass
class TemplateContract:
    template_id: str
    source_template: str
    version: str
    approved: bool
    output_filename: str
    allowed_sheets: list[str]
    candidate_sheets: list[str]
    common_fields: dict[str, str | list[str]]
    sheets: dict[str, dict]
    clear_cells: dict[str, list[str]]
    forbidden_tokens: list[str]
    sha256: str | None

    @classmethod
    def load(cls, path: Path) -> "TemplateContract":
        data = yaml.safe_load(path.read_text(encoding="utf-8"))
        allowed_sheets = list(data["allowed_sheets"])
        contract = cls(
            template_id=data["template_id"],
            source_template=data["source_template"],
            version=str(data["version"]),
            approved=bool(data.get("approved", False)),
            output_filename=data["output_filename"],
            allowed_sheets=allowed_sheets,
            candidate_sheets=list(data.get("candidate_sheets", allowed_sheets)),
            common_fields=dict(data.get("common_fields", {})),
            sheets=dict(data["sheets"]),
            clear_cells={str(sheet): list(cells) for sheet, cells in (data.get("clear_cells") or {}).items()},
            forbidden_tokens=list(data.get("forbidden_tokens", [])),
            sha256=data.get("sha256"),
        )
        if not set(contract.allowed_sheets).issubset(set(contract.candidate_sheets)):
            raise ValueError(f"allowed_sheets выходит за candidate_sheets: {contract.template_id}")
        if not set(contract.candidate_sheets).issubset(set(contract.sheets)):
            raise ValueError(f"Для candidate_sheets нет mappings: {contract.template_id}")
        return contract

    def writable_cells(self) -> set[str]:
        targets: set[str] = set()
        for value in self.common_fields.values():
            targets.update(value if isinstance(value, list) else [value])

        def mapping_cells(value: object) -> list[str]:
            if isinstance(value, str):
                return [value] if re.fullmatch(r"[A-Z]{1,3}[1-9][0-9]*", value) else []
            if isinstance(value, list):
                return [cell for item in value for cell in mapping_cells(item)]
            if isinstance(value, dict):
                return [
                    cell
                    for key, item in value.items()
                    if key not in {"suffix_value", "separator_value"}
                    for cell in mapping_cells(item)
                ]
            return []

        for sheet, mapping in self.sheets.items():
            targets.update(f"{sheet}!{cell}" for cell in mapping_cells(mapping))
        for sheet, cells in self.clear_cells.items():
            targets.update(f"{sheet}!{cell}" for cell in cells)
        return targets


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _resolve_target(base: str, target: str) -> str:
    if target.startswith("/"):
        return target.lstrip("/")
    path = PurePosixPath(base).parent / target
    parts: list[str] = []
    for part in path.parts:
        if part == "..":
            if parts:
                parts.pop()
        elif part != ".":
            parts.append(part)
    return "/".join(parts)


class OOXMLWorkbook:
    """Patch cell values and sheet visibility without reserializing the workbook package."""

    def __init__(self, source: Path | bytes):
        self.source = source
        archive_source = io.BytesIO(source) if isinstance(source, bytes) else source
        with zipfile.ZipFile(archive_source) as archive:
            self.parts = {name: archive.read(name) for name in archive.namelist()}
            self.zip_infos = {name: copy.copy(archive.getinfo(name)) for name in archive.namelist()}
        self.workbook = etree.fromstring(self.parts["xl/workbook.xml"])
        rels = etree.fromstring(self.parts["xl/_rels/workbook.xml.rels"])
        targets = {
            rel.get("Id"): _resolve_target("xl/workbook.xml", rel.get("Target"))
            for rel in rels.findall(f"{{{PKG_REL_NS}}}Relationship")
        }
        self.sheet_parts: dict[str, str] = {}
        for sheet in self.workbook.xpath("//x:sheets/x:sheet", namespaces=NS):
            self.sheet_parts[sheet.get("name")] = targets[sheet.get(f"{{{REL_NS}}}id")]

    def _sheet_root(self, sheet_name: str) -> etree._Element:
        if sheet_name not in self.sheet_parts:
            raise ValueError(f"Лист не найден: {sheet_name}")
        part = self.sheet_parts[sheet_name]
        return etree.fromstring(self.parts[part])

    def _save_sheet_root(self, sheet_name: str, root: etree._Element) -> None:
        self.parts[self.sheet_parts[sheet_name]] = etree.tostring(
            root, xml_declaration=True, encoding="UTF-8", standalone=True
        )

    def _cell(self, sheet_name: str, coordinate: str, *, create: bool) -> tuple[etree._Element, etree._Element | None]:
        if not re.fullmatch(r"[A-Z]{1,3}[1-9][0-9]*", coordinate):
            raise ValueError(f"Некорректная ячейка: {coordinate}")
        root = self._sheet_root(sheet_name)
        sheet_data = root.find(f"{{{MAIN_NS}}}sheetData")
        if sheet_data is None:
            raise ValueError(f"В листе {sheet_name} отсутствует sheetData")
        row_number = int(re.search(r"\d+", coordinate).group())
        row = sheet_data.find(f"{{{MAIN_NS}}}row[@r='{row_number}']")
        if row is None and create:
            row = etree.SubElement(sheet_data, f"{{{MAIN_NS}}}row", r=str(row_number))
        if row is None:
            return root, None
        cell = row.find(f"{{{MAIN_NS}}}c[@r='{coordinate}']")
        if cell is None and create:
            cell = etree.SubElement(row, f"{{{MAIN_NS}}}c", r=coordinate)
        return root, cell

    @staticmethod
    def _clear_cell_payload(cell: etree._Element) -> None:
        for child in list(cell):
            if child.tag in {f"{{{MAIN_NS}}}f", f"{{{MAIN_NS}}}v", f"{{{MAIN_NS}}}is"}:
                cell.remove(child)
        cell.attrib.pop("t", None)

    def set_cell(self, sheet_name: str, coordinate: str, value: str | int | float) -> None:
        root, cell = self._cell(sheet_name, coordinate, create=True)
        assert cell is not None
        self._clear_cell_payload(cell)
        if isinstance(value, (int, float)) and not isinstance(value, bool):
            etree.SubElement(cell, f"{{{MAIN_NS}}}v").text = str(value)
        else:
            cell.set("t", "inlineStr")
            inline = etree.SubElement(cell, f"{{{MAIN_NS}}}is")
            text = etree.SubElement(inline, f"{{{MAIN_NS}}}t")
            text.set("{http://www.w3.org/XML/1998/namespace}space", "preserve")
            text.text = str(value)
        self._save_sheet_root(sheet_name, root)

    def clear_cell(self, sheet_name: str, coordinate: str) -> None:
        """Clear a cell value/formula while preserving its style and surrounding structure."""

        root, cell = self._cell(sheet_name, coordinate, create=False)
        if cell is None:
            return
        self._clear_cell_payload(cell)
        self._save_sheet_root(sheet_name, root)

    def clear_cells(self, targets: dict[str, list[str] | set[str]]) -> None:
        """Clear many cells while parsing each worksheet only once."""

        for sheet_name, coordinates in targets.items():
            root = self._sheet_root(sheet_name)
            sheet_data = root.find(f"{{{MAIN_NS}}}sheetData")
            if sheet_data is None:
                raise ValueError(f"В листе {sheet_name} отсутствует sheetData")
            changed = False
            for coordinate in coordinates:
                if not re.fullmatch(r"[A-Z]{1,3}[1-9][0-9]*", coordinate):
                    raise ValueError(f"Некорректная ячейка: {coordinate}")
                row_number = int(re.search(r"\d+", coordinate).group())
                row = sheet_data.find(f"{{{MAIN_NS}}}row[@r='{row_number}']")
                if row is None:
                    continue
                cell = row.find(f"{{{MAIN_NS}}}c[@r='{coordinate}']")
                if cell is None:
                    continue
                self._clear_cell_payload(cell)
                changed = True
            if changed:
                self._save_sheet_root(sheet_name, root)

    def highlight_cells(
        self,
        targets: dict[str, list[str] | set[str]],
        *,
        rgb: str = "FFFFE699",
    ) -> int:
        """Apply a warning fill while preserving every other style component."""

        rgb = rgb.upper()
        if not re.fullmatch(r"[0-9A-F]{8}", rgb):
            raise ValueError("Цвет заливки должен быть ARGB в формате AARRGGBB")
        styles_part = "xl/styles.xml"
        if styles_part not in self.parts:
            raise ValueError("В книге отсутствует таблица стилей")
        styles = etree.fromstring(self.parts[styles_part])
        fills = styles.find(f"{{{MAIN_NS}}}fills")
        cell_xfs = styles.find(f"{{{MAIN_NS}}}cellXfs")
        if fills is None or cell_xfs is None:
            raise ValueError("В книге отсутствуют fills/cellXfs")

        fill_id: int | None = None
        for index, fill in enumerate(fills):
            pattern = fill.find(f"{{{MAIN_NS}}}patternFill")
            foreground = pattern.find(f"{{{MAIN_NS}}}fgColor") if pattern is not None else None
            if (
                pattern is not None
                and pattern.get("patternType") == "solid"
                and foreground is not None
                and (foreground.get("rgb") or "").upper() == rgb
            ):
                fill_id = index
                break
        if fill_id is None:
            fill = etree.SubElement(fills, f"{{{MAIN_NS}}}fill")
            pattern = etree.SubElement(fill, f"{{{MAIN_NS}}}patternFill", patternType="solid")
            etree.SubElement(pattern, f"{{{MAIN_NS}}}fgColor", rgb=rgb)
            etree.SubElement(pattern, f"{{{MAIN_NS}}}bgColor", indexed="64")
            fill_id = len(fills) - 1
            fills.set("count", str(len(fills)))

        style_remap: dict[int, int] = {}
        changed_cells = 0
        for sheet_name, coordinates in targets.items():
            root = self._sheet_root(sheet_name)
            sheet_data = root.find(f"{{{MAIN_NS}}}sheetData")
            if sheet_data is None:
                raise ValueError(f"В листе {sheet_name} отсутствует sheetData")
            changed_sheet = False
            for coordinate in coordinates:
                if not re.fullmatch(r"[A-Z]{1,3}[1-9][0-9]*", coordinate):
                    raise ValueError(f"Некорректная ячейка: {coordinate}")
                row_number = int(re.search(r"\d+", coordinate).group())
                row = sheet_data.find(f"{{{MAIN_NS}}}row[@r='{row_number}']")
                if row is None:
                    row = etree.SubElement(sheet_data, f"{{{MAIN_NS}}}row", r=str(row_number))
                cell = row.find(f"{{{MAIN_NS}}}c[@r='{coordinate}']")
                if cell is None:
                    cell = etree.SubElement(row, f"{{{MAIN_NS}}}c", r=coordinate)
                old_style = int(cell.get("s", "0"))
                if old_style >= len(cell_xfs):
                    raise ValueError(f"Некорректный индекс стиля {old_style}: {sheet_name}!{coordinate}")
                new_style = style_remap.get(old_style)
                if new_style is None:
                    style = copy.deepcopy(cell_xfs[old_style])
                    style.set("fillId", str(fill_id))
                    style.set("applyFill", "1")
                    cell_xfs.append(style)
                    new_style = len(cell_xfs) - 1
                    style_remap[old_style] = new_style
                cell.set("s", str(new_style))
                changed_cells += 1
                changed_sheet = True
            if changed_sheet:
                self._save_sheet_root(sheet_name, root)
        cell_xfs.set("count", str(len(cell_xfs)))
        self.parts[styles_part] = etree.tostring(
            styles, xml_declaration=True, encoding="UTF-8", standalone=True
        )
        return changed_cells

    def set_formula(self, sheet_name: str, coordinate: str, formula: str) -> None:
        """Set an OOXML formula without a cached result so Excel recalculates it."""

        root, cell = self._cell(sheet_name, coordinate, create=True)
        assert cell is not None
        self._clear_cell_payload(cell)
        etree.SubElement(cell, f"{{{MAIN_NS}}}f").text = formula.removeprefix("=")
        self._save_sheet_root(sheet_name, root)

    def localize_external_sheet_references(self, sheet_names: list[str]) -> int:
        """Point formulas like '[1]Sheet'!A1 at an identically named local sheet."""

        known = set(sheet_names)
        quoted = re.compile(r"'\[(?:[1-9][0-9]*)\](?P<sheet>(?:''|[^'])+)'!")
        unquoted = re.compile(r"\[(?:[1-9][0-9]*)\](?P<sheet>[^'!\[\]]+)!")

        def replace_quoted(match: re.Match[str]) -> str:
            sheet = match.group("sheet").replace("''", "'")
            if sheet not in known:
                return match.group(0)
            return f"'{sheet.replace(chr(39), chr(39) * 2)}'!"

        def replace_unquoted(match: re.Match[str]) -> str:
            sheet = match.group("sheet")
            if sheet not in known:
                return match.group(0)
            return f"'{sheet.replace(chr(39), chr(39) * 2)}'!"

        changed = 0
        for sheet_name in sheet_names:
            root = self._sheet_root(sheet_name)
            changed_sheet = False
            for formula in root.iter(f"{{{MAIN_NS}}}f"):
                if not formula.text:
                    continue
                updated = quoted.sub(replace_quoted, formula.text)
                updated = unquoted.sub(replace_unquoted, updated)
                if updated != formula.text:
                    formula.text = updated
                    changed += 1
                    changed_sheet = True
            if changed_sheet:
                self._save_sheet_root(sheet_name, root)
        return changed

    def rename_sheet(self, old_name: str, new_name: str) -> None:
        """Rename a worksheet and update textual references throughout the OOXML package."""

        if old_name not in self.sheet_parts:
            raise ValueError(f"Лист не найден: {old_name}")
        if (
            not new_name
            or len(new_name) > 31
            or re.search(r"[\[\]:*?/\\]", new_name)
            or new_name in self.sheet_parts
        ):
            raise ValueError(f"Некорректное новое имя листа: {new_name}")
        sheet = next(
            item
            for item in self.workbook.xpath("//x:sheets/x:sheet", namespaces=NS)
            if item.get("name") == old_name
        )
        sheet.set("name", new_name)
        old_bytes = old_name.encode("utf-8")
        new_bytes = new_name.encode("utf-8")
        for part, payload in list(self.parts.items()):
            if part.endswith((".xml", ".rels")) and old_bytes in payload:
                self.parts[part] = payload.replace(old_bytes, new_bytes)
        self.sheet_parts[new_name] = self.sheet_parts.pop(old_name)

    def remove_defined_name(self, name: str) -> None:
        defined_names = self.workbook.find(f"{{{MAIN_NS}}}definedNames")
        if defined_names is None:
            return
        for item in list(defined_names):
            if item.get("name") == name:
                defined_names.remove(item)
        if len(defined_names) == 0:
            self.workbook.remove(defined_names)

    def remove_broken_defined_names(self) -> int:
        """Remove workbook names whose formulas already contain a broken #REF!."""

        defined_names = self.workbook.find(f"{{{MAIN_NS}}}definedNames")
        if defined_names is None:
            return 0
        removed = 0
        for item in list(defined_names):
            if "#REF!" in (item.text or "").upper():
                defined_names.remove(item)
                removed += 1
        if len(defined_names) == 0:
            self.workbook.remove(defined_names)
        return removed

    def remove_broken_data_validations(self) -> int:
        """Remove validation rules containing broken or source-machine references."""

        removed = 0
        for sheet_name in list(self.sheet_parts):
            root = self._sheet_root(sheet_name)
            changed = False
            for validation in list(root.iter()):
                if etree.QName(validation).localname != "dataValidation":
                    continue
                payload = etree.tostring(validation, encoding="unicode")
                if "#REF!" not in payload.upper() and not re.search(
                    r"[A-Z]:\\",
                    payload,
                    flags=re.IGNORECASE,
                ):
                    continue
                parent = validation.getparent()
                if parent is not None:
                    parent.remove(validation)
                    count = parent.get("count")
                    if count and count.isdigit():
                        parent.set("count", str(max(0, int(count) - 1)))
                    removed += 1
                    changed = True
            if changed:
                self._save_sheet_root(sheet_name, root)
        return removed

    def clear_formula_caches(self) -> None:
        """Remove cached formula results to prevent values from a prior object leaking into previews."""

        for sheet_name in self.sheet_parts:
            root = self._sheet_root(sheet_name)
            changed = False
            for cell in root.xpath("//x:c[x:f]", namespaces=NS):
                value = cell.find(f"{{{MAIN_NS}}}v")
                if value is not None:
                    cell.remove(value)
                    changed = True
            if changed:
                self._save_sheet_root(sheet_name, root)

    def clear_fill_colors(
        self,
        *,
        rgb_values: set[str] | None = None,
        indexed_values: set[int] | None = None,
    ) -> int:
        """Remove selected solid fills while preserving every other cell-style component.

        The implementation remaps affected cells to an equivalent ``cellXf`` with
        ``fillId=0``. It avoids reserializing worksheets or rebuilding styles, which
        is important for the legacy pilot templates.
        """

        rgb_values = {value.upper() for value in (rgb_values or {"FFFFFF00", "00FFFF00"})}
        indexed_values = indexed_values or {6, 13}
        styles_part = "xl/styles.xml"
        if styles_part not in self.parts:
            return 0
        styles = etree.fromstring(self.parts[styles_part])
        fills = styles.find(f"{{{MAIN_NS}}}fills")
        cell_xfs = styles.find(f"{{{MAIN_NS}}}cellXfs")
        if fills is None or cell_xfs is None:
            return 0

        selected_fill_ids: set[int] = set()
        for index, fill in enumerate(fills):
            pattern = fill.find(f"{{{MAIN_NS}}}patternFill")
            if pattern is None:
                continue
            foreground = pattern.find(f"{{{MAIN_NS}}}fgColor")
            if foreground is None:
                continue
            rgb = (foreground.get("rgb") or "").upper()
            indexed = foreground.get("indexed")
            if rgb in rgb_values or (indexed is not None and int(indexed) in indexed_values):
                selected_fill_ids.add(index)
        if not selected_fill_ids:
            return 0

        def style_signature(style: etree._Element) -> bytes:
            normalized = copy.deepcopy(style)
            normalized.attrib.pop("fillId", None)
            normalized.attrib.pop("applyFill", None)
            return etree.tostring(normalized)

        no_fill_by_signature: dict[bytes, int] = {}
        for index, style in enumerate(cell_xfs):
            if int(style.get("fillId", "0")) == 0:
                no_fill_by_signature.setdefault(style_signature(style), index)

        style_remap: dict[int, int] = {}
        for index, style in enumerate(list(cell_xfs)):
            if int(style.get("fillId", "0")) not in selected_fill_ids:
                continue
            signature = style_signature(style)
            replacement = no_fill_by_signature.get(signature)
            if replacement is None:
                clean_style = copy.deepcopy(style)
                clean_style.set("fillId", "0")
                clean_style.attrib.pop("applyFill", None)
                cell_xfs.append(clean_style)
                replacement = len(cell_xfs) - 1
                no_fill_by_signature[signature] = replacement
            style_remap[index] = replacement
        if not style_remap:
            return 0
        cell_xfs.set("count", str(len(cell_xfs)))
        self.parts[styles_part] = etree.tostring(
            styles, xml_declaration=True, encoding="UTF-8", standalone=True
        )

        changed_cells = 0
        for sheet_name in self.sheet_parts:
            root = self._sheet_root(sheet_name)
            changed_sheet = False
            for cell in root.xpath("//x:c[@s]", namespaces=NS):
                style_index = int(cell.get("s"))
                replacement = style_remap.get(style_index)
                if replacement is None:
                    continue
                cell.set("s", str(replacement))
                changed_cells += 1
                changed_sheet = True
            if changed_sheet:
                self._save_sheet_root(sheet_name, root)
        return changed_cells

    def remove_external_links(self) -> int:
        """Remove external-link package parts after every dependent formula was localized."""

        external_parts = [
            name for name in self.parts if name.startswith("xl/externalLinks/")
        ]
        external_references = self.workbook.find(f"{{{MAIN_NS}}}externalReferences")
        if external_references is not None:
            self.workbook.remove(external_references)

        rels_name = "xl/_rels/workbook.xml.rels"
        rels = etree.fromstring(self.parts[rels_name])
        for relationship in list(rels):
            if relationship.get("Type", "").endswith("/externalLink"):
                rels.remove(relationship)
        self.parts[rels_name] = etree.tostring(
            rels, xml_declaration=True, encoding="UTF-8", standalone=True
        )

        content_types_name = "[Content_Types].xml"
        content_types = etree.fromstring(self.parts[content_types_name])
        for item in list(content_types):
            if item.get("PartName", "").startswith("/xl/externalLinks/"):
                content_types.remove(item)
        self.parts[content_types_name] = etree.tostring(
            content_types, xml_declaration=True, encoding="UTF-8", standalone=True
        )
        for part in external_parts:
            del self.parts[part]
        return len(
            [
                name
                for name in external_parts
                if re.fullmatch(r"xl/externalLinks/externalLink\d+\.xml", name)
            ]
        )

    def remove_calculation_chain(self) -> None:
        """Drop calcChain so repaired formulas are rebuilt by Excel/LibreOffice."""

        calc_chain = "xl/calcChain.xml"
        if calc_chain not in self.parts:
            return
        del self.parts[calc_chain]
        rels_name = "xl/_rels/workbook.xml.rels"
        rels = etree.fromstring(self.parts[rels_name])
        for relationship in list(rels):
            if relationship.get("Type", "").endswith("/calcChain"):
                rels.remove(relationship)
        self.parts[rels_name] = etree.tostring(
            rels, xml_declaration=True, encoding="UTF-8", standalone=True
        )
        content_types_name = "[Content_Types].xml"
        content_types = etree.fromstring(self.parts[content_types_name])
        for item in list(content_types):
            if item.get("PartName") == "/xl/calcChain.xml":
                content_types.remove(item)
        self.parts[content_types_name] = etree.tostring(
            content_types, xml_declaration=True, encoding="UTF-8", standalone=True
        )

    def prune_shared_strings(self) -> int:
        """Remove unreferenced shared strings, including cleared prior-object values."""

        part = "xl/sharedStrings.xml"
        if part not in self.parts:
            return 0
        roots: dict[str, etree._Element] = {}
        references: list[tuple[etree._Element, int]] = []
        for sheet_name, sheet_part in self.sheet_parts.items():
            root = self._sheet_root(sheet_name)
            roots[sheet_part] = root
            for cell in root.xpath("//x:c[@t='s'][x:v]", namespaces=NS):
                value = cell.find(f"{{{MAIN_NS}}}v")
                if value is not None and value.text is not None:
                    references.append((value, int(value.text)))
        shared = etree.fromstring(self.parts[part])
        items = list(shared.findall(f"{{{MAIN_NS}}}si"))
        used = sorted({index for _, index in references})
        mapping = {old: new for new, old in enumerate(used)}
        for value, old in references:
            value.text = str(mapping[old])
        for item in items:
            shared.remove(item)
        for old in used:
            shared.append(items[old])
        shared.set("count", str(len(references)))
        shared.set("uniqueCount", str(len(used)))
        self.parts[part] = etree.tostring(
            shared, xml_declaration=True, encoding="UTF-8", standalone=True
        )
        for sheet_part, root in roots.items():
            self.parts[sheet_part] = etree.tostring(
                root, xml_declaration=True, encoding="UTF-8", standalone=True
            )
        return len(items) - len(used)

    def scrub_document_properties(self) -> int:
        """Remove author, editor, company, and other source-document metadata."""

        changed = 0
        core_name = "docProps/core.xml"
        if core_name in self.parts:
            core = etree.fromstring(self.parts[core_name])
            for item in list(core):
                core.remove(item)
                changed += 1
            creator = etree.SubElement(
                core,
                "{http://purl.org/dc/elements/1.1/}creator",
            )
            creator.text = "Executive Docs"
            modified_by = etree.SubElement(
                core,
                "{http://schemas.openxmlformats.org/package/2006/metadata/core-properties}lastModifiedBy",
            )
            modified_by.text = "Executive Docs"
            self.parts[core_name] = etree.tostring(
                core,
                xml_declaration=True,
                encoding="UTF-8",
                standalone=True,
            )
        app_name = "docProps/app.xml"
        if app_name in self.parts:
            app = etree.fromstring(self.parts[app_name])
            for item in app.iter():
                if etree.QName(item).localname in {
                    "Company",
                    "Manager",
                    "HyperlinkBase",
                } and item.text:
                    item.text = ""
                    changed += 1
            self.parts[app_name] = etree.tostring(
                app,
                xml_declaration=True,
                encoding="UTF-8",
                standalone=True,
            )
        return changed

    def remove_embedded_custom_data(self) -> int:
        """Drop custom XML/custom properties that can retain prior-project payloads."""

        removed_parts = {
            name
            for name in self.parts
            if name.startswith("customXml/") or name == "docProps/custom.xml"
        }
        if not removed_parts:
            return 0
        for name in removed_parts:
            del self.parts[name]

        content_types_name = "[Content_Types].xml"
        content_types = etree.fromstring(self.parts[content_types_name])
        for item in list(content_types):
            part_name = item.get("PartName", "").lstrip("/")
            if part_name in removed_parts:
                content_types.remove(item)
        self.parts[content_types_name] = etree.tostring(
            content_types,
            xml_declaration=True,
            encoding="UTF-8",
            standalone=True,
        )

        for name, payload in list(self.parts.items()):
            if not name.endswith(".rels"):
                continue
            relationships = etree.fromstring(payload)
            changed = False
            for relationship in list(relationships):
                rel_type = relationship.get("Type", "")
                target = relationship.get("Target", "")
                if (
                    rel_type.endswith(("/customXml", "/customXmlProps", "/custom-properties"))
                    or "customXml/" in target
                    or target.endswith("docProps/custom.xml")
                ):
                    relationships.remove(relationship)
                    changed = True
            if changed:
                self.parts[name] = etree.tostring(
                    relationships,
                    xml_declaration=True,
                    encoding="UTF-8",
                    standalone=True,
                )
        return len(removed_parts)

    def set_visibility(self, selected: set[str], candidates: set[str]) -> None:
        for sheet in self.workbook.xpath("//x:sheets/x:sheet", namespaces=NS):
            name = sheet.get("name")
            if name not in candidates:
                continue
            sheet.set("state", "visible" if name in selected else "hidden")

    def set_only_visible(self, selected: str) -> None:
        if selected not in self.sheet_parts:
            raise ValueError(f"Лист не найден: {selected}")
        selected_index = 0
        for index, sheet in enumerate(self.workbook.xpath("//x:sheets/x:sheet", namespaces=NS)):
            is_selected = sheet.get("name") == selected
            sheet.set("state", "visible" if is_selected else "hidden")
            if is_selected:
                selected_index = index
            root = self._sheet_root(sheet.get("name"))
            for view in root.xpath("//x:sheetViews/x:sheetView", namespaces=NS):
                if is_selected:
                    view.set("tabSelected", "1")
                else:
                    view.attrib.pop("tabSelected", None)
            self.parts[self.sheet_parts[sheet.get("name")]] = etree.tostring(
                root, xml_declaration=True, encoding="UTF-8", standalone=True
            )
        for view in self.workbook.xpath("//x:bookViews/x:workbookView", namespaces=NS):
            view.set("activeTab", str(selected_index))
        defined_names = self.workbook.find(f"{{{MAIN_NS}}}definedNames")
        if defined_names is not None:
            for item in list(defined_names):
                if item.get("name") == "_xlnm.Print_Area" and item.get("localSheetId") != str(selected_index):
                    defined_names.remove(item)

    def configure_printing(self, sheet_name: str) -> None:
        root = self._sheet_root(sheet_name)
        sheet_pr = root.find(f"{{{MAIN_NS}}}sheetPr")
        if sheet_pr is None:
            sheet_pr = etree.Element(f"{{{MAIN_NS}}}sheetPr")
            root.insert(0, sheet_pr)
        setup_pr = sheet_pr.find(f"{{{MAIN_NS}}}pageSetUpPr")
        if setup_pr is None:
            setup_pr = etree.SubElement(sheet_pr, f"{{{MAIN_NS}}}pageSetUpPr")
        setup_pr.set("fitToPage", "1")
        page_setup = root.find(f"{{{MAIN_NS}}}pageSetup")
        if page_setup is None:
            page_setup = etree.SubElement(root, f"{{{MAIN_NS}}}pageSetup")
        page_setup.attrib.pop("scale", None)
        page_setup.set("paperSize", "9")
        page_setup.set("orientation", "portrait")
        page_setup.set("fitToWidth", "1")
        if not page_setup.get("fitToHeight"):
            page_setup.set("fitToHeight", "0")
        self.parts[self.sheet_parts[sheet_name]] = etree.tostring(
            root, xml_declaration=True, encoding="UTF-8", standalone=True
        )

    def enable_full_calculation(self) -> None:
        calc = self.workbook.find(f"{{{MAIN_NS}}}calcPr")
        if calc is None:
            calc = etree.SubElement(self.workbook, f"{{{MAIN_NS}}}calcPr")
        calc.set("calcMode", "auto")
        calc.set("fullCalcOnLoad", "1")
        calc.set("forceFullCalc", "1")

    def save(self, destination: Path) -> None:
        self.parts["xl/workbook.xml"] = etree.tostring(
            self.workbook, xml_declaration=True, encoding="UTF-8", standalone=True
        )
        destination.parent.mkdir(parents=True, exist_ok=True)
        with zipfile.ZipFile(destination, "w", zipfile.ZIP_DEFLATED) as archive:
            for name, data in self.parts.items():
                info = copy.copy(self.zip_infos.get(name))
                if info is None:
                    info = zipfile.ZipInfo(name, date_time=(1980, 1, 1, 0, 0, 0))
                    info.compress_type = zipfile.ZIP_DEFLATED
                archive.writestr(info, data)


class ExcelGenerator:
    def __init__(self, root: Path, contracts_dir: Path, approved_dir: Path):
        self.root = root
        self.contracts_dir = contracts_dir
        self.approved_dir = approved_dir

    def contract(self, template_id: str) -> TemplateContract:
        path = self.contracts_dir / f"{template_id}.yaml"
        if not path.exists():
            raise ValueError(f"Контракт шаблона не найден: {template_id}")
        return TemplateContract.load(path)

    def template_path(self, contract: TemplateContract) -> Path:
        approved = self.approved_dir / Path(contract.source_template).name
        path = approved if approved.exists() else self.root / contract.source_template
        if contract.approved:
            if not approved.exists():
                raise ValueError(f"Утверждённая копия шаблона отсутствует: {approved}")
            if not contract.sha256:
                raise ValueError(f"В утверждённом контракте {contract.template_id} отсутствует SHA-256")
            if sha256(path) != contract.sha256:
                raise ValueError(f"Контрольная сумма шаблона {contract.template_id} не совпадает с контрактом")
        return path

    @staticmethod
    def _claim_map(claims: list[Claim]) -> dict[str, str]:
        return {
            claim.key: claim.normalized_value
            for claim in claims
            if claim.status.value in {"observed", "derived", "human_confirmed"}
        }

    def generate(
        self,
        plan: DocumentPlan,
        work_items: list[WorkItem],
        claims: list[Claim],
        destination: Path,
    ) -> tuple[Path, TemplateContract]:
        contract = self.contract(plan.template_id)
        source = self.template_path(contract)
        if not source.exists():
            raise ValueError(f"Шаблон отсутствует: {source}")
        selected = set(plan.selected_sheets)
        if plan.output_filename != contract.output_filename or Path(plan.output_filename).name != plan.output_filename:
            raise ValueError("Имя выходного файла не соответствует контракту")
        if not selected or not selected.issubset(set(contract.allowed_sheets)):
            raise ValueError("План содержит неразрешённые листы")
        if len(selected) != len(plan.selected_sheets) or len(set(plan.work_item_ids)) != len(plan.work_item_ids):
            raise ValueError("План содержит повторяющиеся листы или работы")
        if len(plan.work_item_ids) != len(plan.selected_sheets):
            raise ValueError("Каждому выбранному листу должна соответствовать одна работа")
        item_map = {item.id: item for item in work_items}
        expected_family = {"aosr_kl_04": "kl_04", "aosr_kl_6": "kl_6", "aosr_vrs": "vrs"}[contract.template_id]
        if any(item_id not in item_map or item_map[item_id].family != expected_family for item_id in plan.work_item_ids):
            raise ValueError("Работа отсутствует или не соответствует семейству шаблона")
        workbook = OOXMLWorkbook(source)
        act_sheets = {
            sheet_name
            for sheet_name in workbook.sheet_parts
            if sheet_name.casefold().startswith("аоср")
        }
        workbook.set_visibility(selected, set(contract.candidate_sheets) | act_sheets)
        claims_by_key = self._claim_map(claims)
        for field in plan.field_values:
            if claims_by_key.get(field.key) != field.value:
                raise ValueError(f"Поле плана не подтверждено Claim: {field.key}")
        for key, target in contract.common_fields.items():
            value = claims_by_key.get(key)
            for cell_target in target if isinstance(target, list) else [target]:
                sheet, cell = cell_target.split("!", 1)
                if value is None:
                    workbook.clear_cell(sheet, cell)
                else:
                    workbook.set_cell(sheet, cell, value)
        for offset, (sheet, item_id) in enumerate(zip(plan.selected_sheets, plan.work_item_ids)):
            item = item_map[item_id]
            mapping = contract.sheets[sheet]
            material_rows = mapping.get("material_rows")
            if material_rows is not None and (
                not isinstance(material_rows, list)
                or any(not isinstance(row, dict) for row in material_rows)
            ):
                raise ValueError(f"Некорректный material_rows в контракте: {contract.template_id}:{sheet}")
            if isinstance(material_rows, list) and len(item.materials) > len(material_rows):
                raise ValueError(
                    f"Для {sheet} предусмотрено материалов: {len(material_rows)}, "
                    f"получено: {len(item.materials)}"
                )
            for cell in contract.clear_cells.get(sheet, []):
                workbook.clear_cell(sheet, cell)
            for field in (
                "installation",
                "volume",
                "unit",
                "start_date",
                "end_date",
                "attachment",
                "material_name",
                "material_document",
                "subsequent_work",
            ):
                if mapping.get(field):
                    workbook.clear_cell(sheet, mapping[field])
            if isinstance(material_rows, list):
                for row in material_rows:
                    for field in ("name", "separator", "quantity", "unit", "quantity_unit", "document"):
                        if row.get(field):
                            workbook.clear_cell(sheet, row[field])
            workbook.set_cell(sheet, mapping.get("number", "C32"), plan.first_number + offset)
            if mapping.get("suffix"):
                workbook.set_cell(sheet, mapping["suffix"], mapping.get("suffix_value", ""))
            workbook.set_cell(sheet, mapping["work_name"], item.work_type)
            if item.installation and mapping.get("installation"):
                workbook.set_cell(sheet, mapping["installation"], item.installation)
            if item.volume is not None and mapping.get("volume"):
                workbook.set_cell(sheet, mapping["volume"], item.volume)
            if item.unit and mapping.get("unit"):
                workbook.set_cell(sheet, mapping["unit"], item.unit)
            if item.actual_start and mapping.get("start_date"):
                workbook.set_cell(sheet, mapping["start_date"], item.actual_start)
            if item.actual_end and mapping.get("end_date"):
                workbook.set_cell(sheet, mapping["end_date"], item.actual_end)
            if item.subsequent_work and mapping.get("subsequent_work"):
                workbook.set_cell(sheet, mapping["subsequent_work"], item.subsequent_work)
            if item.execution_scheme_id and mapping.get("attachment"):
                scheme_name = claims_by_key.get(f"artifact.{item.execution_scheme_id}.name", "Исполнительная схема")
                workbook.set_cell(sheet, mapping["attachment"], scheme_name)
            if isinstance(material_rows, list):
                for material, row in zip(item.materials, material_rows):
                    if row.get("name"):
                        workbook.set_cell(sheet, row["name"], material.name)
                    if row.get("separator"):
                        workbook.set_cell(
                            sheet,
                            row["separator"],
                            str(row.get("separator_value", "-")),
                        )
                    if material.quantity is not None and row.get("quantity"):
                        workbook.set_cell(sheet, row["quantity"], material.quantity)
                    if material.unit and row.get("unit"):
                        workbook.set_cell(sheet, row["unit"], material.unit)
                    if material.quantity is not None and row.get("quantity_unit"):
                        quantity_unit = material.quantity
                        if material.unit and not quantity_unit.casefold().endswith(
                            material.unit.casefold()
                        ):
                            quantity_unit += material.unit
                        workbook.set_cell(sheet, row["quantity_unit"], quantity_unit)
                    if material.quality_document and row.get("document"):
                        workbook.set_cell(sheet, row["document"], material.quality_document)
            elif item.materials and mapping.get("material_name"):
                names = "; ".join(material.name for material in item.materials)
                documents = "; ".join(
                    material.quality_document or "" for material in item.materials if material.quality_document
                )
                workbook.set_cell(sheet, mapping["material_name"], names)
                if documents and mapping.get("material_document"):
                    workbook.set_cell(sheet, mapping["material_document"], documents)
        workbook.enable_full_calculation()
        output = destination / plan.output_filename
        workbook.save(output)
        return output, contract

    @staticmethod
    def _draft_warning_categories(questions: list[NeedInputQuestion]) -> set[str]:
        categories: set[str] = set()
        for question in questions:
            if question.answer:
                continue
            text = f"{question.field_key} {question.prompt} {question.reason}".casefold()
            if any(token in text for token in ("actual.start", "actual.end", "дат")):
                categories.add("dates")
            if any(token in text for token in ("volume", "quantity", "объем", "объём")):
                categories.add("volumes")
            if any(token in text for token in ("material", "паспорт", "сертификат")):
                categories.add("materials")
            if any(token in text for token in ("change", "изменен", "изменён", "отклонен", "отклонён")):
                categories.add("changes")
            if any(token in text for token in ("profile", "профил", "подписант", "реквизит", "полномочи")):
                categories.add("profiles")
        return categories

    def generate_draft(
        self,
        plan: DocumentPlan,
        work_items: list[WorkItem],
        claims: list[Claim],
        questions: list[NeedInputQuestion],
        destination: Path,
    ) -> tuple[Path, TemplateContract]:
        """Generate a visibly marked non-final workbook from admissible values."""

        output, contract = self.generate(plan, work_items, claims, destination)
        workbook = OOXMLWorkbook(output)
        categories = self._draft_warning_categories(questions)
        claims_by_key = self._claim_map(claims)
        item_map = {item.id: item for item in work_items}
        selected_items = [item_map[item_id] for item_id in plan.work_item_ids]
        if any(not item.actual_start or not item.actual_end for item in selected_items):
            categories.add("dates")
        if any(item.volume is None for item in selected_items):
            categories.add("volumes")
        if any(item.change_state.value == "UNKNOWN" for item in selected_items):
            categories.add("changes")
        if any(
            material.quality_document is None
            for item in selected_items
            for material in item.materials
        ):
            categories.add("materials")
        if any(
            key not in claims_by_key
            for key in (
                "contractor.construction_control.name",
                "contractor.work_supervisor.name",
                "customer.construction_control.name",
                "customer.site_representative.name",
            )
        ):
            categories.add("profiles")
        labels = {
            "dates": "не подтверждены фактические даты",
            "volumes": "не подтверждены фактические объёмы",
            "materials": "не указаны паспорта/сертификаты материалов",
            "changes": "не подтверждено наличие отклонений от проекта",
            "profiles": "не утверждены реквизиты и подписанты",
        }
        warning_block = "ЧЕРНОВИК. НЕ ДЛЯ ПОДПИСАНИЯ."
        if categories:
            warning_block += "\n" + "\n".join(
                f"• {labels[category]}" for category in labels if category in categories
            )

        def set_common(key: str, value: str) -> None:
            target = contract.common_fields.get(key)
            if not target:
                return
            for cell_target in target if isinstance(target, list) else [target]:
                sheet, cell = cell_target.split("!", 1)
                workbook.set_cell(sheet, cell, value)

        if "dates" in categories:
            set_common("actual.start", "НЕ ПОДТВЕРЖДЕНО")
            set_common("actual.end", "НЕ ПОДТВЕРЖДЕНО")
        if "profiles" in categories:
            for key in contract.common_fields:
                if not key.startswith(("contractor.", "customer.", "designer.")):
                    continue
                observed = claims_by_key.get(key)
                value = f"ЧЕРНОВИК, НЕ ПОДТВЕРЖДЕНО: {observed}" if observed else "НЕ ПОДТВЕРЖДЕНО"
                set_common(key, value)

        for sheet, item_id in zip(plan.selected_sheets, plan.work_item_ids):
            item = item_map[item_id]
            mapping = contract.sheets[sheet]
            if "dates" in categories:
                if mapping.get("start_date"):
                    workbook.set_cell(sheet, mapping["start_date"], "НЕ ПОДТВЕРЖДЕНО")
                if mapping.get("end_date"):
                    workbook.set_cell(sheet, mapping["end_date"], "НЕ ПОДТВЕРЖДЕНО")
            if "volumes" in categories and mapping.get("volume"):
                project_value = item.volume or ""
                value = (
                    f"НЕ ПОДТВЕРЖДЕНО (по проекту: {project_value})"
                    if project_value
                    else "НЕ ПОДТВЕРЖДЕНО"
                )
                workbook.set_cell(sheet, mapping["volume"], value)
            if "materials" in categories:
                material_rows = mapping.get("material_rows")
                if isinstance(material_rows, list):
                    for index, row in enumerate(material_rows):
                        material = item.materials[index] if index < len(item.materials) else None
                        if row.get("document") and (material is None or not material.quality_document):
                            workbook.set_cell(sheet, row["document"], "ПАСПОРТ/СЕРТИФИКАТ НЕ УКАЗАН")
                elif mapping.get("material_document") and (
                    not item.materials or any(not material.quality_document for material in item.materials)
                ):
                    workbook.set_cell(
                        sheet,
                        mapping["material_document"],
                        "ПАСПОРТ/СЕРТИФИКАТ НЕ УКАЗАН",
                    )
            if mapping.get("attachment"):
                attachment = ""
                if item.execution_scheme_id:
                    attachment = claims_by_key.get(
                        f"artifact.{item.execution_scheme_id}.name",
                        "Исполнительная схема",
                    )
                workbook.set_cell(
                    sheet,
                    mapping["attachment"],
                    f"{attachment}\n{warning_block}".strip(),
                )
        workbook.enable_full_calculation()
        workbook.save(output)
        draft_output = output.with_name(f"ЧЕРНОВИК - {output.name}")
        output.replace(draft_output)
        return draft_output, contract


def workbook_snapshot(path: Path) -> dict:
    workbook = openpyxl.load_workbook(path, read_only=False, data_only=False, keep_links=True)
    formulas: dict[str, str] = {}
    errors: list[str] = []
    sheets: list[dict] = []
    for sheet in workbook.worksheets:
        sheet_errors: list[str] = []
        for row in sheet.iter_rows():
            for cell in row:
                if cell.data_type == "f":
                    formulas[f"{sheet.title}!{cell.coordinate}"] = str(cell.value)
                if cell.value and any(token in str(cell.value) for token in ("#REF!", "#VALUE!", "#N/A", "#DIV/0!")):
                    marker = f"{sheet.title}!{cell.coordinate}={cell.value}"
                    errors.append(marker)
                    sheet_errors.append(marker)
        sheets.append(
            {
                "name": sheet.title,
                "state": sheet.sheet_state,
                "print_area": str(sheet.print_area),
                "page_orientation": sheet.page_setup.orientation,
                "paper_size": sheet.page_setup.paperSize,
                "merged": sorted(str(item) for item in sheet.merged_cells.ranges),
                "styles": {
                    cell.coordinate: cell.style_id
                    for row in sheet.iter_rows()
                    for cell in row
                    if cell.has_style
                },
                "errors": sheet_errors,
            }
        )
    defined_names = {name: str(value) for name, value in workbook.defined_names.items()}
    snapshot = {
        "sha256": sha256(path),
        "sheets": sheets,
        "formulas": formulas,
        "external_links": len(getattr(workbook, "_external_links", [])),
        "defined_names": defined_names,
        "defined_name_errors": [
            f"{name}={value}"
            for name, value in defined_names.items()
            if any(token in value for token in ("#REF!", "#VALUE!", "#N/A", "#DIV/0!"))
        ],
    }
    workbook.close()
    return snapshot
