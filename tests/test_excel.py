from pathlib import Path

import openpyxl
import pytest

from executive_docs.domain import (
    Claim,
    ClaimStatus,
    DocumentPlan,
    FieldValue,
    Material,
    NeedInputQuestion,
    WorkItem,
)
from executive_docs.excel import ExcelGenerator, OOXMLWorkbook, sha256, workbook_snapshot
from executive_docs.validation import validate_workbook


ROOT = Path(__file__).resolve().parents[1]


def generator(tmp_path: Path) -> ExcelGenerator:
    return ExcelGenerator(
        ROOT,
        ROOT / "templates" / "contracts",
        ROOT / "templates" / "approved",
    )


def test_ooxml_generator_changes_only_selected_contract_fields(tmp_path: Path) -> None:
    item = WorkItem(
        id="work-kl04-1",
        family="kl_04",
        work_type="Устройство трубопровода КЛ-0,4 кВ",
        sequence_index=1,
        actual_start="01.06.2026",
        actual_end="02.06.2026",
        volume="18",
        unit="м",
        change_state="NO",
        source_claim_keys=["actual.start", "actual.end"],
    )
    plan = DocumentPlan(
        template_id="aosr_kl_04",
        selected_sheets=["АОСР-3"],
        work_item_ids=[item.id],
        first_number=17,
        output_filename="АОСР КЛ-0,4кВ.xlsx",
    )
    claims = [
        Claim(key="actual.start", raw_value="01.06.2026", normalized_value="01.06.2026", source_kind="human_answer", locator="question:q1", evidence_fragment="confirmed", status=ClaimStatus.HUMAN_CONFIRMED),
        Claim(key="actual.end", raw_value="02.06.2026", normalized_value="02.06.2026", source_kind="human_answer", locator="question:q2", evidence_fragment="confirmed", status=ClaimStatus.HUMAN_CONFIRMED),
        Claim(key="project.code", raw_value="P-42", normalized_value="P-42", source_kind="project", locator="page:1", evidence_fragment="P-42", status=ClaimStatus.OBSERVED),
    ]
    output, contract = generator(tmp_path).generate(plan, [item], claims, tmp_path)
    workbook = openpyxl.load_workbook(output, data_only=False, keep_links=True)
    assert workbook["АОСР-3"]["C32"].value == 17
    assert workbook["АОСР-3"]["A62"].value == item.work_type
    assert workbook["АОСР-3"]["R62"].value == "18"
    assert workbook["Данные объект"]["B43"].value == "P-42"
    assert workbook["Данные объект"]["B11"].value is None
    assert workbook["АОСР-3"]["A69"].value is None
    assert workbook["АОСР-3"]["O69"].value is None
    assert workbook["АОСР-3"].sheet_state == "visible"
    assert workbook["АОСР-1"].sheet_state == "hidden"
    assert workbook["АОСР-пожар"].sheet_state == "hidden"
    workbook.close()
    render_copy = tmp_path / "render-copy.xlsx"
    render_package = OOXMLWorkbook(output)
    render_package.set_only_visible("АОСР-3")
    render_package.configure_printing("АОСР-3")
    render_package.save(render_copy)
    rendered = openpyxl.load_workbook(render_copy, data_only=False, keep_links=True)
    assert rendered.active.title == "АОСР-3"
    assert str(rendered["АОСР-3"].print_area) != ""
    assert str(rendered["АОСР-1"].print_area) == ""
    assert rendered["АОСР-3"].page_setup.fitToWidth == 1
    rendered.close()
    issues = validate_workbook(output, generator(tmp_path).template_path(contract), contract, plan)
    assert not any(issue.code in {"SHEET_STRUCTURE", "FORMULA_CHANGED", "MERGES_CHANGED", "STYLES_CHANGED"} for issue in issues)
    assert sha256(ROOT / contract.source_template) == "0a9e8934639ab411350863346e9f2103371ae95f2a287c83f674b70ef2336acd"


def test_generator_rejects_model_controlled_output_path(tmp_path: Path) -> None:
    item = WorkItem(id="w", family="kl_04", work_type="Работа", sequence_index=1)
    plan = DocumentPlan(
        template_id="aosr_kl_04",
        selected_sheets=["АОСР-3"],
        work_item_ids=["w"],
        first_number=1,
        output_filename="../escape.xlsx",
    )
    with pytest.raises(ValueError, match="Имя выходного файла"):
        generator(tmp_path).generate(plan, [item], [], tmp_path)


def test_draft_generator_marks_missing_values_inside_xlsx(tmp_path: Path) -> None:
    item = WorkItem(
        id="work-kl04-1",
        family="kl_04",
        work_type="Прокладка кабеля КЛ-0,4 кВ",
        sequence_index=1,
        volume="3 м по проекту",
        unit="м",
        materials=[Material(name="Кабель АВБШв 4х95")],
    )
    plan = DocumentPlan(
        template_id="aosr_kl_04",
        selected_sheets=["АОСР-3"],
        work_item_ids=[item.id],
        first_number=4,
        output_filename="АОСР КЛ-0,4кВ.xlsx",
    )
    claims = [
        Claim(
            key="project.code",
            raw_value="P-42",
            normalized_value="P-42",
            source_kind="project",
            source_file_id="project",
            locator="page:1",
            evidence_fragment="P-42",
            status=ClaimStatus.OBSERVED,
        ),
        Claim(
            key="contractor.name",
            raw_value="ООО Подрядчик",
            normalized_value="ООО Подрядчик",
            source_kind="project",
            source_file_id="project",
            locator="page:2",
            evidence_fragment="ООО Подрядчик",
            status=ClaimStatus.OBSERVED,
        ),
    ]
    questions = [
        NeedInputQuestion(id="q-dates", field_key="actual.dates", prompt="Даты?", reason="Нет дат"),
        NeedInputQuestion(id="q-volume", field_key="actual.volume", prompt="Объём?", reason="Нет факта"),
        NeedInputQuestion(
            id="q-materials",
            field_key="materials.quality_documents",
            prompt="Паспорта?",
            reason="Нет паспортов",
        ),
        NeedInputQuestion(id="q-changes", field_key="changes.state", prompt="Изменения?", reason="Неизвестно"),
        NeedInputQuestion(
            id="q-profile",
            field_key="customer.profile_confirmation",
            prompt="Профиль?",
            reason="Не утверждён",
        ),
    ]

    output, contract = generator(tmp_path).generate_draft(
        plan,
        [item],
        claims,
        questions,
        tmp_path,
    )

    assert output.name == "ЧЕРНОВИК - АОСР КЛ-0,4кВ.xlsx"
    workbook = openpyxl.load_workbook(output, data_only=False, keep_links=True)
    assert workbook["Данные объект"]["B43"].value == "P-42"
    assert workbook["Данные объект"]["B11"].value == "ЧЕРНОВИК, НЕ ПОДТВЕРЖДЕНО: ООО Подрядчик"
    assert workbook["АОСР-3"]["P77"].value == "НЕ ПОДТВЕРЖДЕНО"
    assert workbook["АОСР-3"]["P78"].value == "НЕ ПОДТВЕРЖДЕНО"
    assert workbook["АОСР-3"]["R62"].value == "НЕ ПОДТВЕРЖДЕНО (по проекту: 3 м по проекту)"
    assert workbook["АОСР-3"]["O69"].value == "ПАСПОРТ/СЕРТИФИКАТ НЕ УКАЗАН"
    assert "ЧЕРНОВИК. НЕ ДЛЯ ПОДПИСАНИЯ." in workbook["АОСР-3"]["G72"].value
    workbook.close()
    issues = validate_workbook(
        output,
        generator(tmp_path).template_path(contract),
        contract,
        plan,
        claims,
    )
    assert not [issue for issue in issues if issue.severity == "error"]


def test_generator_rejects_field_value_without_matching_claim(tmp_path: Path) -> None:
    item = WorkItem(id="w", family="kl_04", work_type="Работа", sequence_index=1)
    plan = DocumentPlan(
        template_id="aosr_kl_04",
        selected_sheets=["АОСР-3"],
        work_item_ids=["w"],
        first_number=1,
        field_values=[FieldValue(key="project.code", value="INVENTED")],
        output_filename="АОСР КЛ-0,4кВ.xlsx",
    )
    with pytest.raises(ValueError, match="не подтверждено Claim"):
        generator(tmp_path).generate(plan, [item], [], tmp_path)


def test_generator_writes_multiple_material_rows_and_subsequent_work(tmp_path: Path) -> None:
    vrs_item = WorkItem(
        id="vrs-2",
        family="vrs",
        work_type="Монтаж постамента ВРЩ",
        sequence_index=1,
        materials=[
            Material(
                name="Уголок - 63х63х5",
                quantity="10.64",
                unit="м",
                quality_document="Сертификат №1",
            ),
            Material(
                name="Профлист",
                quantity="1.12",
                unit="м2",
                quality_document="Сертификат №2",
            ),
            Material(
                name="Труба профильная 80х80х3",
                quantity="2.75",
                unit="м",
                quality_document="Сертификат №3",
            ),
        ],
    )
    vrs_plan = DocumentPlan(
        template_id="aosr_vrs",
        selected_sheets=["АОСР-2"],
        work_item_ids=[vrs_item.id],
        first_number=1,
        output_filename="АОСР ВРЩ.xlsx",
    )
    vrs_output, _ = generator(tmp_path).generate(vrs_plan, [vrs_item], [], tmp_path)
    workbook = openpyxl.load_workbook(vrs_output, data_only=False, keep_links=True)
    sheet = workbook["АОСР-2"]
    assert [sheet[f"A{row}"].value for row in (69, 70, 71)] == [
        "Уголок - 63х63х5",
        "Профлист",
        "Труба профильная 80х80х3",
    ]
    assert [sheet[f"I{row}"].value for row in (69, 70, 71)] == ["-", "-", "-"]
    assert [sheet[f"J{row}"].value for row in (69, 70, 71)] == ["10.64", "1.12", "2.75"]
    assert [sheet[f"L{row}"].value for row in (69, 70, 71)] == ["м", "м2", "м"]
    assert [sheet[f"N{row}"].value for row in (69, 70, 71)] == [
        "Сертификат №1",
        "Сертификат №2",
        "Сертификат №3",
    ]
    workbook.close()

    kl_item = WorkItem(
        id="kl04-4",
        family="kl_04",
        work_type="Прокладка кабеля",
        sequence_index=1,
        subsequent_work="Пусконаладочные работы",
    )
    kl_plan = DocumentPlan(
        template_id="aosr_kl_04",
        selected_sheets=["АОСР-4"],
        work_item_ids=[kl_item.id],
        first_number=1,
        output_filename="АОСР КЛ-0,4кВ.xlsx",
    )
    kl_output, _ = generator(tmp_path).generate(kl_plan, [kl_item], [], tmp_path)
    workbook = openpyxl.load_workbook(kl_output, data_only=False, keep_links=True)
    assert workbook["АОСР-4"]["A86"].value == "Пусконаладочные работы"
    workbook.close()


def test_generator_rejects_more_materials_than_template_rows(tmp_path: Path) -> None:
    item = WorkItem(
        id="vrs-6",
        family="vrs",
        work_type="Работа",
        sequence_index=1,
        materials=[Material(name=f"Материал {index}") for index in range(3)],
    )
    plan = DocumentPlan(
        template_id="aosr_vrs",
        selected_sheets=["АОСР-6"],
        work_item_ids=[item.id],
        first_number=1,
        output_filename="АОСР ВРЩ.xlsx",
    )
    with pytest.raises(ValueError, match="предусмотрено материалов"):
        generator(tmp_path).generate(plan, [item], [], tmp_path)


@pytest.mark.parametrize(
    ("source_name", "candidate_name"),
    [
        ("9. АОСР КЛ.xlsx", "9. АОСР КЛ.xlsx"),
        ("10. АОСР ВРЩ.xlsx", "10. АОСР ВРЩ.xlsx"),
    ],
)
def test_cleaned_template_candidates_preserve_structure_and_printing(
    source_name: str, candidate_name: str
) -> None:
    baseline = workbook_snapshot(ROOT / "template" / source_name)
    candidate = workbook_snapshot(ROOT / "templates" / "approved" / candidate_name)
    assert [item["name"] for item in candidate["sheets"]] == [item["name"] for item in baseline["sheets"]]
    assert [
        (item["name"], item["print_area"], item["page_orientation"], item["paper_size"])
        for item in candidate["sheets"]
    ] == [
        (item["name"], item["print_area"], item["page_orientation"], item["paper_size"])
        for item in baseline["sheets"]
    ]
    assert candidate["external_links"] == 0
    assert not candidate["defined_name_errors"]
    assert not [error for item in candidate["sheets"] for error in item["errors"]]


def test_ooxml_clear_fill_colors_preserves_formula_and_print_structure(tmp_path: Path) -> None:
    source = ROOT / "template" / "9. АОСР КЛ.xlsx"
    destination = tmp_path / source.name
    before = workbook_snapshot(source)
    package = OOXMLWorkbook(source)
    assert package.clear_fill_colors() > 0
    package.save(destination)
    after = workbook_snapshot(destination)
    assert before["formulas"] == after["formulas"]
    assert before["defined_names"] == after["defined_names"]
    assert [item["name"] for item in before["sheets"]] == [item["name"] for item in after["sheets"]]
    assert [
        (item["name"], item["print_area"], item["page_orientation"], item["paper_size"], item["merged"])
        for item in before["sheets"]
    ] == [
        (item["name"], item["print_area"], item["page_orientation"], item["paper_size"], item["merged"])
        for item in after["sheets"]
    ]
