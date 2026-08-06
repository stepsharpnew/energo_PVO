from __future__ import annotations

import io
import re
import zipfile
from pathlib import Path

import openpyxl
import pytest
import yaml
from lxml import etree
from pypdf import PdfWriter
from pypdf.generic import DecodedStreamObject, DictionaryObject, NameObject

from executive_docs.agent import OpenAIAgent
from executive_docs.config import Settings
from executive_docs.domain import (
    JobStatus,
    ProjectState,
    TemplateCellAssignment,
    TemplateFillAnalysis,
    TemplateUnresolvedFinding,
)
from executive_docs.excel import MAIN_NS, OOXMLWorkbook, sha256
from executive_docs.ingestion import build_inventory
from executive_docs.pipeline import Pipeline
from executive_docs.repository import Repository
from executive_docs.selected_templates import (
    SelectedTemplateGenerator,
    TemplateCatalog,
    validate_selected_template_output,
)
from executive_docs.storage import Storage


ROOT = Path(__file__).resolve().parents[1]
JOB_ID = "55555555-5555-5555-5555-555555555555"


def build_catalog(tmp_path: Path) -> TemplateCatalog:
    approved = tmp_path / "templates" / "approved"
    contracts = tmp_path / "templates" / "fill-contracts"
    approved.mkdir(parents=True)
    contracts.mkdir(parents=True)
    candidate = approved / "sample.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Данные"
    sheet["A1"] = "Шифр проекта"
    sheet["B1"] = None
    sheet["A2"] = "Фактическая дата"
    sheet["B2"] = None
    sheet["C1"] = "=B1"
    support = workbook.create_sheet("Служебный")
    support["A1"] = "служебная структура"
    support.sheet_state = "hidden"
    workbook.save(candidate)
    workbook.close()
    contract = {
        "template_id": "sample",
        "display_name": "Тестовый шаблон",
        "document_kind": "test",
        "version": "1-candidate",
        "status": "READY_FOR_VISUAL_APPROVAL",
        "approved": False,
        "candidate_template": "templates/approved/sample.xlsx",
        "candidate_sha256": sha256(candidate),
        "source_sha256": "0" * 64,
        "etalon_sha256": "1" * 64,
        "output_filename": "sample.xlsx",
        "warning_fill_rgb": "FFFFE699",
        "structural_findings": {"formula_differences": 0},
        "fields": [
            {
                "sheet": "Данные",
                "cell": "B1",
                "label": "Шифр проекта",
                "value_kind": "text",
                "required": True,
                "manual_reason": None,
            },
            {
                "sheet": "Данные",
                "cell": "B2",
                "label": "Фактическая дата",
                "value_kind": "date",
                "required": True,
                "manual_reason": "Требуется подтверждение специалистом",
            },
        ],
    }
    (contracts / "sample.yaml").write_text(
        yaml.safe_dump(contract, allow_unicode=True, sort_keys=False),
        encoding="utf-8",
    )
    return TemplateCatalog(tmp_path, contracts, approved)


def test_selected_template_generator_fills_only_registered_cells_and_highlights_missing(
    tmp_path: Path,
) -> None:
    catalog = build_catalog(tmp_path)
    contract = catalog.get("sample")
    assignment = TemplateCellAssignment(
        sheet="Данные",
        cell="B1",
        value="P-42",
        source_file_id="project",
        locator="page:1",
        evidence_fragment="Шифр проекта P-42",
    )

    output, unresolved = SelectedTemplateGenerator(catalog).generate(
        contract,
        [assignment],
        tmp_path / "output",
    )

    assert output.name == "ЧЕРНОВИК - sample.xlsx"
    assert [(item.sheet, item.cell) for item in unresolved] == [("Данные", "B2")]
    workbook = openpyxl.load_workbook(output, data_only=False)
    assert workbook["Данные"]["B1"].value == "P-42"
    assert workbook["Данные"]["B2"].value is None
    assert workbook["Данные"]["B2"].fill.fill_type == "solid"
    assert workbook["Данные"]["B2"].fill.fgColor.rgb == "FFFFE699"
    assert workbook["Данные"]["C1"].value == "=B1"
    workbook.close()
    issues = validate_selected_template_output(
        output,
        catalog.candidate_path(contract),
        contract,
        [assignment],
        unresolved,
    )
    assert not [issue for issue in issues if issue.severity == "error"]


def test_selected_template_generator_rejects_manual_or_unknown_cells(tmp_path: Path) -> None:
    catalog = build_catalog(tmp_path)
    contract = catalog.get("sample")
    generator = SelectedTemplateGenerator(catalog)
    manual = TemplateCellAssignment(
        sheet="Данные",
        cell="B2",
        value="01.07.2026",
        source_file_id="project",
        locator="page:1",
        evidence_fragment="01.07.2026",
    )
    with pytest.raises(ValueError, match="ручного подтверждения"):
        generator.generate(contract, [manual], tmp_path / "manual")
    unknown = manual.model_copy(update={"cell": "Z99", "value": "X"})
    with pytest.raises(ValueError, match="незарегистрированную"):
        generator.generate(contract, [unknown], tmp_path / "unknown")


def test_selected_template_generator_writes_declared_numbers_as_numbers(
    tmp_path: Path,
) -> None:
    catalog = build_catalog(tmp_path)
    contract_path = tmp_path / "templates" / "fill-contracts" / "sample.yaml"
    data = yaml.safe_load(contract_path.read_text(encoding="utf-8"))
    data["fields"][0]["value_kind"] = "number"
    contract_path.write_text(
        yaml.safe_dump(data, allow_unicode=True, sort_keys=False),
        encoding="utf-8",
    )
    catalog.reload()
    contract = catalog.get("sample")
    assignment = TemplateCellAssignment(
        sheet="Данные",
        cell="B1",
        value="12,5",
        source_file_id="project",
        locator="page:1",
        evidence_fragment="Количество 12,5",
    )

    output, unresolved = SelectedTemplateGenerator(catalog).generate(
        contract,
        [assignment],
        tmp_path / "numeric",
    )

    workbook = openpyxl.load_workbook(output, data_only=False)
    try:
        assert workbook["Данные"]["B1"].value == 12.5
        assert workbook["Данные"]["B1"].data_type == "n"
    finally:
        workbook.close()
    issues = validate_selected_template_output(
        output,
        catalog.candidate_path(contract),
        contract,
        [assignment],
        unresolved,
    )
    assert not [issue for issue in issues if issue.severity == "error"]


def test_selected_template_catalog_rejects_candidate_changed_after_load(
    tmp_path: Path,
) -> None:
    catalog = build_catalog(tmp_path)
    contract = catalog.get("sample")
    candidate = catalog.candidate_path(contract)
    workbook = openpyxl.load_workbook(candidate)
    workbook["Данные"]["D9"] = "чужое значение"
    workbook.save(candidate)
    workbook.close()

    with pytest.raises(ValueError, match="изменился"):
        catalog.get("sample")


def test_selected_template_generator_rejects_contract_changed_after_load(
    tmp_path: Path,
) -> None:
    catalog = build_catalog(tmp_path)
    contract = catalog.get("sample")
    contract.path.write_text(
        contract.path.read_text(encoding="utf-8") + "\n# hot change\n",
        encoding="utf-8",
    )

    with pytest.raises(ValueError, match="Контракт шаблона изменился"):
        SelectedTemplateGenerator(catalog).generate(
            contract,
            [],
            tmp_path / "output",
        )


def test_unresolved_conflict_keeps_template_and_pdf_provenance(
    tmp_path: Path,
) -> None:
    catalog = build_catalog(tmp_path)
    contract = catalog.get("sample")
    finding = TemplateUnresolvedFinding(
        sheet="Данные",
        cell="B1",
        category="conflict",
        reason="На страницах указаны два разных шифра",
        source_locators=["page:1", "page:2"],
        source_values=["P-42", "P-77"],
        evidence_fragments=["Шифр P-42", "Шифр P-77"],
    )

    unresolved = SelectedTemplateGenerator.unresolved_cells(
        contract,
        [],
        [finding],
    )

    item = next(cell for cell in unresolved if cell.cell == "B1")
    assert item.template_id == "sample"
    assert item.category == "conflict"
    assert item.blocking is True
    assert item.source_locators == ["page:1", "page:2"]
    assert "два разных" in item.reason
    assert (
        TemplateUnresolvedFinding(
            sheet="Данные",
            cell="B1",
            category="unapproved_rule",
            reason="Правило переноса ещё не утверждено специалистом",
        ).category
        == "unapproved_rule"
    )


def test_template_assignment_must_match_text_on_the_cited_pdf_page(tmp_path: Path) -> None:
    catalog = build_catalog(tmp_path)
    contract = catalog.get("sample")
    settings = Settings(
        root=tmp_path,
        data_dir=tmp_path / "data",
        runs_dir=tmp_path / "data" / "runs",
        db_path=tmp_path / "data" / "app.db",
    )
    settings.ensure_directories()
    storage = Storage(settings)
    storage.initialize_job(JOB_ID)
    pdf = io.BytesIO()
    writer = PdfWriter()
    page = writer.add_blank_page(width=595, height=842)
    font = writer._add_object(
        DictionaryObject(
            {
                NameObject("/Type"): NameObject("/Font"),
                NameObject("/Subtype"): NameObject("/Type1"),
                NameObject("/BaseFont"): NameObject("/Helvetica"),
            }
        )
    )
    page[NameObject("/Resources")] = DictionaryObject(
        {
            NameObject("/Font"): DictionaryObject(
                {NameObject("/F1"): font}
            )
        }
    )
    content = DecodedStreamObject()
    content.set_data(
        b"BT /F1 12 Tf 72 720 Td "
        b"(Project code P-42. Alternative P-77) Tj ET"
    )
    page[NameObject("/Contents")] = writer._add_object(content)
    writer.add_blank_page(width=595, height=842)
    writer.write(pdf)
    pdf.seek(0)
    artifact = storage.save_upload(
        JOB_ID,
        "project.pdf",
        pdf,
        "application/pdf",
    )
    state = ProjectState(
        job_id=JOB_ID,
        operator_name="Специалист",
        flow_version="selected-template-v2",
        selected_template_id=contract.template_id,
        artifacts=[artifact],
    )
    state.artifacts, _ = build_inventory(storage.job_dir(JOB_ID), state.artifacts)
    valid = TemplateFillAnalysis(
        summary="ok",
        assignments=[
            TemplateCellAssignment(
                sheet="Данные",
                cell="B1",
                value="P-42",
                source_file_id=artifact.id,
                locator="page:1",
                evidence_fragment="Project code P-42",
            )
        ],
    )
    invented = valid.model_copy(
        update={
            "assignments": [
                valid.assignments[0].model_copy(
                    update={
                        "value": "P-99",
                        "evidence_fragment": "Project code P-99",
                    }
                )
            ]
        }
    )
    truncated = valid.model_copy(
        update={
            "assignments": [
                valid.assignments[0].model_copy(
                    update={
                        "value": "P-4",
                        "evidence_fragment": "Project code P-42",
                    }
                )
            ]
        }
    )
    visual_only = valid.model_copy(
        update={
            "assignments": [
                valid.assignments[0].model_copy(
                    update={
                        "value": "P-77",
                        "locator": "page:2",
                        "evidence_fragment": "Image code P-77",
                    }
                )
            ]
        }
    )

    assert (
        OpenAIAgent._template_fill_rejection(
            state,
            contract,
            valid,
            storage.job_dir(JOB_ID),
        )
        is None
    )
    assert "not found" in (
        OpenAIAgent._template_fill_rejection(
            state,
            contract,
            invented,
            storage.job_dir(JOB_ID),
        )
        or ""
    )
    assert "value is not present" in (
        OpenAIAgent._template_fill_rejection(
            state,
            contract,
            truncated,
            storage.job_dir(JOB_ID),
        )
        or ""
    )
    assert (
        OpenAIAgent._template_fill_rejection(
            state,
            contract,
            visual_only,
            storage.job_dir(JOB_ID),
        )
        is None
    )
    conflict = TemplateFillAnalysis(
        summary="conflict",
        unresolved=[
            TemplateUnresolvedFinding(
                sheet="Данные",
                cell="B1",
                category="conflict",
                reason="На странице указаны P-42 и P-77",
                source_locators=["page:1", "page:1"],
                source_values=["P-42", "P-77"],
                evidence_fragments=[
                    "Project code P-42",
                    "Alternative P-77",
                ],
            )
        ],
    )
    assert (
        OpenAIAgent._template_fill_rejection(
            state,
            contract,
            conflict,
            storage.job_dir(JOB_ID),
        )
        is None
    )
    conflict_claims = Pipeline._selected_template_claims(
        contract,
        artifact,
        conflict,
    )
    assert [claim.status for claim in conflict_claims] == [
        "conflict",
        "conflict",
    ]
    assert [claim.normalized_value for claim in conflict_claims] == [
        "P-42",
        "P-77",
    ]
    assert [claim.locator for claim in conflict_claims] == [
        "page:1",
        "page:1",
    ]
    empty_conflict_value = conflict.model_copy(
        update={
            "unresolved": [
                conflict.unresolved[0].model_copy(
                    update={"source_values": ["", "P-77"]}
                )
            ]
        }
    )
    assert "empty unresolved source value" in (
        OpenAIAgent._template_fill_rejection(
            state,
            contract,
            empty_conflict_value,
            storage.job_dir(JOB_ID),
        )
        or ""
    )
    incomplete = TemplateFillAnalysis(summary="incomplete")
    assert "Every writable template cell" in (
        OpenAIAgent._template_fill_rejection(
            state,
            contract,
            incomplete,
            storage.job_dir(JOB_ID),
        )
        or ""
    )


def test_invalid_missing_finding_returns_model_correction_without_pydantic_leak() -> None:
    candidate, error = OpenAIAgent._validate_template_fill_call(
        {
            "summary": "Нужна проверка",
            "assignments": [],
            "unresolved": [
                {
                    "sheet": "Акт опрессовки",
                    "cell": "B9",
                    "category": "missing_from_pdf",
                    "reason": "Не подтверждено",
                    "source_locators": ["page:3"],
                    "source_values": ["12.07.2026"],
                    "evidence_fragments": ["Дата установки 12.07.2026"],
                }
            ],
        }
    )

    assert candidate is None
    assert error is not None
    assert "call submit_template_fill again" in error
    assert "MUST all be empty" in error
    assert "unresolved.0" in error
    assert "errors.pydantic.dev" not in error
    assert "input_value" not in error


def test_selected_template_pipeline_creates_one_draft_in_one_pass(tmp_path: Path) -> None:
    catalog = build_catalog(tmp_path)
    contract = catalog.get("sample")
    settings = Settings(
        root=tmp_path,
        data_dir=tmp_path / "data",
        runs_dir=tmp_path / "data" / "runs",
        db_path=tmp_path / "data" / "app.db",
        skill_dir=ROOT / "agent-skill" / "prepare-executive-docs",
        contracts_dir=ROOT / "templates" / "contracts",
        fill_contracts_dir=tmp_path / "templates" / "fill-contracts",
        approved_templates_dir=tmp_path / "templates" / "approved",
        source_templates_dir=ROOT / "template",
        profiles_dir=ROOT / "profiles",
        agent_mode="heuristic",
        soffice_path="missing-soffice",
    )
    settings.ensure_directories()
    repository = Repository(settings.db_path)
    repository.initialize()
    storage = Storage(settings)
    storage.initialize_job(JOB_ID)
    pdf_bytes = io.BytesIO()
    writer = PdfWriter()
    writer.add_blank_page(width=595, height=842)
    writer.write(pdf_bytes)
    pdf_bytes.seek(0)
    artifact = storage.save_upload(
        JOB_ID,
        "project.pdf",
        pdf_bytes,
        "application/pdf",
    )
    state = ProjectState(
        job_id=JOB_ID,
        operator_name="Специалист",
        flow_version="selected-template-v2",
        selected_template_id=contract.template_id,
        selected_template_name=contract.display_name,
        selected_template_status=contract.status,
        selected_template_version=contract.version,
        selected_template_sha256=contract.candidate_sha256,
        selected_template_contract_sha256=contract.contract_sha256,
        status=JobStatus.FILES_UPLOADED,
        artifacts=[artifact],
    )
    repository.create(state)

    Pipeline(settings, repository, storage).process(JOB_ID)

    result = repository.get(JOB_ID)
    assert result is not None
    assert result.status == JobStatus.NEEDS_INPUT
    assert result.template_analysis_complete is True
    assert result.draft_report_ready is True
    assert len(result.draft_excel_files) == 1
    assert len(result.unresolved_template_cells) == 2
    assert (
        storage.job_dir(JOB_ID) / result.draft_excel_files[0]
    ).is_file()
    snapshot = (
        storage.job_dir(JOB_ID)
        / "output"
        / "selected"
        / "r1"
        / "source"
        / "template.xlsx"
    )
    assert snapshot.is_file()
    assert sha256(snapshot) == contract.candidate_sha256


def test_selected_template_pipeline_fails_closed_without_all_pins(
    tmp_path: Path,
) -> None:
    catalog = build_catalog(tmp_path)
    contract = catalog.get("sample")
    settings = Settings(
        root=tmp_path,
        data_dir=tmp_path / "data",
        runs_dir=tmp_path / "data" / "runs",
        db_path=tmp_path / "data" / "app.db",
        skill_dir=ROOT / "agent-skill" / "prepare-executive-docs",
        contracts_dir=ROOT / "templates" / "contracts",
        fill_contracts_dir=tmp_path / "templates" / "fill-contracts",
        approved_templates_dir=tmp_path / "templates" / "approved",
        source_templates_dir=ROOT / "template",
        profiles_dir=ROOT / "profiles",
        agent_mode="heuristic",
        soffice_path="missing-soffice",
    )
    settings.ensure_directories()
    repository = Repository(settings.db_path)
    repository.initialize()
    storage = Storage(settings)
    storage.initialize_job(JOB_ID)
    state = ProjectState(
        job_id=JOB_ID,
        operator_name="Специалист",
        flow_version="selected-template-v2",
        selected_template_id=contract.template_id,
        selected_template_version=contract.version,
        selected_template_sha256=contract.candidate_sha256,
        selected_template_contract_sha256=None,
        status=JobStatus.FILES_UPLOADED,
    )
    repository.create(state)

    Pipeline(settings, repository, storage).process(JOB_ID)

    result = repository.get(JOB_ID)
    assert result is not None
    assert result.status == JobStatus.FAILED_ANALYSIS
    assert "SHA" in (result.error or "")
    assert result.draft_excel_files == []


def test_real_candidates_are_discovery_only_and_do_not_retain_known_project_values() -> None:
    catalog = TemplateCatalog(
        ROOT,
        ROOT / "templates" / "fill-contracts",
        ROOT / "templates" / "approved",
    )
    forbidden = (
        "Шатковский",
        "Чернявская",
        "Алексанян",
        "25.09.2024",
        "15.01.2025",
        "26.05.2025",
    )
    assert {item.template_id for item in catalog.list()} == {
        "emr",
        "protocols",
        "ojr",
        "avk",
        "aosr_vl",
    }
    for contract in catalog.list():
        assert contract.status == "DISCOVERY_REVIEW_REQUIRED"
        assert contract.approved is False
        expected_derivation = (
            "source_structure_plus_reviewed_aosr_overrides"
            if contract.template_id == "aosr_vl"
            else "source_only_discovery"
        )
        assert contract.structural_findings["target_derivation"] == expected_derivation
        assert contract.structural_findings["remaining_sensitive_value_count"] == 0
        assert contract.structural_findings["candidate_external_links"] == 0
        assert contract.structural_findings["package_forbidden_token_count"] == 0
        contract_text = contract.path.read_text(encoding="utf-8").casefold()
        assert not any(
            token in contract_text
            for token in (
                "гефест",
                "энергосистем",
                "солнечногорск",
                "высоцк",
                "алексанян",
                "чернявск",
                "шатковск",
                "трушин",
                "бараночников",
            )
        )
        with zipfile.ZipFile(catalog.candidate_path(contract)) as archive:
            names = archive.namelist()
            assert not any(name.startswith("xl/externalLinks/") for name in names)
            assert not any(name.startswith("customXml/") for name in names)
            core = archive.read("docProps/core.xml").decode(
                "utf-8",
                errors="ignore",
            )
            assert "Executive Docs" in core
            assert not any(
                token in core.casefold()
                for token in (
                    "алексанян",
                    "чернявск",
                    "шатковск",
                    "elena camarillo",
                )
            )
        workbook = openpyxl.load_workbook(
            catalog.candidate_path(contract),
            data_only=False,
            keep_links=True,
        )
        try:
            values = "\n".join(
                str(cell.value)
                for worksheet in workbook.worksheets
                for row in worksheet.iter_rows()
                for cell in row
                if cell.value not in (None, "") and cell.data_type != "f"
            )
            assert not any(token in values for token in forbidden)
            for field in contract.fields:
                cell = workbook[field.sheet][field.cell]
                assert cell.value in (None, "")
                for merged in workbook[field.sheet].merged_cells.ranges:
                    if field.cell in merged:
                        assert field.cell == merged.start_cell.coordinate
        finally:
            workbook.close()

    emr = catalog.get("emr")
    emr_workbook = openpyxl.load_workbook(
        catalog.candidate_path(emr),
        data_only=False,
        keep_links=True,
    )
    try:
        assert emr_workbook["Главная "]["A16"].value.startswith(
            "ФОРМЫ ДОКУМЕНТОВ"
        )
        assert ("Главная ", "A16") not in emr.field_map
        assert emr_workbook["связ.данные (Работы)"]["A2"].value == "Монтаж ТП"
        assert ("связ.данные (Работы)", "A2") not in emr.field_map
        assert (
            emr_workbook["связ.данные(оборудование)"]["A2"].value
            == "Подстанция трансформаторная"
        )
    finally:
        emr_workbook.close()

    protocols = catalog.get("protocols")
    protocol_workbook = openpyxl.load_workbook(
        catalog.candidate_path(protocols),
        data_only=False,
        keep_links=True,
    )
    try:
        for coordinate in (
            ("прот.№4", "B18"),
            ("Протокол _ВНА", "B12"),
        ):
            assert protocol_workbook[coordinate[0]][coordinate[1]].value is None
            assert protocols.field_map[coordinate].manual_reason
        assert "Если установка работает" in (
            protocol_workbook["Протокол _ВНА"]["L18"].value or ""
        )
        assert ("Протокол _ВНА", "L18") not in protocols.field_map
    finally:
        protocol_workbook.close()

    avk = catalog.get("avk")
    avk_workbook = openpyxl.load_workbook(
        catalog.candidate_path(avk),
        data_only=False,
        keep_links=True,
    )
    try:
        for coordinate in (("КТП", "A41"), ("КТП", "G45")):
            assert avk_workbook[coordinate[0]][coordinate[1]].value is None
            assert avk.field_map[coordinate].manual_reason
    finally:
        avk_workbook.close()

    aosr = catalog.get("aosr_vl")
    aosr_workbook = openpyxl.load_workbook(
        catalog.candidate_path(aosr),
        data_only=False,
        keep_links=True,
    )
    try:
        assert (
            aosr_workbook["данные строй"]["B2"].value
            == "Самонесущий изолированный повод"
        )
        assert ("данные строй", "B2") not in aosr.field_map
        assert not [
            field
            for field in aosr.fields
            if field.sheet == "Данные организации"
        ]
        assert len(aosr.model_fields()) >= 20
        assert aosr.field_map[("Данные объект", "B2")].semantic_id == (
            "project.sap_number"
        )
        assert aosr.field_map[("Данные объект", "B43")].semantic_id == (
            "project.design_document_code"
        )
        assert aosr_workbook["АОСР-2"]["P63"].value is None
        assert aosr_workbook["АОСР-3"]["X62"].value is None
        assert aosr_workbook["АОСР-4"]["H69"].value is None
        assert aosr_workbook["АОСР-6"]["H69"].value is None
        assert aosr_workbook["АОСР-7"]["H69"].value is None
        for coordinate in (
            ("АОСР-2", "P63"),
            ("АОСР-3", "T62"),
            ("АОСР-3", "X62"),
            ("АОСР-4", "H69"),
            ("АОСР-6", "H69"),
            ("АОСР-7", "H69"),
        ):
            assert coordinate in aosr.field_map
            assert not aosr.field_map[coordinate].manual_reason
        assert aosr_workbook["АОСР-1"]["O63"].value.startswith("=IF(")
        formulas = [
            str(cell.value)
            for worksheet in aosr_workbook.worksheets
            for row in worksheet.iter_rows()
            for cell in row
            if cell.data_type == "f"
        ]
        assert not any("#REF!" in formula.upper() for formula in formulas)
        assert not any(re.search(r"\[[1-9][0-9]*\]", formula) for formula in formulas)
    finally:
        aosr_workbook.close()

    assert aosr.structural_findings["formula_errors_observed"] == 0
    assert aosr.structural_findings["raw_ref_error_count"] == 0
    assert aosr.structural_findings["remaining_external_formula_reference_count"] == 0
    assert aosr.structural_findings["unsafe_blank_formula_count"] == 0
    with pytest.raises(ValueError, match="смыслу поля"):
        SelectedTemplateGenerator._typed_value(
            aosr.field_map[("Данные объект", "B2")],
            "5557-354783-68-03/26",
        )


def test_real_ojr_draft_changes_only_warning_fills(tmp_path: Path) -> None:
    catalog = TemplateCatalog(
        ROOT,
        ROOT / "templates" / "fill-contracts",
        ROOT / "templates" / "approved",
    )
    contract = catalog.get("ojr")
    output, unresolved = SelectedTemplateGenerator(catalog).generate(
        contract,
        [],
        tmp_path,
    )

    issues = validate_selected_template_output(
        output,
        catalog.candidate_path(contract),
        contract,
        [],
        unresolved,
    )

    assert not [
        issue
        for issue in issues
        if issue.code
        in {
            "UNAUTHORIZED_STYLE_CHANGE",
            "UNAUTHORIZED_CELL_CHANGE",
            "UNRESOLVED_CELL_NOT_EMPTY",
            "UNRESOLVED_CELL_NOT_HIGHLIGHTED",
        }
    ]
    assert not [issue for issue in issues if issue.severity == "error"]


def test_real_aosr_vl_candidate_is_technically_clean(tmp_path: Path) -> None:
    catalog = TemplateCatalog(
        ROOT,
        ROOT / "templates" / "fill-contracts",
        ROOT / "templates" / "approved",
    )
    contract = catalog.get("aosr_vl")
    output, unresolved = SelectedTemplateGenerator(catalog).generate(
        contract,
        [],
        tmp_path,
    )

    issues = validate_selected_template_output(
        output,
        catalog.candidate_path(contract),
        contract,
        [],
        unresolved,
    )

    assert not [issue for issue in issues if issue.severity == "error"]
    assert {issue.code for issue in issues} == {"TEMPLATE_NOT_APPROVED"}


def test_selected_validator_rejects_visibility_and_row_structure_changes(
    tmp_path: Path,
) -> None:
    catalog = build_catalog(tmp_path)
    contract = catalog.get("sample")
    output, unresolved = SelectedTemplateGenerator(catalog).generate(
        contract,
        [],
        tmp_path / "generated",
    )

    changed_visibility = tmp_path / "changed-visibility.xlsx"
    package = OOXMLWorkbook(output)
    package.set_visibility(
        {"Данные", "Служебный"},
        {"Данные", "Служебный"},
    )
    package.save(changed_visibility)
    issues = validate_selected_template_output(
        changed_visibility,
        catalog.candidate_path(contract),
        contract,
        [],
        unresolved,
    )
    assert any(
        issue.code == "SHEET_VISIBILITY_CHANGED"
        and issue.severity == "error"
        for issue in issues
    )

    changed_row = tmp_path / "changed-row.xlsx"
    package = OOXMLWorkbook(output)
    root = package._sheet_root("Данные")
    namespace = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    row = root.find(
        f"{{{namespace}}}sheetData/{{{namespace}}}row[@r='1']"
    )
    assert row is not None
    row.set("ht", "99")
    row.set("customHeight", "1")
    package._save_sheet_root("Данные", root)
    package.save(changed_row)
    issues = validate_selected_template_output(
        changed_row,
        catalog.candidate_path(contract),
        contract,
        [],
        unresolved,
    )
    assert any(
        issue.code == "WORKSHEET_STRUCTURE_CHANGED"
        and issue.severity == "error"
        for issue in issues
    )


def test_selected_validator_rejects_changed_style_definitions(
    tmp_path: Path,
) -> None:
    catalog = build_catalog(tmp_path)
    contract = catalog.get("sample")
    output, unresolved = SelectedTemplateGenerator(catalog).generate(
        contract,
        [],
        tmp_path / "generated",
    )
    changed_styles = tmp_path / "changed-styles.xlsx"
    package = OOXMLWorkbook(output)
    styles = etree.fromstring(package.parts["xl/styles.xml"])
    name = styles.find(
        f"{{{MAIN_NS}}}fonts/{{{MAIN_NS}}}font/{{{MAIN_NS}}}name"
    )
    assert name is not None
    name.set("val", "Courier New")
    package.parts["xl/styles.xml"] = etree.tostring(
        styles,
        xml_declaration=True,
        encoding="UTF-8",
        standalone=True,
    )
    package.save(changed_styles)

    issues = validate_selected_template_output(
        changed_styles,
        catalog.candidate_path(contract),
        contract,
        [],
        unresolved,
    )
    assert any(
        issue.code == "STYLE_DEFINITIONS_CHANGED"
        and issue.severity == "error"
        for issue in issues
    )


def test_selected_validator_rejects_changed_theme(
    tmp_path: Path,
) -> None:
    catalog = build_catalog(tmp_path)
    contract = catalog.get("sample")
    output, unresolved = SelectedTemplateGenerator(catalog).generate(
        contract,
        [],
        tmp_path / "generated",
    )
    changed_theme = tmp_path / "changed-theme.xlsx"
    package = OOXMLWorkbook(output)
    theme_name = "xl/theme/theme1.xml"
    original_theme = package.parts[theme_name]
    mutated_theme = original_theme.replace(b"Calibri", b"Courier New", 1)
    if mutated_theme == original_theme:
        mutated_theme = original_theme.replace(b"Cambria", b"Courier New", 1)
    assert mutated_theme != original_theme
    package.parts[theme_name] = mutated_theme
    package.save(changed_theme)

    issues = validate_selected_template_output(
        changed_theme,
        catalog.candidate_path(contract),
        contract,
        [],
        unresolved,
    )
    assert any(
        issue.code == "IMMUTABLE_PACKAGE_PART_CHANGED"
        and issue.severity == "error"
        for issue in issues
    )
