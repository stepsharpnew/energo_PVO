from __future__ import annotations

import csv
import hashlib
import io
import json
import re
import zipfile
from pathlib import Path

import openpyxl
from lxml import etree
from pypdf import PdfReader, PdfWriter

from .domain import Artifact


ALLOWED_EXTENSIONS = {".pdf", ".xlsx", ".docx", ".png", ".jpg", ".jpeg", ".csv", ".txt"}
MAX_OFFICE_UNCOMPRESSED_BYTES = 300 * 1024 * 1024
MAX_OFFICE_COMPRESSION_RATIO = 300
EXTRACTOR_VERSION = "2"

PILOT_PATTERNS = {
    "kl_04": (r"(?<![а-яa-z])кл\s*[-–—]?\s*0[,.]4", r"кабельн\w*\s+лини\w*\s+0[,.]4"),
    "kl_6": (r"(?<![а-яa-z])кл\s*[-–—]?\s*6(?:\s*кв)?", r"кабельн\w*\s+лини\w*\s+6\s*кв"),
    "vrs": (r"\bврщ\b", r"вводно[- ]распределительн\w*\s+щит"),
}
OUT_OF_SCOPE_PATTERNS = {
    "ktp": (r"\bктп\b", r"трансформаторн\w*\s+подстанц"),
    "vl": (r"\bвл\s*[-–—]?\s*6", r"воздушн\w*\s+лини"),
    "geo": (r"\bгео\b", r"геодез"),
    "gnb": (r"\bгнб\b",),
    "avk": (r"\bавк\b",),
    "emr": (r"\bэмр\b",),
}
EVIDENCE_TERMS = (
    "рабочий проект",
    "технические условия",
    "наименование объекта",
    "адрес",
    "шифр",
    "ведомость объемов",
    "ведомость объёмов",
    "спецификация",
    "кабельная линия",
    "кл-0,4",
    "кл 0,4",
    "кл-6",
    "кл 6",
    "врщ",
    "исполнительная схема",
    "сертификат",
    "паспорт",
)


def validate_signature(path: Path) -> None:
    ext = path.suffix.lower()
    if ext not in ALLOWED_EXTENSIONS:
        raise ValueError(f"Неподдерживаемый формат: {ext}")
    head = path.read_bytes()[:16]
    if ext == ".pdf" and not head.startswith(b"%PDF-"):
        raise ValueError("Расширение PDF не соответствует содержимому")
    if ext in {".xlsx", ".docx"} and not head.startswith(b"PK"):
        raise ValueError("Файл Office не является OOXML ZIP")
    if ext in {".xlsx", ".docx"}:
        try:
            with zipfile.ZipFile(path) as archive:
                members = archive.infolist()
                total = sum(item.file_size for item in members)
                compressed = sum(item.compress_size for item in members)
                if any(item.flag_bits & 0x1 for item in members):
                    raise ValueError("Зашифрованные Office-файлы не поддерживаются")
                if total > MAX_OFFICE_UNCOMPRESSED_BYTES or (compressed and total / compressed > MAX_OFFICE_COMPRESSION_RATIO):
                    raise ValueError("Подозрительный коэффициент сжатия Office-файла")
                required = "xl/workbook.xml" if ext == ".xlsx" else "word/document.xml"
                if required not in archive.namelist() or archive.testzip():
                    raise ValueError("Повреждённая структура OOXML")
        except zipfile.BadZipFile as exc:
            raise ValueError("Повреждённый OOXML ZIP") from exc
    if ext == ".png" and not head.startswith(b"\x89PNG"):
        raise ValueError("Некорректный PNG")
    if ext in {".jpg", ".jpeg"} and not head.startswith(b"\xff\xd8"):
        raise ValueError("Некорректный JPEG")


def classify(path: Path, preview: str = "", original_name: str | None = None) -> str:
    name = (original_name or path.name).lower()
    text = (name + " " + preview[:4000]).lower()
    if path.suffix.lower() in {".png", ".jpg", ".jpeg"}:
        return "image_evidence"
    if "сертифик" in text:
        return "certificate"
    if "паспорт" in text:
        return "passport"
    if "аттест" in text:
        return "attestation"
    if path.suffix.lower() == ".xlsx" and ("акт освидетельствования скрытых работ" in text or name.startswith("аоср")):
        return "filled_aosr"
    if "исполнительн" in text or (path.suffix.lower() == ".pdf" and name.startswith("аоср")) or "ис гео" in name:
        return "execution_scheme"
    if "техническ" in text and "услов" in text:
        return "technical_conditions"
    if "рабочий проект" in text or "состав проекта" in text:
        return "project"
    if path.suffix.lower() == ".xlsx":
        return "spreadsheet"
    return "source"


def extract_pdf(path: Path, page_numbers: list[int] | None = None, max_chars: int = 60_000) -> tuple[str, int]:
    reader = PdfReader(str(path))
    if reader.is_encrypted:
        raise ValueError("Зашифрованные PDF не поддерживаются")
    pages = page_numbers or list(range(1, len(reader.pages) + 1))
    chunks: list[str] = []
    for number in pages:
        if number < 1 or number > len(reader.pages):
            continue
        text = reader.pages[number - 1].extract_text() or ""
        chunks.append(f"\n--- PAGE {number} ---\n{text}")
        if sum(map(len, chunks)) >= max_chars:
            chunks.append("\n[TRUNCATED]")
            break
    return "".join(chunks)[:max_chars], len(reader.pages)


def extract_docx(path: Path, max_chars: int = 60_000) -> str:
    with zipfile.ZipFile(path) as archive:
        xml = archive.read("word/document.xml")
    root = etree.fromstring(xml)
    text = "\n".join(root.itertext())
    return text[:max_chars]


def extract_xlsx(path: Path, sheet_names: list[str] | None = None, max_chars: int = 60_000) -> str:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=False, keep_links=True)
    names = sheet_names or workbook.sheetnames
    chunks: list[str] = []
    for name in names:
        if name not in workbook.sheetnames:
            continue
        sheet = workbook[name]
        chunks.append(f"\n--- SHEET {name} [{sheet.sheet_state}] ---")
        for row in sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 1000)):
            values = [f"{cell.coordinate}={cell.value}" for cell in row if cell.value not in (None, "")]
            if values:
                chunks.append(" | ".join(values))
            if sum(map(len, chunks)) >= max_chars:
                chunks.append("[TRUNCATED]")
                return "\n".join(chunks)[:max_chars]
    return "\n".join(chunks)[:max_chars]


def extract_csv(path: Path, max_chars: int = 60_000) -> str:
    raw = path.read_bytes()
    for encoding in ("utf-8-sig", "cp1251"):
        try:
            text = raw.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        text = raw.decode("utf-8", errors="replace")
    rows = csv.reader(io.StringIO(text))
    return "\n".join(" | ".join(row) for row in rows)[:max_chars]


def extract_source(path: Path, *, pages: list[int] | None = None, sheets: list[str] | None = None) -> tuple[str, int | None]:
    ext = path.suffix.lower()
    if ext == ".pdf":
        return extract_pdf(path, pages)
    if ext == ".xlsx":
        return extract_xlsx(path, sheets), None
    if ext == ".docx":
        return extract_docx(path), None
    if ext == ".csv":
        return extract_csv(path), None
    if ext == ".txt":
        return path.read_text(encoding="utf-8", errors="replace")[:60_000], None
    return f"Binary image {path.name}; use the uploaded image as visual evidence.", None


def _normalized(value: str) -> str:
    return re.sub(r"\s+", " ", value.lower().replace("ё", "е")).strip()


def detect_scope(text: str, original_name: str) -> str:
    """Route paid context; this is not accepted as factual document evidence."""
    content = _normalized(text)
    name = _normalized(original_name)
    # Explicit names are a cheap routing hint, never a Claim. Content remains
    # available in the inventory, and unknown/conflicting files stay eligible
    # for model inspection rather than becoming document facts here.
    for source in (name, content[:2_500], content):
        for family, patterns in PILOT_PATTERNS.items():
            if any(re.search(pattern, source) for pattern in patterns):
                return family
        for family, patterns in OUT_OF_SCOPE_PATTERNS.items():
            if any(re.search(pattern, source) for pattern in patterns):
                return family
    return "unknown"


def _segment_score(text: str, *, first: bool = False) -> int:
    normalized = _normalized(text)
    score = 20 if first else 0
    score += sum(12 for term in EVIDENCE_TERMS if term in normalized)
    score += min(len(text) // 1000, 12)
    return score


def _build_segments(path: Path) -> tuple[list[dict], int | None]:
    ext = path.suffix.lower()
    if ext == ".pdf":
        reader = PdfReader(str(path))
        if reader.is_encrypted:
            raise ValueError("Зашифрованные PDF не поддерживаются")
        segments = []
        for number, page in enumerate(reader.pages, 1):
            text = (page.extract_text() or "")[:25_000]
            segments.append(
                {
                    "locator": f"page:{number}",
                    "page": number,
                    "sheet": None,
                    "text": text,
                    "char_count": len(text),
                    "visual_required": len(re.sub(r"\s+", "", text)) < 80,
                    "score": _segment_score(text, first=number == 1),
                }
            )
        return segments, len(reader.pages)
    if ext == ".xlsx":
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=False, keep_links=True)
        segments = []
        try:
            for position, sheet in enumerate(workbook.worksheets):
                lines = [f"--- SHEET {sheet.title} [{sheet.sheet_state}] ---"]
                for row in sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 1000)):
                    values = [f"{cell.coordinate}={cell.value}" for cell in row if cell.value not in (None, "")]
                    if values:
                        lines.append(" | ".join(values))
                    if sum(map(len, lines)) >= 25_000:
                        lines.append("[TRUNCATED]")
                        break
                text = "\n".join(lines)[:25_000]
                segments.append(
                    {
                        "locator": f"sheet:{sheet.title}",
                        "page": None,
                        "sheet": sheet.title,
                        "text": text,
                        "char_count": len(text),
                        "visual_required": False,
                        "score": _segment_score(text, first=position == 0),
                    }
                )
        finally:
            workbook.close()
        return segments, None
    text, pages = extract_source(path)
    return (
        [
            {
                "locator": "file",
                "page": None,
                "sheet": None,
                "text": text,
                "char_count": len(text),
                "visual_required": ext in {".png", ".jpg", ".jpeg"},
                "score": _segment_score(text, first=True),
            }
        ],
        pages,
    )


def _file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def source_index(root: Path, artifact: Artifact) -> dict:
    path = root / "input" / artifact.stored_name
    cache_dir = root / "extracted"
    cache_dir.mkdir(parents=True, exist_ok=True)
    cache_path = cache_dir / f"{artifact.sha256}-{EXTRACTOR_VERSION}.json"
    digest = _file_sha256(path)
    if digest != artifact.sha256:
        raise ValueError(f"Файл {artifact.original_name} изменился после загрузки")
    if cache_path.exists():
        try:
            cached = json.loads(cache_path.read_text(encoding="utf-8"))
            if cached.get("sha256") == artifact.sha256 and cached.get("extractor_version") == EXTRACTOR_VERSION:
                joined = "\n".join(item.get("text", "") for item in cached.get("segments", []))
                # Extraction is content-addressed, while a filename is only a
                # routing hint. Rebind routing metadata for duplicate bytes
                # uploaded under different names.
                cached["original_name"] = artifact.original_name
                cached["scope"] = detect_scope(joined[:120_000], artifact.original_name)
                return cached
        except (OSError, json.JSONDecodeError):
            pass
    segments, pages = _build_segments(path)
    joined = "\n".join(item["text"] for item in segments)
    payload = {
        "extractor_version": EXTRACTOR_VERSION,
        "sha256": artifact.sha256,
        "original_name": artifact.original_name,
        "pages": pages,
        "scope": detect_scope(joined[:120_000], artifact.original_name),
        "segments": segments,
    }
    temporary = cache_path.with_suffix(".tmp")
    temporary.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")
    temporary.replace(cache_path)
    return payload


def read_indexed_source(
    root: Path,
    artifact: Artifact,
    *,
    pages: list[int] | None = None,
    sheets: list[str] | None = None,
    max_chars: int = 18_000,
) -> tuple[str, int | None]:
    index = source_index(root, artifact)
    selected = []
    page_set = set(pages or [])
    sheet_set = set(sheets or [])
    for segment in index["segments"]:
        if page_set and segment.get("page") not in page_set:
            continue
        if sheet_set and segment.get("sheet") not in sheet_set:
            continue
        selected.append(f"\n--- {segment['locator']} ---\n{segment['text']}")
    text = "".join(selected)
    if len(text) > max_chars:
        text = text[:max_chars] + "\n[TRUNCATED BY SOURCE BUDGET]"
    return text, index.get("pages")


def _pilot_match_count(text: str) -> int:
    normalized = _normalized(text)
    return sum(
        len(re.findall(pattern, normalized))
        for patterns in PILOT_PATTERNS.values()
        for pattern in patterns
    )


def _segment_context_score(segment: dict, *, project: bool) -> int:
    score = int(segment.get("score") or 0)
    matches = _pilot_match_count(segment.get("text", ""))
    if project:
        if segment.get("visual_required"):
            score += 1_200
        if matches:
            score += 900 + min(matches, 20) * 40
        if segment.get("page") == 1:
            score += 800
    elif segment.get("visual_required"):
        score += 600
    return score


def build_compact_evidence(root: Path, artifacts: list[Artifact], max_chars: int) -> list[dict]:
    candidates: list[tuple[int, Artifact, dict, str]] = []
    for artifact in artifacts:
        if artifact.category == "filled_aosr":
            continue
        index = source_index(root, artifact)
        scope = index.get("scope", "unknown")
        if artifact.category == "execution_scheme" and scope in OUT_OF_SCOPE_PATTERNS:
            continue
        is_project = artifact.category in {"project", "technical_conditions"}
        per_file = 32 if is_project else 3
        ranked = sorted(
            index["segments"],
            key=lambda item: (-_segment_context_score(item, project=is_project), str(item["locator"])),
        )[:per_file]
        category_bonus = {
            "project": 50,
            "technical_conditions": 45,
            "execution_scheme": 40,
            "passport": 35,
            "certificate": 35,
            "attestation": 30,
        }.get(artifact.category, 10)
        for segment in ranked:
            candidates.append(
                (
                    category_bonus + _segment_context_score(segment, project=is_project),
                    artifact,
                    segment,
                    scope,
                )
            )
    packet: list[dict] = []
    used = 0
    for _, artifact, segment, scope in sorted(candidates, key=lambda item: -item[0]):
        if used >= max_chars:
            break
        record_overhead = 180
        remaining = max_chars - used - record_overhead
        if remaining <= 0:
            break
        excerpt = segment["text"][: min(8_000, remaining)]
        if not excerpt.strip():
            continue
        packet.append(
            {
                "file_id": artifact.id,
                "category": artifact.category,
                "scope_hint": scope,
                "locator": segment["locator"],
                "visual_required": segment["visual_required"],
                "text": excerpt,
            }
        )
        used += len(excerpt) + record_overhead
    return packet


def _selected_pdf_pages(index: dict, *, project: bool, limit: int | None = None) -> list[int]:
    segments = [item for item in index["segments"] if item.get("page")]
    effective_limit = limit if limit is not None else (len(segments) if project else 4)
    ranked = sorted(
        segments,
        key=lambda item: (-_segment_context_score(item, project=project), int(item["page"])),
    )
    return sorted(int(item["page"]) for item in ranked[:effective_limit])


def _pdf_subset(source: Path, destination: Path, pages: list[int]) -> Path:
    if destination.exists():
        try:
            cached = PdfReader(str(destination))
            if not cached.is_encrypted and len(cached.pages) == len(pages):
                return destination
        except Exception:
            pass
    reader = PdfReader(str(source))
    writer = PdfWriter()
    for page in pages:
        if 1 <= page <= len(reader.pages):
            writer.add_page(reader.pages[page - 1])
    destination.parent.mkdir(parents=True, exist_ok=True)
    temporary = destination.with_suffix(".tmp")
    with temporary.open("wb") as stream:
        writer.write(stream)
    temporary.replace(destination)
    return destination


def select_visual_sources(
    root: Path,
    artifacts: list[Artifact],
    *,
    max_pages: int,
    include_project: bool,
) -> list[dict]:
    page_candidates: list[tuple[int, str, int, Artifact, Path, str, bool]] = []
    for artifact in artifacts:
        path = root / "input" / artifact.stored_name
        ext = path.suffix.lower()
        index = source_index(root, artifact)
        scope = index.get("scope", "unknown")
        if artifact.category == "filled_aosr":
            continue
        if artifact.category == "execution_scheme" and scope in OUT_OF_SCOPE_PATTERNS:
            continue
        if ext in {".png", ".jpg", ".jpeg"}:
            page_candidates.append((1_250, artifact.original_name.casefold(), 1, artifact, path, "image evidence", True))
            continue
        if ext != ".pdf":
            continue
        is_project = artifact.category in {"project", "technical_conditions"}
        if is_project and not include_project:
            continue
        visually_relevant = artifact.category == "execution_scheme" or is_project or any(
            item.get("visual_required") for item in index["segments"]
        )
        if not visually_relevant:
            continue
        segments = {int(item["page"]): item for item in index["segments"] if item.get("page")}
        if is_project:
            pages = _selected_pdf_pages(index, project=True, limit=None)
        elif artifact.category == "execution_scheme":
            pages = _selected_pdf_pages(index, project=False, limit=len(index["segments"]))
        else:
            visual_pages = [int(item["page"]) for item in index["segments"] if item.get("page") and item.get("visual_required")]
            pages = visual_pages or _selected_pdf_pages(index, project=False, limit=4)
        for page in pages:
            segment = segments[page]
            required = False
            if is_project:
                if segment.get("visual_required"):
                    priority, reason = 1_300, "project page without reliable text layer"
                    required = True
                elif _pilot_match_count(segment.get("text", "")):
                    priority, reason = 1_100 + min(_pilot_match_count(segment.get("text", "")), 20), "pilot-family project evidence"
                elif page == 1:
                    priority, reason = 1_000, "project title page"
                else:
                    priority, reason = 500 + int(segment.get("score") or 0), "ranked project evidence"
            elif artifact.category == "execution_scheme":
                priority = 1_200 if scope in PILOT_PATTERNS else 900
                reason = "pilot execution scheme" if scope in PILOT_PATTERNS else "unclassified execution scheme"
                required = True
            else:
                priority, reason = 800 + int(segment.get("visual_required") or 0) * 100, "visual source evidence"
                required = bool(segment.get("visual_required"))
            page_candidates.append((priority, artifact.original_name.casefold(), page, artifact, path, reason, required))

    ranked = sorted(page_candidates, key=lambda item: (-item[0], item[1], item[2]))
    required_candidates = [item for item in ranked if item[6]]
    if len(required_candidates) > max_pages:
        raise ValueError(
            f"Минимально необходимый визуальный контекст требует {len(required_candidates)} страниц, "
            f"а выбранный профиль разрешает {max_pages}. Используйте более высокий профиль или увеличьте лимит."
        )
    required_keys = {(item[3].id, item[2]) for item in required_candidates}
    optional = [item for item in ranked if (item[3].id, item[2]) not in required_keys]
    chosen = required_candidates + optional[: max_pages - len(required_candidates)]
    chosen.sort(key=lambda item: (-item[0], item[1], item[2]))
    grouped: dict[str, dict] = {}
    for priority, _, page, artifact, path, reason, _ in chosen:
        item = grouped.setdefault(
            artifact.id,
            {"artifact": artifact, "source_path": path, "pages": [], "reasons": set(), "priority": priority},
        )
        item["pages"].append(page)
        item["reasons"].add(reason)
        item["priority"] = max(item["priority"], priority)

    selected: list[dict] = []
    for item in sorted(grouped.values(), key=lambda value: (-value["priority"], value["artifact"].original_name.casefold())):
        artifact = item["artifact"]
        path = item["source_path"]
        pages = sorted(set(item["pages"]))
        visual_path = path
        if path.suffix.lower() == ".pdf":
            total_pages = int(source_index(root, artifact).get("pages") or len(pages))
            if pages != list(range(1, total_pages + 1)):
                suffix = hashlib.sha256(",".join(map(str, pages)).encode()).hexdigest()[:10]
                visual_path = _pdf_subset(path, root / "extracted" / "visual" / f"{artifact.sha256}-{suffix}.pdf", pages)
        selected.append(
            {
                "artifact": artifact,
                "path": visual_path,
                "pages": pages,
                "reason": "; ".join(sorted(item["reasons"])),
            }
        )
    return selected


def build_inventory(root: Path, artifacts: list[Artifact]) -> tuple[list[Artifact], str]:
    updated: list[Artifact] = []
    manifest: list[dict] = []
    for artifact in artifacts:
        path = root / "input" / artifact.stored_name
        validate_signature(path)
        try:
            index = source_index(root, artifact)
            preview = "\n".join(item["text"] for item in index["segments"][:2])
            pages = index.get("pages")
        except Exception as exc:
            raise ValueError(f"Не удалось прочитать {artifact.original_name}: {exc}") from exc
        item = artifact.model_copy(update={"category": classify(path, preview, artifact.original_name), "pages": pages})
        updated.append(item)
        manifest.append(
            {
                "id": item.id,
                "name": item.original_name,
                "stored_name": item.stored_name,
                "category": item.category,
                "media_type": item.media_type,
                "size": item.size,
                "sha256": item.sha256,
                "pages": item.pages,
                "scope_hint": index.get("scope", "unknown"),
                "preview": re.sub(r"\s+", " ", preview)[:1200],
            }
        )
    return updated, json.dumps(manifest, ensure_ascii=False, indent=2)
