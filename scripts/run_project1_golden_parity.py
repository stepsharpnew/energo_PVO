#!/usr/bin/env python3
"""Generate the project1 pilot workbooks and compare them with the golden XLSX files.

This is a diagnostic fixture-parity run, not a blind agent run:

* the cleaned candidates in ``templates/approved`` are the source templates;
* project1 golden workbooks supply the facts that would normally be provided by
  source documents or confirmed through NEEDS_INPUT;
* generation still goes through ExcelGenerator and template contracts;
* generated workbooks are never promoted to a final/approved job;
* known golden defects are reported and are not treated as production truth.
"""

from __future__ import annotations

import json
import re
import sys
import unicodedata
import warnings
import zipfile
from collections import Counter
from dataclasses import dataclass
from datetime import date, datetime, timezone
from decimal import Decimal, InvalidOperation
from html import escape
from pathlib import Path
from typing import Any, Iterable

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "src"))

from executive_docs.config import settings  # noqa: E402
from executive_docs.domain import (  # noqa: E402
    ChangeState,
    Claim,
    ClaimStatus,
    DocumentPlan,
    Material,
    WorkItem,
)
from executive_docs.excel import ExcelGenerator, TemplateContract, sha256  # noqa: E402
from executive_docs.packaging import render_selected_sheets  # noqa: E402
from executive_docs.validation import validate_semantics, validate_workbook  # noqa: E402


warnings.filterwarnings(
    "ignore",
    message="Data Validation extension is not supported and will be removed",
)


@dataclass(frozen=True)
class GoldenSpec:
    family: str
    template_id: str
    golden: Path
    selected_sheets: tuple[str, ...]


SPECS = (
    GoldenSpec(
        family="kl_04",
        template_id="aosr_kl_04",
        golden=ROOT / "project1" / "9. АОСР КЛ.xlsx",
        selected_sheets=("АОСР-3", "АОСР-4"),
    ),
    GoldenSpec(
        family="kl_6",
        template_id="aosr_kl_6",
        golden=ROOT / "project1" / "9. АОСР КЛ 6кВ.xlsx",
        selected_sheets=tuple(f"АОСР-{index}" for index in range(1, 8)),
    ),
    GoldenSpec(
        family="vrs",
        template_id="aosr_vrs",
        golden=ROOT / "project1" / "10. АОСР ВРЩ.xlsx",
        selected_sheets=tuple(f"АОСР-{index}" for index in range(1, 7)),
    ),
)


class WorkbookReader:
    """Read raw formulas and their cached semantic values without saving a workbook."""

    def __init__(self, path: Path):
        self.path = path
        self.raw = openpyxl.load_workbook(path, data_only=False, keep_links=True)
        self.cached = openpyxl.load_workbook(path, data_only=True, keep_links=True)

    def close(self) -> None:
        self.raw.close()
        self.cached.close()

    def raw_value(self, sheet: str, cell: str) -> Any:
        return self.raw[sheet][cell].value

    def semantic_value(self, sheet: str, cell: str) -> Any:
        raw = self.raw_value(sheet, cell)
        cached = self.cached[sheet][cell].value
        if isinstance(raw, str) and raw.startswith("="):
            return cached if cached is not None else raw
        return cached if cached is not None else raw

    def sheet_names(self) -> list[str]:
        return list(self.raw.sheetnames)


def value_text(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%d.%m.%Y")
    if isinstance(value, date):
        return value.strftime("%d.%m.%Y")
    if isinstance(value, bool):
        return "да" if value else "нет"
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        return format(Decimal(str(value)), "f").rstrip("0").rstrip(".")
    return str(value)


def canonical(value: Any) -> str:
    text = value_text(value)
    if text is None:
        return ""
    text = unicodedata.normalize("NFKC", text).replace("\xa0", " ")
    text = re.sub(r"\s+", " ", text).strip()
    numeric = text.replace(" ", "").replace(",", ".")
    if re.fullmatch(r"[-+]?\d+(?:\.\d+)?", numeric):
        try:
            normalized = format(Decimal(numeric).normalize(), "f")
            return "0" if normalized in {"-0", ""} else normalized
        except InvalidOperation:
            pass
    return text


def split_target(target: str) -> tuple[str, str]:
    return tuple(target.split("!", 1))  # type: ignore[return-value]


def iter_targets(value: str | list[str]) -> Iterable[str]:
    return value if isinstance(value, list) else [value]


def iter_mapping_cells(
    mapping: dict[str, Any],
    prefix: str = "",
) -> Iterable[tuple[str, str]]:
    for key, value in mapping.items():
        if key in {"suffix_value", "separator_value"}:
            continue
        field = f"{prefix}.{key}" if prefix else key
        if isinstance(value, str) and re.fullmatch(r"[A-Z]{1,3}[1-9][0-9]*", value):
            yield field, value
        elif isinstance(value, dict):
            yield from iter_mapping_cells(value, field)
        elif isinstance(value, list):
            for index, item in enumerate(value, 1):
                if isinstance(item, dict):
                    yield from iter_mapping_cells(item, f"{field}.{index}")


def split_quantity_unit(value: str | None) -> tuple[str | None, str | None]:
    if value is None or not canonical(value):
        return None, None
    match = re.fullmatch(r"\s*([-+]?\d+(?:[.,]\d+)?)\s*(.*?)\s*", value)
    if not match:
        return value, None
    return match.group(1), match.group(2) or None


def observed_claim(
    *,
    key: str,
    value: Any,
    golden: Path,
    locator: str,
    affected: list[str],
) -> Claim:
    text = value_text(value)
    if text is None:
        raise ValueError(f"Cannot create a Claim without a value: {key}")
    return Claim(
        key=key,
        raw_value=text,
        normalized_value=text,
        source_kind="golden_fixture",
        source_file_id=f"golden:{golden.name}",
        locator=f"{golden.name}:{locator}",
        evidence_fragment=(
            "Значение прочитано из утверждённого семантического примера project1 "
            "только для диагностического parity-теста."
        ),
        status=ClaimStatus.OBSERVED,
        affected_documents=affected,
    )


def build_fixture(
    spec: GoldenSpec,
    contract: TemplateContract,
    reader: WorkbookReader,
    sequence_start: int,
) -> tuple[list[Claim], list[WorkItem], DocumentPlan, dict[str, str], int]:
    claims: list[Claim] = []
    common_values: dict[str, str] = {}
    for key, target_value in contract.common_fields.items():
        targets = list(iter_targets(target_value))
        value = None
        source_target = targets[0]
        for target in targets:
            sheet, cell = split_target(target)
            candidate = reader.semantic_value(sheet, cell)
            if canonical(candidate):
                value = candidate
                source_target = target
                break
        text = value_text(value)
        if text is None or not canonical(text):
            continue
        common_values[key] = text
        claims.append(
            observed_claim(
                key=key,
                value=text,
                golden=spec.golden,
                locator=source_target,
                affected=[spec.template_id],
            )
        )

    work_items: list[WorkItem] = []
    attachment_ids: list[str] = []
    sequence = sequence_start
    for index, sheet in enumerate(spec.selected_sheets, 1):
        mapping = contract.sheets[sheet]

        def field(name: str) -> str | None:
            cell = mapping.get(name)
            if not isinstance(cell, str):
                return None
            return value_text(reader.semantic_value(sheet, cell))

        work_type = field("work_name")
        if not work_type:
            raise ValueError(f"Golden work name is empty: {spec.golden.name}:{sheet}")
        source_keys: list[str] = []

        def add_source(kind: str, value: str | None, cell_name: str | None = None) -> None:
            if value is None or not canonical(value):
                return
            key = f"golden.{spec.template_id}.{index}.{kind}"
            source_keys.append(key)
            claims.append(
                observed_claim(
                    key=key,
                    value=value,
                    golden=spec.golden,
                    locator=f"{sheet}!{cell_name or mapping.get(kind, '?')}",
                    affected=[spec.template_id],
                )
            )

        start_date = field("start_date")
        end_date = field("end_date")
        volume = field("volume")
        unit = field("unit")
        installation = field("installation")
        subsequent_work = field("subsequent_work")
        attachment_name = field("attachment")
        add_source("work", work_type, mapping.get("work_name"))
        add_source("start", start_date, mapping.get("start_date"))
        add_source("end", end_date, mapping.get("end_date"))
        add_source("volume", volume, mapping.get("volume"))
        add_source("installation", installation, mapping.get("installation"))
        add_source("subsequent_work", subsequent_work, mapping.get("subsequent_work"))

        materials: list[Material] = []
        material_rows = mapping.get("material_rows")
        if isinstance(material_rows, list):
            for row in material_rows:
                if not isinstance(row, dict) or not isinstance(row.get("name"), str):
                    continue
                material_name = value_text(reader.semantic_value(sheet, row["name"]))
                if not material_name or not canonical(material_name):
                    continue
                material_quantity = (
                    value_text(reader.semantic_value(sheet, row["quantity"]))
                    if isinstance(row.get("quantity"), str)
                    else None
                )
                material_unit = (
                    value_text(reader.semantic_value(sheet, row["unit"]))
                    if isinstance(row.get("unit"), str)
                    else None
                )
                if isinstance(row.get("quantity_unit"), str):
                    material_quantity, material_unit = split_quantity_unit(
                        value_text(reader.semantic_value(sheet, row["quantity_unit"]))
                    )
                material_document = (
                    value_text(reader.semantic_value(sheet, row["document"]))
                    if isinstance(row.get("document"), str)
                    else None
                )
                materials.append(
                    Material(
                        name=material_name,
                        quantity=material_quantity,
                        unit=material_unit,
                        quality_document=material_document,
                        source_file_id=f"golden:{spec.golden.name}",
                    )
                )
        else:
            material_name = field("material_name")
            material_document = field("material_document")
            if material_name:
                materials.append(
                    Material(
                        name=material_name,
                        quality_document=material_document,
                        source_file_id=f"golden:{spec.golden.name}",
                    )
                )
        scheme_id = None
        if attachment_name:
            scheme_id = f"golden-scheme-{spec.template_id}-{index}"
            attachment_ids.append(scheme_id)
            claims.append(
                observed_claim(
                    key=f"artifact.{scheme_id}.name",
                    value=attachment_name,
                    golden=spec.golden,
                    locator=f"{sheet}!{mapping['attachment']}",
                    affected=[spec.template_id],
                )
            )
        work_items.append(
            WorkItem(
                id=f"golden-{spec.template_id}-{index}",
                family=spec.family,
                work_type=work_type,
                sequence_index=sequence,
                actual_start=start_date,
                actual_end=end_date,
                volume=volume,
                unit=unit,
                installation=installation,
                subsequent_work=subsequent_work,
                materials=materials,
                change_state=ChangeState.YES if scheme_id else ChangeState.NO,
                execution_scheme_id=scheme_id,
                source_claim_keys=source_keys,
            )
        )
        sequence += 1

    first_mapping = contract.sheets[spec.selected_sheets[0]]
    first_number_value = reader.semantic_value(spec.selected_sheets[0], first_mapping.get("number", "C32"))
    try:
        first_number = int(Decimal(canonical(first_number_value)))
    except (ValueError, InvalidOperation):
        first_number = 1
    plan = DocumentPlan(
        template_id=spec.template_id,
        selected_sheets=list(spec.selected_sheets),
        work_item_ids=[item.id for item in work_items],
        first_number=first_number,
        attachments=attachment_ids,
        output_filename=contract.output_filename,
    )
    return claims, work_items, plan, common_values, sequence


def mapped_comparison(
    contract: TemplateContract,
    plan: DocumentPlan,
    golden: WorkbookReader,
    generated: WorkbookReader,
) -> dict[str, Any]:
    rows: list[dict[str, Any]] = []
    for field, target_value in contract.common_fields.items():
        for target in iter_targets(target_value):
            sheet, cell = split_target(target)
            expected = golden.semantic_value(sheet, cell)
            actual = generated.semantic_value(sheet, cell)
            rows.append(
                {
                    "kind": "common",
                    "field": field,
                    "location": target,
                    "golden_raw": value_text(golden.raw_value(sheet, cell)),
                    "expected": value_text(expected),
                    "actual": value_text(actual),
                    "match": canonical(expected) == canonical(actual),
                }
            )
    for sheet in plan.selected_sheets:
        for field, cell in iter_mapping_cells(contract.sheets[sheet]):
            expected = golden.semantic_value(sheet, cell)
            actual = generated.semantic_value(sheet, cell)
            rows.append(
                {
                    "kind": "act",
                    "field": field,
                    "location": f"{sheet}!{cell}",
                    "golden_raw": value_text(golden.raw_value(sheet, cell)),
                    "expected": value_text(expected),
                    "actual": value_text(actual),
                    "match": canonical(expected) == canonical(actual),
                }
            )
    return {
        "total": len(rows),
        "matches": sum(row["match"] for row in rows),
        "mismatches": [row for row in rows if not row["match"]],
        "rows": rows,
    }


def color_signature(color: Any) -> dict[str, Any]:
    return {
        "type": color.type,
        "rgb": color.rgb if color.type == "rgb" else None,
        "indexed": color.indexed if color.type == "indexed" else None,
        "theme": color.theme if color.type == "theme" else None,
        "tint": color.tint,
    }


def fill_signature(cell: Any) -> dict[str, Any]:
    return {
        "fill_type": cell.fill.fill_type,
        "pattern_type": cell.fill.patternType,
        "foreground": color_signature(cell.fill.fgColor),
        "background": color_signature(cell.fill.bgColor),
    }


def mapped_fill_comparison(
    contract: TemplateContract,
    plan: DocumentPlan,
    golden: WorkbookReader,
    generated: WorkbookReader,
) -> dict[str, Any]:
    locations = {
        target
        for target_value in contract.common_fields.values()
        for target in iter_targets(target_value)
    }
    locations.update(
        f"{sheet}!{cell}"
        for sheet in plan.selected_sheets
        for _, cell in iter_mapping_cells(contract.sheets[sheet])
    )
    mismatches: list[dict[str, Any]] = []
    for location in sorted(locations):
        sheet, cell = split_target(location)
        expected = fill_signature(golden.raw[sheet][cell])
        actual = fill_signature(generated.raw[sheet][cell])
        if expected != actual:
            mismatches.append(
                {
                    "location": location,
                    "golden": expected,
                    "generated": actual,
                }
            )
    return {
        "total": len(locations),
        "matches": len(locations) - len(mismatches),
        "mismatches": mismatches,
    }


def variable_locations(spec: GoldenSpec) -> list[tuple[str, str]]:
    locations = [("Данные объект", f"B{row}") for row in range(2, 45)]
    rows = (32, 62, *range(69, 81), *range(85, 89))
    for sheet in spec.selected_sheets:
        for row in rows:
            for column in range(1, 37):
                locations.append((sheet, openpyxl.utils.get_column_letter(column) + str(row)))
    return locations


def contract_gaps(
    spec: GoldenSpec,
    contract: TemplateContract,
    baseline: WorkbookReader,
    golden: WorkbookReader,
    generated: WorkbookReader,
) -> list[dict[str, Any]]:
    writable = contract.writable_cells()
    gaps: list[dict[str, Any]] = []
    for sheet, cell in variable_locations(spec):
        location = f"{sheet}!{cell}"
        if location in writable:
            continue
        golden_raw = golden.raw_value(sheet, cell)
        if isinstance(golden_raw, str) and golden_raw.startswith("="):
            continue
        expected = golden.semantic_value(sheet, cell)
        if not canonical(expected):
            continue
        baseline_value = baseline.semantic_value(sheet, cell)
        if canonical(expected) == canonical(baseline_value):
            continue
        row_number = int(re.search(r"\d+", cell).group())
        if sheet == "Данные объект":
            category = "object_card"
        elif row_number in {69, 70, 71}:
            category = "additional_material_or_quality_document"
        elif row_number in {72, 74}:
            category = "attachment"
        elif 76 <= row_number <= 80:
            category = "date"
        elif 85 <= row_number <= 88:
            category = "subsequent_work"
        else:
            category = "act_content"
        gaps.append(
            {
                "location": location,
                "category": category,
                "baseline": value_text(baseline_value),
                "golden": value_text(expected),
                "generated": value_text(generated.semantic_value(sheet, cell)),
            }
        )
    existing = {item["location"] for item in gaps}
    mapped = {
        f"{sheet}!{cell}"
        for sheet in spec.selected_sheets
        for _, cell in iter_mapping_cells(contract.sheets[sheet])
    }
    for sheet in spec.selected_sheets:
        for cell in contract.clear_cells.get(sheet, []):
            location = f"{sheet}!{cell}"
            if location in existing or location in mapped:
                continue
            expected = golden.semantic_value(sheet, cell)
            if not canonical(expected):
                continue
            gaps.append(
                {
                    "location": location,
                    "category": "cleared_without_semantic_field",
                    "baseline": value_text(baseline.semantic_value(sheet, cell)),
                    "golden": value_text(expected),
                    "generated": value_text(generated.semantic_value(sheet, cell)),
                }
            )
    return gaps


def non_whitelist_changes(
    contract: TemplateContract,
    baseline: WorkbookReader,
    generated: WorkbookReader,
) -> list[dict[str, Any]]:
    writable = contract.writable_cells()
    changes: list[dict[str, Any]] = []
    for sheet in baseline.sheet_names():
        base_sheet = baseline.raw[sheet]
        generated_sheet = generated.raw[sheet]
        max_row = max(base_sheet.max_row, generated_sheet.max_row)
        max_column = max(base_sheet.max_column, generated_sheet.max_column)
        for row in range(1, max_row + 1):
            for column in range(1, max_column + 1):
                cell = openpyxl.utils.get_column_letter(column) + str(row)
                location = f"{sheet}!{cell}"
                if location in writable:
                    continue
                before = base_sheet[cell].value
                after = generated_sheet[cell].value
                if before != after:
                    changes.append(
                        {
                            "location": location,
                            "before": value_text(before),
                            "after": value_text(after),
                        }
                    )
    return changes


def placeholder_quality_documents(work_items: list[WorkItem]) -> list[dict[str, str]]:
    findings: list[dict[str, str]] = []
    for item in work_items:
        for material in item.materials:
            value = (material.quality_document or "").strip()
            collapsed = re.sub(r"\s+", " ", value).casefold()
            if (
                not collapsed
                or re.search(r"\bб\s*/\s*[нд]\b", collapsed)
                or re.search(r"№\s*(?:от|$|\))", collapsed)
            ):
                findings.append(
                    {
                        "work": item.work_type,
                        "material": material.name,
                        "quality_document": value,
                    }
                )
    return findings


def report_html(report: dict[str, Any]) -> str:
    family_rows = ""
    for result in report["workbooks"]:
        comparison = result["mapped_comparison"]
        structural_errors = [
            issue
            for issue in result["technical_validation"]
            if issue["code"] != "TEMPLATE_NOT_APPROVED"
        ]
        family_rows += (
            "<tr>"
            f"<td>{escape(result['template_id'])}</td>"
            f"<td>{comparison['matches']} / {comparison['total']}</td>"
            f"<td>{len(comparison['mismatches'])}</td>"
            f"<td>{len(result['mapped_fill_comparison']['mismatches'])}</td>"
            f"<td>{len(result['contract_gaps'])}</td>"
            f"<td>{len(result['non_whitelist_changes'])}</td>"
            f"<td>{len(structural_errors)}</td>"
            f"<td>{result['preview_count']}</td>"
            "</tr>"
        )
    mismatch_rows = ""
    for result in report["workbooks"]:
        for row in result["mapped_comparison"]["mismatches"]:
            mismatch_rows += (
                "<tr>"
                f"<td>{escape(result['template_id'])}</td>"
                f"<td>{escape(row['location'])}</td>"
                f"<td>{escape(str(row['expected']))}</td>"
                f"<td>{escape(str(row['actual']))}</td>"
                "</tr>"
            )
    gap_rows = ""
    for result in report["workbooks"]:
        for row in result["contract_gaps"][:100]:
            gap_rows += (
                "<tr>"
                f"<td>{escape(result['template_id'])}</td>"
                f"<td>{escape(row['category'])}</td>"
                f"<td>{escape(row['location'])}</td>"
                f"<td>{escape(str(row['golden']))}</td>"
                "</tr>"
            )
    semantic_rows = "".join(
        "<tr>"
        f"<td>{escape(issue['severity'])}</td>"
        f"<td>{escape(issue['code'])}</td>"
        f"<td>{escape(issue['message'])}</td>"
        "</tr>"
        for issue in report["semantic_validation"]
    )
    placeholder_rows = "".join(
        "<tr>"
        f"<td>{escape(item['work'])}</td>"
        f"<td>{escape(item['material'])}</td>"
        f"<td>{escape(item['quality_document'])}</td>"
        "</tr>"
        for item in report["placeholder_quality_documents"]
    )
    return f"""<!doctype html>
<html lang="ru"><head><meta charset="utf-8"><title>project1 — golden parity</title>
<style>
body{{font:14px Arial,sans-serif;max-width:1250px;margin:28px auto;color:#172033;line-height:1.45}}
h1,h2{{color:#153a67}} table{{border-collapse:collapse;width:100%;margin:14px 0 28px}}
th,td{{border:1px solid #d6deea;padding:7px;vertical-align:top;text-align:left}}
th{{background:#eef4fb}} .warning{{background:#fff7dd;border:1px solid #e7c860;padding:12px}}
.ok{{color:#157347}} .bad{{color:#b42318}} code{{background:#f2f4f7;padding:2px 4px}}
</style></head><body>
<h1>Диагностический прогон project1 → очищенные шаблоны</h1>
<p class="warning"><b>Не для подписания.</b> Это fixture-parity тест: значения, которых не было
в исходных PDF или подтверждённых профилях, взяты из готовых XLSX project1 как тестовые ответы.
NEEDS_INPUT не изменён и не отключён в рабочем приложении.</p>
<p><b>Создано:</b> {escape(report['created_at'])}. <b>Вызовы модели:</b> 0.</p>
<h2>Результат по книгам</h2>
<table><tr><th>Контракт</th><th>Совпало mapped</th><th>Не совпало</th>
<th>Отличия заливки</th><th>Пробелы контракта</th><th>Изменения вне whitelist</th>
<th>Технические ошибки*</th><th>PDF АОСР</th></tr>{family_rows}</table>
<p>* Без ожидаемого блокера <code>TEMPLATE_NOT_APPROVED</code>.</p>
<h2>Несовпадения в уже поддержанных полях</h2>
<table><tr><th>Книга</th><th>Ячейка</th><th>Эталон</th><th>Результат</th></tr>
{mismatch_rows or "<tr><td colspan='4' class='ok'>Нет</td></tr>"}</table>
<h2>Значения эталона, которых ещё нет в контракте</h2>
<table><tr><th>Книга</th><th>Категория</th><th>Ячейка</th><th>Эталон</th></tr>
{gap_rows or "<tr><td colspan='4' class='ok'>Нет</td></tr>"}</table>
<h2>Проблемные документы качества в самом эталоне</h2>
<table><tr><th>Работа</th><th>Материал</th><th>Документ</th></tr>
{placeholder_rows or "<tr><td colspan='3' class='ok'>Не найдены</td></tr>"}</table>
<h2>Смысловые блокеры, которые были пропущены только для теста</h2>
<table><tr><th>Уровень</th><th>Код</th><th>Сообщение</th></tr>
{semantic_rows or "<tr><td colspan='3' class='ok'>Нет</td></tr>"}</table>
</body></html>"""


def main() -> int:
    stamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    run_root = settings.runs_dir / f"project1-golden-parity-{stamp}"
    xlsx_dir = run_root / "output" / "xlsx"
    preview_dir = run_root / "preview"
    report_dir = run_root / "report"
    for directory in (xlsx_dir, preview_dir, report_dir):
        directory.mkdir(parents=True, exist_ok=True)

    generator = ExcelGenerator(settings.root, settings.contracts_dir, settings.approved_templates_dir)
    all_claims: list[Claim] = []
    all_work_items: list[WorkItem] = []
    all_plans: list[DocumentPlan] = []
    artifact_categories: dict[str, str] = {}
    results: list[dict[str, Any]] = []
    sequence = 1

    for spec in SPECS:
        contract = generator.contract(spec.template_id)
        golden_reader = WorkbookReader(spec.golden)
        try:
            claims, work_items, plan, common_values, sequence = build_fixture(
                spec, contract, golden_reader, sequence
            )
            output, _ = generator.generate(plan, work_items, claims, xlsx_dir)
            baseline_path = generator.template_path(contract)
            baseline_reader = WorkbookReader(baseline_path)
            generated_reader = WorkbookReader(output)
            try:
                comparison = mapped_comparison(contract, plan, golden_reader, generated_reader)
                fill_comparison = mapped_fill_comparison(
                    contract,
                    plan,
                    golden_reader,
                    generated_reader,
                )
                gaps = contract_gaps(
                    spec,
                    contract,
                    baseline_reader,
                    golden_reader,
                    generated_reader,
                )
                outside = non_whitelist_changes(contract, baseline_reader, generated_reader)
            finally:
                baseline_reader.close()
                generated_reader.close()
            technical = validate_workbook(output, baseline_path, contract, plan, claims)
            previews, render_issues = render_selected_sheets(
                output,
                plan.selected_sheets,
                preview_dir / spec.template_id,
                settings.soffice_path,
            )
            technical.extend(render_issues)
            results.append(
                {
                    "family": spec.family,
                    "template_id": spec.template_id,
                    "golden": str(spec.golden.relative_to(ROOT)),
                    "golden_sha256": sha256(spec.golden),
                    "source_template": str(baseline_path.relative_to(ROOT)),
                    "source_template_sha256": sha256(baseline_path),
                    "generated": str(output.relative_to(run_root)),
                    "generated_sha256": sha256(output),
                    "selected_sheets": plan.selected_sheets,
                    "common_fixture_values": common_values,
                    "mapped_comparison": comparison,
                    "mapped_fill_comparison": fill_comparison,
                    "contract_gaps": gaps,
                    "contract_gap_counts": dict(Counter(item["category"] for item in gaps)),
                    "non_whitelist_changes": outside,
                    "technical_validation": [issue.model_dump(mode="json") for issue in technical],
                    "technical_issue_counts": dict(Counter(issue.code for issue in technical)),
                    "preview_count": len(previews),
                    "preview_files": [str(path.relative_to(run_root)) for path in previews],
                }
            )
            all_claims.extend(claims)
            all_work_items.extend(work_items)
            all_plans.append(plan)
            artifact_categories.update({artifact_id: "execution_scheme" for artifact_id in plan.attachments})
            for item in work_items:
                for material in item.materials:
                    if material.source_file_id:
                        artifact_categories[material.source_file_id] = "certificate"
        finally:
            golden_reader.close()

    semantic = validate_semantics(
        all_work_items,
        all_claims,
        all_plans,
        artifact_categories=artifact_categories,
    )
    report = {
        "diagnostic_only": True,
        "fixture_parity": True,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "notice": (
            "NEEDS_INPUT пропущен только путём использования заполненных XLSX project1 "
            "как тестового источника. Рабочие правила приложения не изменены."
        ),
        "model_calls": 0,
        "run_root": str(run_root),
        "workbooks": results,
        "claims": [claim.model_dump(mode="json") for claim in all_claims],
        "work_items": [item.model_dump(mode="json") for item in all_work_items],
        "document_plans": [plan.model_dump(mode="json") for plan in all_plans],
        "semantic_validation": [issue.model_dump(mode="json") for issue in semantic],
        "semantic_issue_counts": dict(Counter(issue.code for issue in semantic)),
        "placeholder_quality_documents": placeholder_quality_documents(all_work_items),
        "summary": {
            "workbook_count": len(results),
            "act_count": sum(len(plan.selected_sheets) for plan in all_plans),
            "mapped_total": sum(item["mapped_comparison"]["total"] for item in results),
            "mapped_matches": sum(item["mapped_comparison"]["matches"] for item in results),
            "mapped_mismatches": sum(
                len(item["mapped_comparison"]["mismatches"]) for item in results
            ),
            "mapped_fill_mismatches": sum(
                len(item["mapped_fill_comparison"]["mismatches"]) for item in results
            ),
            "contract_gaps": sum(len(item["contract_gaps"]) for item in results),
            "non_whitelist_changes": sum(
                len(item["non_whitelist_changes"]) for item in results
            ),
            "technical_errors_excluding_unapproved": sum(
                issue["severity"] == "error" and issue["code"] != "TEMPLATE_NOT_APPROVED"
                for item in results
                for issue in item["technical_validation"]
            ),
            "preview_count": sum(item["preview_count"] for item in results),
        },
    }
    report_json = report_dir / "comparison.json"
    report_json.write_text(
        json.dumps(report, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    report_path = report_dir / "comparison.html"
    report_path.write_text(report_html(report), encoding="utf-8")

    zip_path = run_root / "project1-golden-parity.zip"
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as archive:
        for folder, prefix in (
            (xlsx_dir, "xlsx"),
            (preview_dir, "preview"),
            (report_dir, "report"),
        ):
            for path in sorted(folder.rglob("*")):
                if path.is_file():
                    archive.write(path, f"{prefix}/{path.relative_to(folder).as_posix()}")

    print(f"run_root={run_root}")
    print(f"report_html={report_path}")
    print(f"report_json={report_json}")
    print(f"result_zip={zip_path}")
    print(json.dumps(report["summary"], ensure_ascii=False, sort_keys=True))
    print(f"semantic_issue_counts={json.dumps(report['semantic_issue_counts'], ensure_ascii=False, sort_keys=True)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
