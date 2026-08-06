from __future__ import annotations

import copy
import hashlib
import re
import zipfile
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path

import openpyxl
import yaml
from lxml import etree

from .domain import (
    TemplateCellAssignment,
    TemplateUnresolvedFinding,
    UnresolvedTemplateCell,
    ValidationIssue,
)
from .excel import MAIN_NS, OOXMLWorkbook, sha256


TEMPLATE_ID = re.compile(r"[a-z][a-z0-9_-]{1,63}")
CELL_ADDRESS = re.compile(r"[A-Z]{1,3}[1-9][0-9]*")
SEMANTIC_ID = re.compile(r"[a-z][a-z0-9_.-]{2,127}")
WARNING_FILL_RGB = "FFFFE699"


@dataclass(frozen=True)
class TemplateField:
    sheet: str
    cell: str
    label: str
    semantic_id: str | None
    description: str
    evidence_rule: str
    value_kind: str
    required: bool
    manual_reason: str | None
    value_pattern: str | None

    @property
    def coordinate(self) -> tuple[str, str]:
        return self.sheet, self.cell

    @classmethod
    def load(cls, data: dict) -> "TemplateField":
        sheet = str(data["sheet"])
        cell = str(data["cell"]).upper()
        if not sheet or not CELL_ADDRESS.fullmatch(cell):
            raise ValueError(f"Некорректная цель шаблона: {sheet}!{cell}")
        semantic_id = str(data["semantic_id"]) if data.get("semantic_id") else None
        if semantic_id and not SEMANTIC_ID.fullmatch(semantic_id):
            raise ValueError(f"Некорректный semantic_id: {semantic_id}")
        value_pattern = str(data["value_pattern"]) if data.get("value_pattern") else None
        if value_pattern:
            try:
                re.compile(value_pattern)
            except re.error as exc:
                raise ValueError(
                    f"Некорректный value_pattern: {sheet}!{cell}: {exc}"
                ) from exc
        label = str(data.get("label") or f"{sheet}!{cell}")
        return cls(
            sheet=sheet,
            cell=cell,
            label=label,
            semantic_id=semantic_id,
            description=str(data.get("description") or label),
            evidence_rule=str(data.get("evidence_rule") or "direct_pdf"),
            value_kind=str(data.get("value_kind") or "text"),
            required=bool(data.get("required", True)),
            manual_reason=str(data["manual_reason"]) if data.get("manual_reason") else None,
            value_pattern=value_pattern,
        )

    def validate_raw_value(self, raw_value: str) -> None:
        if self.value_pattern and not re.fullmatch(
            self.value_pattern,
            raw_value.strip(),
        ):
            raise ValueError(
                f"Значение не соответствует смыслу поля {self.label}: "
                f"{self.sheet}!{self.cell}"
            )


@dataclass(frozen=True)
class SelectedTemplateContract:
    template_id: str
    display_name: str
    document_kind: str
    version: str
    status: str
    approved: bool
    candidate_template: str
    candidate_sha256: str
    source_sha256: str
    etalon_sha256: str
    output_filename: str
    warning_fill_rgb: str
    fields: tuple[TemplateField, ...]
    structural_findings: dict
    path: Path
    contract_sha256: str

    @classmethod
    def load(cls, path: Path) -> "SelectedTemplateContract":
        data = yaml.safe_load(path.read_text(encoding="utf-8"))
        template_id = str(data["template_id"])
        if not TEMPLATE_ID.fullmatch(template_id):
            raise ValueError(f"Некорректный template_id: {template_id}")
        fields = tuple(TemplateField.load(item) for item in data.get("fields", []))
        coordinates = [field.coordinate for field in fields]
        if not fields or len(coordinates) != len(set(coordinates)):
            raise ValueError(f"Поля шаблона пусты или содержат повторы: {template_id}")
        output_filename = Path(str(data["output_filename"])).name
        if output_filename != str(data["output_filename"]) or Path(output_filename).suffix.lower() != ".xlsx":
            raise ValueError(f"Некорректное имя результата: {template_id}")
        warning_fill_rgb = str(data.get("warning_fill_rgb") or WARNING_FILL_RGB).upper()
        if not re.fullmatch(r"[0-9A-F]{8}", warning_fill_rgb):
            raise ValueError(f"Некорректный warning_fill_rgb: {template_id}")
        digests = {
            "candidate_sha256": str(data["candidate_sha256"]),
            "source_sha256": str(data["source_sha256"]),
            "etalon_sha256": str(data["etalon_sha256"]),
        }
        for key, value in digests.items():
            if not re.fullmatch(r"[0-9a-f]{64}", value):
                raise ValueError(f"Некорректный {key}: {template_id}")
        return cls(
            template_id=template_id,
            display_name=str(data["display_name"]),
            document_kind=str(data.get("document_kind") or template_id),
            version=str(data["version"]),
            status=str(data.get("status") or "DISCOVERY_REVIEW_REQUIRED"),
            approved=bool(data.get("approved", False)),
            candidate_template=str(data["candidate_template"]),
            candidate_sha256=digests["candidate_sha256"],
            source_sha256=digests["source_sha256"],
            etalon_sha256=digests["etalon_sha256"],
            output_filename=output_filename,
            warning_fill_rgb=warning_fill_rgb,
            fields=fields,
            structural_findings=dict(data.get("structural_findings") or {}),
            path=path,
            contract_sha256=sha256(path),
        )

    @property
    def field_map(self) -> dict[tuple[str, str], TemplateField]:
        return {field.coordinate: field for field in self.fields}

    def public_metadata(self) -> dict:
        manual_count = sum(bool(field.manual_reason) for field in self.fields)
        return {
            "id": self.template_id,
            "display_name": self.display_name,
            "document_kind": self.document_kind,
            "status": self.status,
            "approved": self.approved,
            "target_count": len(self.fields),
            "manual_count": manual_count,
            "warning": (
                None
                if self.approved
                else (
                    "Контракт обнаружен автоматически и доступен только для проверочного "
                    "черновика. Итоговый выпуск заблокирован до проверки и утверждения специалистом."
                )
            ),
        }

    def model_fields(self) -> list[dict]:
        return [
            {
                "sheet": field.sheet,
                "cell": field.cell,
                "label": field.label,
                "semantic_id": field.semantic_id,
                "description": field.description,
                "evidence_rule": field.evidence_rule,
                "value_kind": field.value_kind,
            }
            for field in self.fields
            if not field.manual_reason
        ]


class TemplateCatalog:
    def __init__(self, root: Path, contracts_dir: Path, approved_dir: Path):
        self.root = root.resolve()
        self.contracts_dir = contracts_dir
        self.approved_dir = approved_dir.resolve()
        self._contracts: dict[str, SelectedTemplateContract] = {}
        self.reload()

    def reload(self) -> None:
        contracts: dict[str, SelectedTemplateContract] = {}
        if self.contracts_dir.exists():
            for path in sorted(self.contracts_dir.glob("*.yaml")):
                contract = SelectedTemplateContract.load(path)
                if contract.template_id in contracts:
                    raise ValueError(f"Повтор template_id: {contract.template_id}")
                candidate = self.candidate_path(contract)
                if not candidate.is_file():
                    raise ValueError(f"Кандидат шаблона отсутствует: {candidate}")
                if sha256(candidate) != contract.candidate_sha256:
                    raise ValueError(f"SHA-256 кандидата не совпадает: {contract.template_id}")
                contracts[contract.template_id] = contract
        self._contracts = contracts

    def candidate_path(self, contract: SelectedTemplateContract) -> Path:
        candidate = (self.root / contract.candidate_template).resolve()
        if self.approved_dir != candidate.parent and self.approved_dir not in candidate.parents:
            raise ValueError(f"Кандидат шаблона вне templates/approved: {contract.template_id}")
        return candidate

    def list(self) -> list[SelectedTemplateContract]:
        return list(self._contracts.values())

    def public_list(self) -> list[dict]:
        return [contract.public_metadata() for contract in self.list()]

    def get(self, template_id: str) -> SelectedTemplateContract:
        if not TEMPLATE_ID.fullmatch(template_id):
            raise KeyError(template_id)
        try:
            contract = self._contracts[template_id]
        except KeyError:
            raise KeyError(template_id) from None
        self.assert_current(contract)
        return contract

    def assert_current(self, contract: SelectedTemplateContract) -> None:
        self.assert_contract_current(contract)
        candidate = self.candidate_path(contract)
        if sha256(candidate) != contract.candidate_sha256:
            raise ValueError(
                f"Кандидат шаблона изменился после загрузки каталога: {contract.template_id}"
            )

    def assert_contract_current(self, contract: SelectedTemplateContract) -> None:
        cached = self._contracts.get(contract.template_id)
        if cached is None or cached.contract_sha256 != contract.contract_sha256:
            raise ValueError(
                f"Контракт шаблона изменился после загрузки каталога: {contract.template_id}"
            )
        if sha256(contract.path) != contract.contract_sha256:
            raise ValueError(
                f"Контракт шаблона изменился после загрузки каталога: {contract.template_id}"
            )


class SelectedTemplateGenerator:
    def __init__(self, catalog: TemplateCatalog):
        self.catalog = catalog

    @staticmethod
    def _assignment_map(
        contract: SelectedTemplateContract,
        assignments: list[TemplateCellAssignment],
    ) -> dict[tuple[str, str], TemplateCellAssignment]:
        field_map = contract.field_map
        assignment_map: dict[tuple[str, str], TemplateCellAssignment] = {}
        for assignment in assignments:
            coordinate = assignment.sheet, assignment.cell.upper()
            field = field_map.get(coordinate)
            if field is None:
                raise ValueError(f"Модель попыталась записать незарегистрированную ячейку: {assignment.sheet}!{assignment.cell}")
            if field.manual_reason:
                raise ValueError(f"Поле требует ручного подтверждения: {assignment.sheet}!{assignment.cell}")
            if coordinate in assignment_map:
                raise ValueError(f"Повтор записи: {assignment.sheet}!{assignment.cell}")
            if not assignment.value.strip() or assignment.value.lstrip().startswith("="):
                raise ValueError(f"Пустое значение или формула запрещены: {assignment.sheet}!{assignment.cell}")
            assignment_map[coordinate] = assignment
        return assignment_map

    @classmethod
    def unresolved_cells(
        cls,
        contract: SelectedTemplateContract,
        assignments: list[TemplateCellAssignment],
        findings: list[TemplateUnresolvedFinding] | None = None,
    ) -> list[UnresolvedTemplateCell]:
        assignment_map = cls._assignment_map(contract, assignments)
        finding_map: dict[
            tuple[str, str],
            TemplateUnresolvedFinding,
        ] = {}
        for finding in findings or []:
            coordinate = finding.sheet, finding.cell.upper()
            field = contract.field_map.get(coordinate)
            if field is None:
                raise ValueError(
                    "Модель вернула причину для незарегистрированной ячейки: "
                    f"{finding.sheet}!{finding.cell}"
                )
            if field.manual_reason:
                raise ValueError(
                    "Причина ручного поля определяется серверным контрактом: "
                    f"{finding.sheet}!{finding.cell}"
                )
            if coordinate in assignment_map:
                raise ValueError(
                    "Заполненная ячейка не может одновременно быть нерешённой: "
                    f"{finding.sheet}!{finding.cell}"
                )
            if coordinate in finding_map:
                raise ValueError(
                    f"Повтор причины: {finding.sheet}!{finding.cell}"
                )
            finding_map[coordinate] = finding
        unresolved: list[UnresolvedTemplateCell] = []
        for field in contract.fields:
            if field.coordinate in assignment_map:
                continue
            finding = finding_map.get(field.coordinate)
            reason = (
                field.manual_reason
                or (finding.reason if finding else None)
                or "В PDF нет надёжно подтверждённого значения"
            )
            category = (
                "manual_confirmation"
                if field.manual_reason
                else finding.category if finding else "missing_from_pdf"
            )
            unresolved.append(
                UnresolvedTemplateCell(
                    template_id=contract.template_id,
                    sheet=field.sheet,
                    cell=field.cell,
                    label=field.label,
                    reason=reason,
                    category=category,
                    required=field.required,
                    blocking=field.required,
                    source_locators=(
                        list(finding.source_locators)
                        if finding
                        else []
                    ),
                )
            )
        return unresolved

    def generate(
        self,
        contract: SelectedTemplateContract,
        assignments: list[TemplateCellAssignment],
        destination: Path,
        *,
        findings: list[TemplateUnresolvedFinding] | None = None,
        source_snapshot: Path | None = None,
    ) -> tuple[Path, list[UnresolvedTemplateCell]]:
        assignment_map = self._assignment_map(contract, assignments)

        self.catalog.assert_contract_current(contract)
        source = self.catalog.candidate_path(contract)
        source_bytes = source.read_bytes()
        if hashlib.sha256(source_bytes).hexdigest() != contract.candidate_sha256:
            raise ValueError(
                f"Кандидат шаблона изменился перед генерацией: {contract.template_id}"
            )
        if source_snapshot is not None:
            source_snapshot.parent.mkdir(parents=True, exist_ok=True)
            source_snapshot.write_bytes(source_bytes)
        workbook = OOXMLWorkbook(source_bytes)
        for coordinate, assignment in assignment_map.items():
            field = contract.field_map[coordinate]
            workbook.set_cell(
                assignment.sheet,
                assignment.cell.upper(),
                self._typed_value(field, assignment.value),
            )

        unresolved = self.unresolved_cells(contract, assignments, findings)
        warning_targets: dict[str, list[str]] = {}
        for item in unresolved:
            warning_targets.setdefault(item.sheet, []).append(item.cell)
        if warning_targets:
            workbook.highlight_cells(warning_targets, rgb=contract.warning_fill_rgb)
        workbook.enable_full_calculation()
        destination.mkdir(parents=True, exist_ok=True)
        output = destination / f"ЧЕРНОВИК - {contract.output_filename}"
        workbook.save(output)
        return output, unresolved

    @staticmethod
    def _typed_value(field: TemplateField, raw_value: str) -> str | int | float:
        field.validate_raw_value(raw_value)
        if field.value_kind == "text":
            return raw_value
        if field.value_kind == "date":
            raise ValueError(
                f"Дата может быть внесена только после отдельного подтверждения: "
                f"{field.sheet}!{field.cell}"
            )
        if field.value_kind != "number":
            raise ValueError(
                f"Неизвестный тип значения {field.value_kind}: {field.sheet}!{field.cell}"
            )
        normalized = raw_value.strip().replace("\u00a0", "").replace(" ", "").replace(",", ".")
        if not re.fullmatch(r"-?(?:0|[1-9][0-9]*)(?:\.[0-9]+)?", normalized):
            raise ValueError(
                f"Ожидалось числовое значение: {field.sheet}!{field.cell}"
            )
        try:
            value = Decimal(normalized)
        except InvalidOperation as exc:
            raise ValueError(
                f"Ожидалось числовое значение: {field.sheet}!{field.cell}"
            ) from exc
        return int(value) if value == value.to_integral_value() else float(value)


def validate_selected_template_output(
    output: Path,
    source: Path,
    contract: SelectedTemplateContract,
    assignments: list[TemplateCellAssignment],
    unresolved: list[UnresolvedTemplateCell],
) -> list[ValidationIssue]:
    issues: list[ValidationIssue] = []
    if sha256(source) != contract.candidate_sha256:
        return [
            ValidationIssue(
                code="TEMPLATE_SHA_CHANGED",
                severity="error",
                message="Кандидат шаблона изменился после закрепления задания",
                artifact=output.name,
            )
        ]
    try:
        with zipfile.ZipFile(output) as archive:
            if archive.testzip():
                raise ValueError("повреждённый OOXML ZIP")
            if "xl/workbook.xml" not in archive.namelist():
                raise ValueError("xl/workbook.xml отсутствует")
    except Exception as exc:
        return [
            ValidationIssue(
                code="INVALID_XLSX",
                severity="error",
                message=f"Результат не является корректной книгой XLSX: {exc}",
                artifact=output.name,
            )
        ]

    source_book = openpyxl.load_workbook(source, data_only=False, keep_links=True)
    output_book = openpyxl.load_workbook(output, data_only=False, keep_links=True)
    try:
        if source_book.sheetnames != output_book.sheetnames:
            issues.append(
                ValidationIssue(
                    code="SHEET_STRUCTURE",
                    severity="error",
                    message="Порядок или состав листов отличается от выбранного шаблона",
                    artifact=output.name,
                )
            )
        source_visibility = {
            sheet.title: sheet.sheet_state
            for sheet in source_book.worksheets
        }
        output_visibility = {
            sheet.title: sheet.sheet_state
            for sheet in output_book.worksheets
        }
        if source_visibility != output_visibility:
            issues.append(
                ValidationIssue(
                    code="SHEET_VISIBILITY_CHANGED",
                    severity="error",
                    message="Видимость листов отличается от выбранного шаблона",
                    artifact=output.name,
                )
            )
        assignment_map = {
            (item.sheet, item.cell.upper()): SelectedTemplateGenerator._typed_value(
                contract.field_map[(item.sheet, item.cell.upper())],
                item.value,
            )
            for item in assignments
        }
        unresolved_set = {(item.sheet, item.cell.upper()) for item in unresolved}
        expected_unresolved = set(contract.field_map) - set(assignment_map)
        if unresolved_set != expected_unresolved:
            issues.append(
                ValidationIssue(
                    code="UNRESOLVED_REGISTER_MISMATCH",
                    severity="error",
                    message="Реестр незаполненных ячеек не совпадает с контрактом шаблона",
                    artifact=output.name,
                )
            )
        for sheet_name in set(source_book.sheetnames) & set(output_book.sheetnames):
            before = source_book[sheet_name]
            after = output_book[sheet_name]
            max_row = max(before.max_row, after.max_row)
            max_column = max(before.max_column, after.max_column)
            for row in range(1, max_row + 1):
                for column in range(1, max_column + 1):
                    source_cell = before.cell(row, column)
                    output_cell = after.cell(row, column)
                    coordinate = (sheet_name, source_cell.coordinate)
                    if source_cell.data_type == "f" or output_cell.data_type == "f":
                        if source_cell.value != output_cell.value:
                            issues.append(
                                ValidationIssue(
                                    code="FORMULA_CHANGED",
                                    severity="error",
                                    message="Формула изменилась вне разрешённого контракта",
                                    artifact=output.name,
                                    locator=f"{sheet_name}!{source_cell.coordinate}",
                                )
                            )
                    elif source_cell.value != output_cell.value:
                        expected = assignment_map.get(coordinate)
                        if expected is None or output_cell.value != expected:
                            issues.append(
                                ValidationIssue(
                                    code="UNAUTHORIZED_CELL_CHANGE",
                                    severity="error",
                                    message="Изменена незарегистрированная ячейка",
                                    artifact=output.name,
                                    locator=f"{sheet_name}!{source_cell.coordinate}",
                                )
                            )
                    default_style = (0, 0, 0, 0, 0, 0, 0, 0, 0)
                    source_style = list(source_cell._style or default_style)
                    output_style = list(output_cell._style or default_style)
                    if coordinate in unresolved_set and len(source_style) == len(output_style):
                        # StyleArray index 1 is fillId. Every other style component
                        # must remain byte-for-byte equivalent.
                        output_style[1] = source_style[1]
                    if source_style != output_style:
                        issues.append(
                            ValidationIssue(
                                code="UNAUTHORIZED_STYLE_CHANGE",
                                severity="error",
                                message="Помимо разрешённой заливки изменён стиль ячейки",
                                artifact=output.name,
                                locator=f"{sheet_name}!{source_cell.coordinate}",
                            )
                        )
        for sheet_name, cell in unresolved_set:
            target = output_book[sheet_name][cell]
            color = (target.fill.fgColor.rgb or "").upper()
            if target.value not in (None, ""):
                issues.append(
                    ValidationIssue(
                        code="UNRESOLVED_CELL_NOT_EMPTY",
                        severity="error",
                        message="Неподтверждённая ячейка должна оставаться пустой",
                        artifact=output.name,
                        locator=f"{sheet_name}!{cell}",
                    )
                )
            if target.fill.fill_type != "solid" or color != contract.warning_fill_rgb:
                issues.append(
                    ValidationIssue(
                        code="UNRESOLVED_CELL_NOT_HIGHLIGHTED",
                        severity="error",
                        message="Неподтверждённая ячейка не выделена предупреждающим цветом",
                        artifact=output.name,
                        locator=f"{sheet_name}!{cell}",
                    )
                )
        if len(getattr(source_book, "_external_links", [])) != len(
            getattr(output_book, "_external_links", [])
        ):
            issues.append(
                ValidationIssue(
                    code="EXTERNAL_LINKS_CHANGED",
                    severity="error",
                    message="Набор внешних связей изменился при заполнении",
                    artifact=output.name,
                )
            )
    finally:
        source_book.close()
        output_book.close()

    try:
        source_structure = _package_structure(source)
        output_structure = _package_structure(output)
        if source_structure["workbook"] != output_structure["workbook"]:
            issues.append(
                ValidationIssue(
                    code="WORKBOOK_STRUCTURE_CHANGED",
                    severity="error",
                    message=(
                        "Изменены свойства книги, именованные диапазоны "
                        "или параметры листов вне контракта"
                    ),
                    artifact=output.name,
                )
            )
        for sheet_name in sorted(
            set(source_structure["worksheets"])
            | set(output_structure["worksheets"])
        ):
            if source_structure["worksheets"].get(
                sheet_name
            ) == output_structure["worksheets"].get(sheet_name):
                continue
            issues.append(
                ValidationIssue(
                    code="WORKSHEET_STRUCTURE_CHANGED",
                    severity="error",
                    message=(
                        "Изменены объединения, размеры строк/столбцов, "
                        "гиперссылки или параметры печати листа"
                    ),
                    artifact=output.name,
                    locator=sheet_name,
                )
            )
        styles_preserved, style_message = _style_definitions_preserved(
            source,
            output,
            warning_rgb=contract.warning_fill_rgb,
            unresolved=unresolved_set,
        )
        if not styles_preserved:
            issues.append(
                ValidationIssue(
                    code="STYLE_DEFINITIONS_CHANGED",
                    severity="error",
                    message=style_message,
                    artifact=output.name,
                )
            )
        package_preserved, package_message = _immutable_package_parts_preserved(
            source,
            output,
        )
        if not package_preserved:
            issues.append(
                ValidationIssue(
                    code="IMMUTABLE_PACKAGE_PART_CHANGED",
                    severity="error",
                    message=package_message,
                    artifact=output.name,
                )
            )
    except Exception as exc:
        issues.append(
            ValidationIssue(
                code="STRUCTURE_CHECK_FAILED",
                severity="error",
                message=f"Не удалось сравнить структуру книги: {exc}",
                artifact=output.name,
            )
        )

    if not contract.approved:
        issues.append(
            ValidationIssue(
                code="TEMPLATE_NOT_APPROVED",
                severity="info",
                message=(
                    f"Шаблон остаётся кандидатом со статусом {contract.status}. "
                    "Черновик доступен, финальный выпуск требует отдельного утверждения специалистом."
                ),
                artifact=output.name,
            )
        )
    formula_errors = int(contract.structural_findings.get("formula_errors_observed", 0))
    if formula_errors:
        issues.append(
            ValidationIssue(
                code="TEMPLATE_FORMULA_ERRORS",
                severity="error",
                message=f"В исходном кандидате обнаружены формулы с ошибками: {formula_errors}",
                artifact=output.name,
            )
        )
    external_links = int(
        contract.structural_findings.get(
            "candidate_external_links",
            contract.structural_findings.get("source_external_links", 0),
        )
    )
    if external_links:
        issues.append(
            ValidationIssue(
                code="TEMPLATE_EXTERNAL_LINKS",
                severity="error",
                message=f"В кандидате остаются внешние связи: {external_links}",
                artifact=output.name,
            )
        )
    remaining_external_formula_references = int(
        contract.structural_findings.get(
            "remaining_external_formula_reference_count",
            0,
        )
    )
    if remaining_external_formula_references:
        issues.append(
            ValidationIssue(
                code="TEMPLATE_EXTERNAL_FORMULAS",
                severity="error",
                message=(
                    "После удаления внешних кэшей остались формулы со ссылками "
                    f"на другие книги: {remaining_external_formula_references}"
                ),
                artifact=output.name,
            )
        )
    raw_ref_errors = int(
        contract.structural_findings.get("raw_ref_error_count", 0)
    )
    if raw_ref_errors:
        issues.append(
            ValidationIssue(
                code="TEMPLATE_RAW_REF_ERRORS",
                severity="error",
                message=(
                    "В OOXML кандидата обнаружены ссылки #REF! в формулах, "
                    f"именованных диапазонах или правилах проверки: {raw_ref_errors}"
                ),
                artifact=output.name,
            )
        )
    unsafe_blank_formulas = int(
        contract.structural_findings.get("unsafe_blank_formula_count", 0)
    )
    if unsafe_blank_formulas:
        issues.append(
            ValidationIssue(
                code="TEMPLATE_FALSE_BLANK_FORMULAS",
                severity="error",
                message=(
                    "В кандидате остались формулы, которые превращают пустые "
                    f"исходные ячейки в ложный ноль или текст: {unsafe_blank_formulas}"
                ),
                artifact=output.name,
            )
        )
    remaining_sensitive = int(
        contract.structural_findings.get("remaining_sensitive_value_count", 0)
    )
    if remaining_sensitive:
        issues.append(
            ValidationIssue(
                code="TEMPLATE_SENSITIVE_VALUES",
                severity="error",
                message=(
                    "После автоматической очистки остались значения, требующие "
                    f"проверки специалистом: {remaining_sensitive}"
                ),
                artifact=output.name,
            )
        )
    package_sensitive = int(
        contract.structural_findings.get("package_forbidden_token_count", 0)
    )
    if package_sensitive:
        issues.append(
            ValidationIssue(
                code="TEMPLATE_PACKAGE_SENSITIVE_VALUES",
                severity="error",
                message=(
                    "Внутри OOXML-пакета остались маркеры данных прежнего "
                    f"проекта: {package_sensitive}"
                ),
                artifact=output.name,
            )
        )
    formula_differences = int(
        contract.structural_findings.get(
            "unreviewed_formula_difference_count",
            contract.structural_findings.get(
                "candidate_vs_etalon_formula_differences",
                contract.structural_findings.get(
                    "source_vs_etalon_formula_differences",
                    contract.structural_findings.get("formula_differences", 0),
                ),
            ),
        )
    )
    if formula_differences:
        issues.append(
            ValidationIssue(
                code="TEMPLATE_FORMULA_FINDINGS",
                severity="info",
                message=f"В сравнении кандидата с ETALON зафиксированы различия формул: {formula_differences}",
                artifact=output.name,
            )
        )
    return issues


def _style_definitions_preserved(
    source: Path,
    output: Path,
    *,
    warning_rgb: str,
    unresolved: set[tuple[str, str]],
) -> tuple[bool, str]:
    """Allow only warning-fill styles derived from original cellXfs."""

    parser = etree.XMLParser(remove_blank_text=True)
    source_package = OOXMLWorkbook(source)
    output_package = OOXMLWorkbook(output)
    source_styles = etree.fromstring(
        source_package.parts["xl/styles.xml"],
        parser,
    )
    output_styles = etree.fromstring(
        output_package.parts["xl/styles.xml"],
        parser,
    )

    def canonical(element: etree._Element) -> bytes:
        return etree.tostring(element, method="c14n", with_comments=False)

    def without_mutable_tables(root: etree._Element) -> bytes:
        normalized = copy.deepcopy(root)
        for local_name in ("fills", "cellXfs"):
            element = normalized.find(f"{{{MAIN_NS}}}{local_name}")
            if element is not None:
                normalized.remove(element)
        return canonical(normalized)

    if without_mutable_tables(source_styles) != without_mutable_tables(
        output_styles
    ):
        return (
            False,
            "Изменены определения шрифтов, границ, числовых форматов или другие базовые стили",
        )

    source_fills = source_styles.find(f"{{{MAIN_NS}}}fills")
    output_fills = output_styles.find(f"{{{MAIN_NS}}}fills")
    source_xfs = source_styles.find(f"{{{MAIN_NS}}}cellXfs")
    output_xfs = output_styles.find(f"{{{MAIN_NS}}}cellXfs")
    if any(
        element is None
        for element in (source_fills, output_fills, source_xfs, output_xfs)
    ):
        return False, "В книге отсутствуют обязательные таблицы fills/cellXfs"
    assert source_fills is not None
    assert output_fills is not None
    assert source_xfs is not None
    assert output_xfs is not None

    def declared_count_matches(element: etree._Element) -> bool:
        declared = element.get("count")
        return declared is None or (
            declared.isdigit() and int(declared) == len(element)
        )

    if not all(
        declared_count_matches(element)
        for element in (source_fills, output_fills, source_xfs, output_xfs)
    ):
        return False, "Счётчики таблиц стилей не соответствуют их содержимому"

    for index, source_fill in enumerate(source_fills):
        if index >= len(output_fills) or canonical(source_fill) != canonical(
            output_fills[index]
        ):
            return False, "Изменено исходное определение заливки"

    def is_warning_fill(fill: etree._Element) -> bool:
        pattern = fill.find(f"{{{MAIN_NS}}}patternFill")
        if (
            pattern is None
            or pattern.get("patternType") != "solid"
            or len(pattern) != 2
        ):
            return False
        foreground = pattern.find(f"{{{MAIN_NS}}}fgColor")
        background = pattern.find(f"{{{MAIN_NS}}}bgColor")
        return (
            foreground is not None
            and (foreground.get("rgb") or "").upper() == warning_rgb.upper()
            and background is not None
            and background.get("indexed") == "64"
        )

    warning_fill_id = next(
        (
            index
            for index, fill in enumerate(source_fills)
            if is_warning_fill(fill)
        ),
        None,
    )
    expected_fill_count = len(source_fills)
    if warning_fill_id is None and unresolved:
        expected_fill_count += 1
        warning_fill_id = len(source_fills)
        if (
            len(output_fills) != expected_fill_count
            or not is_warning_fill(output_fills[-1])
        ):
            return False, "Добавлена заливка, не соответствующая предупреждающему цвету"
    elif len(output_fills) != expected_fill_count:
        return False, "Набор заливок изменён вне разрешённого предупреждения"
    if warning_fill_id is None:
        if unresolved:
            return False, "Не найдено определение предупреждающей заливки"
        warning_fill_id = -1

    for index, source_xf in enumerate(source_xfs):
        if index >= len(output_xfs) or canonical(source_xf) != canonical(
            output_xfs[index]
        ):
            return False, "Изменено исходное определение стиля ячейки"

    def xf_signature(style: etree._Element) -> bytes:
        normalized = copy.deepcopy(style)
        normalized.attrib.pop("fillId", None)
        normalized.attrib.pop("applyFill", None)
        return canonical(normalized)

    source_signatures = {
        xf_signature(style)
        for style in source_xfs
    }
    for style in list(output_xfs)[len(source_xfs):]:
        if (
            style.get("fillId") != str(warning_fill_id)
            or style.get("applyFill") != "1"
            or xf_signature(style) not in source_signatures
        ):
            return False, "Добавлен стиль, не являющийся копией исходного с предупреждающей заливкой"

    expected_extra_ids = set(range(len(source_xfs), len(output_xfs)))
    referenced_extra_ids: set[int] = set()
    for sheet_name, coordinate in unresolved:
        root = output_package._sheet_root(sheet_name)
        cells = root.xpath(
            f"//x:c[@r='{coordinate}']",
            namespaces={"x": MAIN_NS},
        )
        if len(cells) != 1:
            return False, f"Не найдена единственная целевая ячейка {sheet_name}!{coordinate}"
        style_id = int(cells[0].get("s", "0"))
        if style_id >= len(source_xfs):
            referenced_extra_ids.add(style_id)
    if referenced_extra_ids != expected_extra_ids:
        return False, "Добавленные предупреждающие стили не совпадают с целевыми ячейками"
    return True, ""


def _immutable_package_parts_preserved(
    source: Path,
    output: Path,
) -> tuple[bool, str]:
    """Pin every OOXML part not intentionally changed by selected-template fill."""

    source_package = OOXMLWorkbook(source)
    output_package = OOXMLWorkbook(output)
    mutable_parts = {
        "xl/workbook.xml",
        "xl/styles.xml",
        *source_package.sheet_parts.values(),
        *output_package.sheet_parts.values(),
    }
    source_names = set(source_package.parts) - mutable_parts
    output_names = set(output_package.parts) - mutable_parts
    if source_names != output_names:
        added = sorted(output_names - source_names)
        removed = sorted(source_names - output_names)
        return (
            False,
            "Изменён состав неизменяемых частей XLSX"
            f" (добавлены: {added[:3] or 'нет'}; удалены: {removed[:3] or 'нет'})",
        )
    changed = sorted(
        name
        for name in source_names
        if source_package.parts[name] != output_package.parts[name]
    )
    if changed:
        return (
            False,
            "Изменены неизменяемые части XLSX: "
            + ", ".join(changed[:5]),
        )
    return True, ""


def _package_structure(path: Path) -> dict[str, object]:
    """Canonical workbook structure with writable cell payloads removed."""

    package = OOXMLWorkbook(path)
    workbook = etree.fromstring(package.parts["xl/workbook.xml"])
    for calc_properties in workbook.findall(f"{{{MAIN_NS}}}calcPr"):
        workbook.remove(calc_properties)
    workbook_signature = etree.tostring(
        workbook,
        method="c14n",
        with_comments=False,
    )

    worksheets: dict[str, bytes] = {}
    for sheet_name in package.sheet_parts:
        root = package._sheet_root(sheet_name)
        sheet_data = root.find(f"{{{MAIN_NS}}}sheetData")
        if sheet_data is not None:
            for row in list(sheet_data):
                for cell in list(row):
                    if etree.QName(cell).localname == "c":
                        row.remove(cell)
                row.text = None
                if len(row) == 0 and set(row.attrib) <= {"r"}:
                    sheet_data.remove(row)
            sheet_data.text = None
        worksheets[sheet_name] = etree.tostring(
            root,
            method="c14n",
            with_comments=False,
        )
    return {
        "workbook": workbook_signature,
        "worksheets": worksheets,
    }
