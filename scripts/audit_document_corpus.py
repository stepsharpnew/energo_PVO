#!/usr/bin/env python3
"""Audit a directory of executive-documentation Excel workbooks without changing it."""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import sys
import zipfile
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "src"))

from executive_docs.excel import workbook_snapshot  # noqa: E402


ERROR_TOKENS = ("#REF!", "#VALUE!", "#N/A", "#DIV/0!", "#NAME?")
OBJECT_FIELDS = {
    "sap": "B2",
    "object_name": "B3",
    "district": "B4",
    "object_address": "B5",
    "actual_start": "B6",
    "actual_end": "B7",
    "installation": "B9",
    "project_code": "B43",
    "object_name_copy": "B44",
}
PILOT_FAMILIES = {"KL"}


def relative(path: Path) -> str:
    try:
        return path.relative_to(ROOT).as_posix()
    except ValueError:
        return path.as_posix()


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for chunk in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def family_for(path: Path) -> str:
    name = path.name.upper()
    if "ГНБ" in name:
        return "GNB"
    if "АВК" in name:
        return "AVK"
    if "ОЖР" in name:
        return "OJR"
    if "ЭМР" in name:
        return "EMR"
    if "БРТП" in name or "КТП" in name:
        return "BRTP"
    if "КЛ" in name:
        return "KL"
    if "ВРЩ" in name:
        return "VRS"
    return "UNKNOWN"


def xml_link_targets(path: Path) -> list[str]:
    targets: set[str] = set()
    with zipfile.ZipFile(path) as archive:
        for member in archive.namelist():
            if not member.endswith(".rels"):
                continue
            try:
                root = ET.fromstring(archive.read(member))
            except ET.ParseError:
                continue
            for relationship in root:
                target = relationship.attrib.get("Target", "")
                mode = relationship.attrib.get("TargetMode", "")
                if mode == "External" or target.startswith(("http:", "https:", "file:")):
                    targets.add(target)
    return sorted(targets)


def normalized_value(value: Any) -> str | None:
    if value is None:
        return None
    if hasattr(value, "isoformat"):
        return value.isoformat()
    text = re.sub(r"\s+", " ", str(value)).strip()
    return text or None


def object_card(path: Path) -> dict[str, str | None]:
    formulas = load_workbook(path, read_only=True, data_only=False, keep_links=False)
    values = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    try:
        if "Данные объект" not in formulas.sheetnames:
            return {}
        formula_sheet = formulas["Данные объект"]
        value_sheet = values["Данные объект"]
        result: dict[str, str | None] = {}
        for key, cell in OBJECT_FIELDS.items():
            raw = formula_sheet[cell].value
            cached = value_sheet[cell].value
            result[key] = normalized_value(cached if cached is not None else raw)
        return result
    finally:
        formulas.close()
        values.close()


def aosr_register(path: Path, sheet_names: list[str]) -> list[dict[str, str | int | float | None]]:
    workbook = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    try:
        items = []
        for sheet_name in sheet_names:
            if not sheet_name.startswith("АОСР"):
                continue
            sheet = workbook[sheet_name]
            items.append(
                {
                    "sheet": sheet_name,
                    "number": sheet["C32"].value,
                    "suffix": normalized_value(sheet["E32"].value),
                    "work_name": normalized_value(sheet["A62"].value),
                }
            )
        return items
    finally:
        workbook.close()


def print_configuration(path: Path, sheet_names: list[str]) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=False, data_only=False, keep_links=False)
    try:
        items = []
        for sheet_name in sheet_names:
            if not (sheet_name.startswith("АОСР") or sheet_name.startswith("Форма")):
                continue
            sheet = workbook[sheet_name]
            print_area = str(sheet.print_area) if sheet.print_area else None
            paper_size = str(sheet.page_setup.paperSize) if sheet.page_setup.paperSize else None
            orientation = sheet.page_setup.orientation
            items.append(
                {
                    "sheet": sheet_name,
                    "state": sheet.sheet_state,
                    "print_area": print_area,
                    "paper_size": paper_size,
                    "orientation": orientation,
                    "fit_to_width": sheet.page_setup.fitToWidth,
                    "fit_to_height": sheet.page_setup.fitToHeight,
                    "a4_configured": paper_size == "9",
                    "has_print_area": bool(print_area),
                }
            )
        return items
    finally:
        workbook.close()


def duplicate_aosr_bases(sheet_names: list[str]) -> dict[str, list[str]]:
    groups: dict[str, list[str]] = defaultdict(list)
    for name in sheet_names:
        if not name.startswith("АОСР"):
            continue
        base = re.sub(r"\s*\(\d+\)$", "", name)
        groups[base].append(name)
    return {base: names for base, names in groups.items() if len(names) > 1}


def external_link_count(value: Any) -> int:
    if isinstance(value, int):
        return value
    if isinstance(value, (list, tuple, set, dict)):
        return len(value)
    return int(bool(value))


def audit_workbook(path: Path) -> dict[str, Any]:
    snapshot = workbook_snapshot(path)
    sheets = [item["name"] for item in snapshot["sheets"]]
    cell_errors = [error for item in snapshot["sheets"] for error in item["errors"]]
    all_errors = cell_errors + snapshot["defined_name_errors"]
    external_targets = xml_link_targets(path)
    print_items = print_configuration(path, sheets)
    register = aosr_register(path, sheets)
    print_blockers = [
        f"{item['sheet']}: {'нет Print_Area' if not item['has_print_area'] else 'не задан A4'}"
        for item in print_items
        if not item["has_print_area"] or not item["a4_configured"]
    ]
    duplicates = duplicate_aosr_bases(sheets)
    blockers = []
    if all_errors:
        blockers.append("FORMULA_ERRORS")
    if snapshot["external_links"] or external_targets:
        blockers.append("EXTERNAL_LINKS")
    if print_blockers:
        blockers.append("PRINT_CONFIGURATION")
    if duplicates:
        blockers.append("DUPLICATE_WORK_SHEETS_REQUIRE_RULE")
    return {
        "file": relative(path),
        "sha256": sha256(path),
        "family": family_for(path),
        "mvp_scope": family_for(path) in PILOT_FAMILIES,
        "sheet_count": len(sheets),
        "sheet_names": sheets,
        "aosr_sheets": [name for name in sheets if name.startswith("АОСР")],
        "aosr_register": register,
        "duplicate_aosr_bases": duplicates,
        "formula_error_count": len(all_errors),
        "formula_errors": all_errors,
        "external_link_parts": snapshot["external_links"],
        "external_targets": external_targets,
        "object_card": object_card(path),
        "print_configuration": print_items,
        "print_blockers": print_blockers,
        "blockers": blockers,
    }


def consistency_findings(workbooks: list[dict[str, Any]]) -> list[dict[str, Any]]:
    findings = []
    for field in OBJECT_FIELDS:
        values: dict[str, list[str]] = defaultdict(list)
        for workbook in workbooks:
            value = workbook["object_card"].get(field)
            if value and not value.startswith("="):
                values[value].append(workbook["file"])
        if len(values) > 1:
            findings.append(
                {
                    "field": field,
                    "status": "CONFLICT_OR_MULTIPLE_SEGMENTS",
                    "values": [{"value": value, "files": files} for value, files in sorted(values.items())],
                    "action": "Эксперт должен подтвердить: это сегменты одного объекта или смешение проектов.",
                }
            )
    by_project: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for workbook in workbooks:
        project_code = workbook["object_card"].get("project_code")
        if project_code:
            by_project[project_code].append(workbook)
    for project_code, project_workbooks in by_project.items():
        identities: dict[str, list[str]] = defaultdict(list)
        for workbook in project_workbooks:
            for item in workbook["aosr_register"]:
                if item["number"] is None:
                    continue
                identity = f"{item['number']}{item['suffix'] or ''}"
                identities[identity].append(workbook["file"])
        duplicates = {identity: files for identity, files in identities.items() if len(set(files)) > 1}
        if duplicates:
            findings.append(
                {
                    "field": "aosr_number",
                    "status": "DUPLICATE_ACROSS_WORKBOOKS",
                    "project_code": project_code,
                    "values": duplicates,
                    "action": "Подтвердить область нумерации: единая по проекту или отдельная по участку/книге.",
                }
            )

        installations = {
            workbook["file"]: workbook["object_card"].get("installation")
            for workbook in project_workbooks
            if workbook["object_card"].get("installation")
        }
        for workbook in project_workbooks:
            current = installations.get(workbook["file"])
            if not current:
                continue
            for item in workbook["aosr_register"]:
                work_name = item.get("work_name") or ""
                foreign = [
                    {"file": file, "installation": installation}
                    for file, installation in installations.items()
                    if file != workbook["file"] and installation in work_name and current not in work_name
                ]
                if foreign:
                    findings.append(
                        {
                            "field": "work_name.installation",
                            "status": "POSSIBLE_CROSS_SEGMENT_COPY_ERROR",
                            "project_code": project_code,
                            "file": workbook["file"],
                            "sheet": item["sheet"],
                            "work_name": work_name,
                            "foreign_installations": foreign,
                            "action": "Специалист должен подтвердить трассу/участок для этого акта.",
                        }
                    )
    return findings


def render_markdown(report: dict[str, Any]) -> str:
    summary = report["summary"]
    lines = [
        "# Аудит корпуса исполнительной документации",
        "",
        f"Источник: `{report['source_directory']}`. Проверено книг: **{summary['workbook_count']}**.",
        "",
        "## Итог",
        "",
        f"- Книг в текущем MVP: {summary['mvp_workbook_count']}.",
        f"- Книг вне текущего MVP: {summary['out_of_scope_workbook_count']}.",
        f"- Формульных/ссылочных ошибок: {summary['formula_error_count']}.",
        f"- Книг с внешними связями: {summary['workbooks_with_external_links']}.",
        f"- Листов АОСР: {summary['aosr_sheet_count']}.",
        f"- Нарушений/неполноты печатной конфигурации: {summary['print_blocker_count']}.",
        "",
        "## Книги",
        "",
        "| Файл | Семейство | Листы | АОСР | Ошибки | Внешние связи | Статус |",
        "|---|---:|---:|---:|---:|---:|---|",
    ]
    for item in report["workbooks"]:
        status = "MVP" if item["mvp_scope"] else "вне MVP"
        if item["blockers"]:
            status += "; есть блокеры"
        lines.append(
            f"| `{Path(item['file']).name}` | {item['family']} | {item['sheet_count']} | "
            f"{len(item['aosr_sheets'])} | {item['formula_error_count']} | "
            f"{external_link_count(item['external_link_parts']) or len(item['external_targets'])} | {status} |"
        )
    lines.extend(
        [
            "",
            "## Вывод для MVP",
            "",
            "`project2` следует использовать как регрессионный корпус готовой ИД, а не как входной проект и не как автоматически утверждённый golden set. "
            "Две книги КЛ пригодны для расширения правил после экспертного подтверждения состава работ, значений и порядка листов. "
            "Остальные семейства фиксируют будущий контур продукта, но не должны автоматически включаться в пилот.",
            "",
            "## Требующие решения наблюдения",
            "",
        ]
    )
    duplicate_files = [item for item in report["workbooks"] if item["duplicate_aosr_bases"]]
    if duplicate_files:
        for item in duplicate_files:
            lines.append(f"- `{Path(item['file']).name}`: повторные типы листов {item['duplicate_aosr_bases']}.")
    else:
        lines.append("- Повторные типы листов АОСР не найдены.")
    for finding in report["cross_workbook_findings"]:
        if finding["status"] == "DUPLICATE_ACROSS_WORKBOOKS":
            lines.append(
                f"- В проекте `{finding['project_code']}` номера АОСР повторяются между книгами; "
                "нужно подтвердить область нумерации."
            )
        elif finding["status"] == "POSSIBLE_CROSS_SEGMENT_COPY_ERROR":
            lines.append(
                f"- `{Path(finding['file']).name}`, `{finding['sheet']}`: название работы ссылается на трассу другой книги; "
                "возможна ошибка копирования."
            )
        else:
            lines.append(f"- Поле `{finding['field']}` имеет несколько значений; требуется определить границы объекта/сегментов.")
    lines.extend(
        [
            "- Формульная ошибка в готовом примере не становится правилом и не должна переноситься в шаблон.",
            "- Факты организации и подписантов из готовых книг не становятся профилем без отдельного подтверждения специалиста.",
            "",
        ]
    )
    return "\n".join(lines)


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("source", type=Path, help="Directory with .xlsx workbooks")
    parser.add_argument("--output", type=Path, required=True, help="Output directory for JSON and Markdown reports")
    args = parser.parse_args()

    source = args.source.resolve()
    output = args.output.resolve()
    files = sorted(source.glob("*.xlsx"), key=lambda item: item.name.casefold())
    if not files:
        raise SystemExit(f"No .xlsx workbooks found in {source}")

    workbooks = [audit_workbook(path) for path in files]
    family_counts = Counter(item["family"] for item in workbooks)
    report = {
        "source_directory": relative(source),
        "method": "Read-only OOXML/openpyxl inspection; no source workbook was modified.",
        "summary": {
            "workbook_count": len(workbooks),
            "families": dict(sorted(family_counts.items())),
            "mvp_workbook_count": sum(item["mvp_scope"] for item in workbooks),
            "out_of_scope_workbook_count": sum(not item["mvp_scope"] for item in workbooks),
            "sheet_count": sum(item["sheet_count"] for item in workbooks),
            "aosr_sheet_count": sum(len(item["aosr_sheets"]) for item in workbooks),
            "formula_error_count": sum(item["formula_error_count"] for item in workbooks),
            "workbooks_with_external_links": sum(
                bool(external_link_count(item["external_link_parts"]) or item["external_targets"]) for item in workbooks
            ),
            "print_blocker_count": sum(len(item["print_blockers"]) for item in workbooks),
        },
        "cross_workbook_findings": consistency_findings(workbooks),
        "workbooks": workbooks,
    }
    output.mkdir(parents=True, exist_ok=True)
    (output / "project-corpus-audit.json").write_text(
        json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    (output / "project-corpus-audit.md").write_text(render_markdown(report), encoding="utf-8")
    print(json.dumps(report["summary"], ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
